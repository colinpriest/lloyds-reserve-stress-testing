"""LLM PDF/HTML extraction: Gemini vs GPT-5-mini with inline adjudication.

Extracts structured reserve data from each Lloyd's syndicate report using two
independent LLMs (Gemini and GPT), compares outputs field-by-field, and when
they disagree on material fields, immediately pauses to adjudicate:

  1. Sends the PDF to a third model (Claude) for targeted verification
  2. Presents the finding to the human operator
  3. Human approves, overrides (pick a model or enter custom value), or stops
  4. All decisions are recorded in the audit trail

Usage:
    python test_gemini.py              # run with inline adjudication on failures
    python test_gemini.py --clean      # wipe outputs, re-run from scratch
    python test_gemini.py --batch      # old behaviour: no interactive adjudication
    python test_gemini.py --table-backend azure     # use Azure Document Intelligence (default)
    python test_gemini.py --table-backend nutrient  # use Nutrient.io
    python test_gemini.py --table-backend adobe     # use Adobe PDF Extract
    python test_gemini.py --azure-free              # free F0 tier: one page per request

LLM outputs are cached in pdf_extraction/llm_cache/ using a SHA-256 hash of
(model_name, prompt_version, prompt_text, syndicate_num, report_year[, page_num]).
Changing the prompt wording or bumping PROMPT_VERSION auto-invalidates the cache.
Delete the llm_cache/ directory to force re-extraction from all LLMs.
"""

import os
import re
import sys
import json
import time
import traceback
import base64
import hashlib
import logging
import signal
import tempfile
import zipfile
import requests
from datetime import datetime, timezone
from pathlib import Path
from dotenv import load_dotenv
from google import genai
from google.genai.types import GenerateContentConfig
from openai import OpenAI

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
    from adobe.pdfservices.operation.pdf_services import PDFServices
    from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
    from adobe.pdfservices.operation.pdfjobs.jobs.extract_pdf_job import ExtractPDFJob
    from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_pdf_params import ExtractPDFParams
    from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_element_type import ExtractElementType
    from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_renditions_element_type import ExtractRenditionsElementType
    from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.table_structure_type import TableStructureType
    from adobe.pdfservices.operation.pdfjobs.result.extract_pdf_result import ExtractPDFResult
    from adobe.pdfservices.operation.io.cloud_asset import CloudAsset
    from adobe.pdfservices.operation.io.stream_asset import StreamAsset
    HAS_ADOBE = True
except ImportError:
    HAS_ADOBE = False

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    from pdf2image import convert_from_path
    from PIL import Image
    import io
    HAS_PDF2IMAGE = True
except ImportError:
    HAS_PDF2IMAGE = False

try:
    import pytesseract
    # Try common Windows install locations if not on PATH
    _tesseract_paths = [
        r"C:\Users\colin\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    ]
    try:
        pytesseract.get_tesseract_version()
    except pytesseract.TesseractNotFoundError:
        for tp in _tesseract_paths:
            if os.path.exists(tp):
                pytesseract.pytesseract.tesseract_cmd = tp
                break
    pytesseract.get_tesseract_version()
    HAS_TESSERACT = True
except Exception:
    HAS_TESSERACT = False

from adjudicate import (
    ask_human,
    build_verification_prompt,
    call_adjudicator,
    sanitize_for_json,
    determine_correct_model,
    present_adjudication,
    present_report_decision,
    load_disagreement_log,
    save_disagreement_log,
    load_rejection_log,
    save_rejection_log,
    ADJUDICATOR_MODEL,
)
from table_extraction import extract_tables, TableBackend, _extract_pages_to_pdf

load_dotenv()

logger = logging.getLogger(__name__)

# Table extraction backend (can be overridden with --table-backend)
# Priority: azure (default) > nutrient > adobe
TABLE_BACKEND = TableBackend.AZURE
AZURE_PAID = True  # default: paid S0 tier (all pages in one request); --azure-free to disable


# ---------------------------------------------------------------------------
# Unicode-to-ASCII sanitiser — applied to all JSON output
# ---------------------------------------------------------------------------

# Common Unicode -> ASCII replacements for financial/PDF text
_UNICODE_MAP = {
    "\u2014": "-",       # em dash
    "\u2013": "-",       # en dash
    "\u2012": "-",       # figure dash
    "\u2015": "-",       # horizontal bar
    "\u2018": "'",       # left single quote
    "\u2019": "'",       # right single quote / apostrophe
    "\u201c": '"',       # left double quote
    "\u201d": '"',       # right double quote
    "\u2026": "...",     # ellipsis
    "\u00a3": "GBP ",    # pound sign
    "\u20ac": "EUR ",    # euro sign
    "\u00a0": " ",       # non-breaking space
    "\u00ad": "",         # soft hyphen
    "\u00b7": ".",       # middle dot
    "\u2022": "-",       # bullet
    "\u2023": "-",       # triangular bullet
    "\u00d7": "x",       # multiplication sign
    "\u00f7": "/",       # division sign
    "\u2264": "<=",      # less than or equal
    "\u2265": ">=",      # greater than or equal
    "\u00b1": "+/-",     # plus-minus
    "\u0141": "",         # L-stroke (PDF garbage)
    "\u0142": "",         # l-stroke (PDF garbage)
    "\ufb01": "fi",      # fi ligature
    "\ufb02": "fl",      # fl ligature
    "\u2019s": "'s",     # possessive
}

# Double-encoded UTF-8 patterns (e.g. \u00c2\u00a3 = double-encoded £)
_DOUBLE_ENCODED = {
    "\u00c2\u00a3": "GBP ",   # double-encoded £
    "\u00c3\u00a9": "e",       # double-encoded e-acute
    "\u00c3\u00a8": "e",       # double-encoded e-grave
    "\u00c3\u00bc": "u",       # double-encoded u-umlaut
    "\u00c3\u00b6": "o",       # double-encoded o-umlaut
}


def _gemini_call_with_retry(fn, *, max_retries=5, base_delay=30):
    """Call a Gemini API function with exponential backoff on transient errors.

    Retries on: 503 UNAVAILABLE, 429 RESOURCE_EXHAUSTED, and 400 "Upload has
    already been terminated" (transient Gemini Files API error).

    fn: zero-argument callable that makes the API request and returns a response.
    Returns the response on success, raises on non-retryable errors.
    """
    for attempt in range(1, max_retries + 1):
        try:
            return fn()
        except Exception as e:
            err_str = str(e)
            retryable = (
                any(code in err_str for code in ("503", "429", "UNAVAILABLE", "RESOURCE_EXHAUSTED"))
                or "Upload has already been terminated" in err_str
            )
            if not retryable or attempt == max_retries:
                raise
            delay = base_delay * (2 ** (attempt - 1))
            print(f"  [retry] Attempt {attempt}/{max_retries} failed ({err_str[:80]}). "
                  f"Retrying in {delay}s...")
            time.sleep(delay)


def _sanitize_ascii(s):
    """Replace Unicode characters with nearest ASCII equivalents."""
    if not isinstance(s, str):
        return s

    # Fix double-encoded sequences first
    for pattern, replacement in _DOUBLE_ENCODED.items():
        s = s.replace(pattern, replacement)

    # Apply known replacements
    for char, replacement in _UNICODE_MAP.items():
        s = s.replace(char, replacement)

    # Strip any remaining non-ASCII characters
    s = s.encode("ascii", errors="ignore").decode("ascii")

    # Clean up double spaces from replacements
    while "  " in s:
        s = s.replace("  ", " ")

    return s


def sanitize_json_ascii(obj):
    """Recursively replace Unicode characters with ASCII in a JSON-serializable object."""
    if isinstance(obj, str):
        return _sanitize_ascii(obj)
    elif isinstance(obj, dict):
        return {_sanitize_ascii(k) if isinstance(k, str) else k:
                sanitize_json_ascii(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_json_ascii(item) for item in obj]
    else:
        return obj


# Pricing per 1M tokens (USD)
PRICING = {
    "gemini-2.5-flash": {"input": 0.30, "output": 2.50},
    "gemini-3-flash":   {"input": 0.50, "output": 3.00},
    "gpt-5-mini":       {"input": 0.25, "output": 2.00},
}

# Standard Lloyd's LOBs (from config.py)
LLOYDS_LOBS = [
    "Property", "Casualty", "Marine", "Energy", "Motor",
    "Aviation", "Reinsurance - Property", "Reinsurance - Casualty",
    "Reinsurance - Specialty", "Professional Lines", "Accident & Health",
    "Cyber", "Aggregate", "Treaty",
]

# Standard causal categories (from config.py CauseCategory enum)
CAUSAL_CATEGORIES = [
    "Natural catastrophe events",
    "Man-made catastrophe / large losses",
    "Social inflation / litigation trends",
    "Economic inflation / claims cost inflation",
    "Regulatory changes",
    "Court rulings / legal developments",
    "Ogden discount rate",
    "COVID-19 / pandemic effects",
    "Geopolitical events",
    "Favorable claims development",
    "Adverse claims development",
    "Reinsurance recoveries",
    "IBNR recalibration",
    "Reserve methodology change",
    "Large loss development",
    "Other",
]

REPORTS_DIR = Path("syndicate_reports/pdfs")
OUTPUT_DIR = Path("pdf_extraction")
HTML_PDF_CACHE = Path("pdf_extraction/html_converted")
SPEC_DIR = Path("pdf_extraction/spec")
AUDIT_DIR = Path("pdf_extraction/audit")

GEMINI_MODEL = "gemini-2.5-flash"
OPENAI_MODEL = "gpt-5-mini"

# Frozen spec versions -- bump these when spec files change
PROMPT_VERSION = "2.10"  # 2.10: add monoline LOB extraction, direction forced from triangle PYD
FIELD_DEFINITIONS_VERSION = "1.0"
TOLERANCE_RULES_VERSION = "1.0"

# ---------------------------------------------------------------------------
# LLM output caching
# ---------------------------------------------------------------------------
LLM_CACHE_DIR = Path("pdf_extraction/llm_cache")

# ---------------------------------------------------------------------------
# Syndicate inception year tracking
# ---------------------------------------------------------------------------
INCEPTION_YEARS_FILE = Path("pdf_extraction/syndicate_inception_years.json")


def _load_inception_years() -> dict:
    """Load syndicate inception years from JSON file, then backfill from
    already-extracted output JSONs that contain triangle data.

    This ensures that syndicates processed in previous runs (whose triangles
    contain ``underwriting_years``) are reflected in the cache even if
    Perplexity was never called or returned nothing for them.

    Entries listed in ``_manual_overrides`` in the JSON file are protected
    from any automatic updates (Perplexity, triangle backfill, etc.).
    """
    cache: dict[int, int] = {}
    manual: set[int] = set()
    if INCEPTION_YEARS_FILE.exists():
        with open(INCEPTION_YEARS_FILE) as f:
            data = json.load(f)
        cache = {int(k): int(v) for k, v in data.items()
                 if k not in ("_meta", "_manual_overrides")}
        # Load manual overrides list
        for s in data.get("_manual_overrides", []):
            manual.add(int(s))

    # Backfill from extraction JSONs that have triangle data or inception_year
    output_dir = Path("pdf_extraction")
    dirty = False
    for jf in output_dir.glob("syndicate_*.json"):
        try:
            stem_parts = jf.stem.split("_")  # syndicate_NNNN_YYYY
            syn_num = int(stem_parts[1])
        except (IndexError, ValueError):
            continue
        # Never overwrite manually-corrected entries
        if syn_num in manual:
            continue
        # Only backfill if syndicate not already in cache
        if syn_num in cache:
            continue
        try:
            with open(jf) as f:
                doc = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue
        # Check 1: explicit inception_year field (from first-year stubs / Perplexity)
        if "inception_year" in doc:
            try:
                cache[syn_num] = int(doc["inception_year"])
                dirty = True
                continue
            except (ValueError, TypeError):
                pass
        # Check 2: _rag_triangle.underwriting_years in model outputs
        for model_key in doc.get("models", {}).values():
            tri = model_key.get("_rag_triangle") if isinstance(model_key, dict) else None
            if tri and tri.get("underwriting_years"):
                uw_years = [int(y) for y in tri["underwriting_years"]]
                inception = min(uw_years)
                if syn_num not in cache or inception < cache[syn_num]:
                    cache[syn_num] = inception
                    dirty = True
                break

    if dirty:
        _save_inception_years(cache)
        print(f"  Backfilled inception years from existing extractions -> {len(cache)} syndicates total")

    return cache


def _save_inception_years(inception: dict) -> None:
    """Save syndicate inception years to JSON file, preserving _meta and _manual_overrides."""
    existing = {}
    if INCEPTION_YEARS_FILE.exists():
        with open(INCEPTION_YEARS_FILE) as f:
            existing = json.load(f)
    meta = existing.get("_meta", {
        "description": "First underwriting year for each Lloyd's syndicate. "
                       "Used to skip reports from the first two years of operation "
                       "where prior year development cannot be meaningfully computed.",
        "source": "Auto-populated from claims development triangles in extraction JSONs. "
                  "Unknown syndicates are looked up via Perplexity API.",
        "last_updated": datetime.now().strftime("%Y-%m-%d"),
    })
    meta["last_updated"] = datetime.now().strftime("%Y-%m-%d")
    manual_overrides = existing.get("_manual_overrides", [])
    output = {"_meta": meta, "_manual_overrides": manual_overrides}
    for s in sorted(inception.keys()):
        output[str(s)] = inception[s]
    INCEPTION_YEARS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(INCEPTION_YEARS_FILE, "w") as f:
        json.dump(output, f, indent=2)



def _load_manual_overrides() -> set[int]:
    """Load the set of syndicate numbers whose inception years are manually
    verified and must never be overwritten by automated processes."""
    if INCEPTION_YEARS_FILE.exists():
        with open(INCEPTION_YEARS_FILE) as f:
            data = json.load(f)
        return {int(s) for s in data.get("_manual_overrides", [])}
    return set()


def _lookup_inception_year_perplexity(syndicate_num: int) -> int | None:
    """Query Perplexity API for the first underwriting year of a syndicate.

    Returns the year as int, or None if lookup fails.
    Uses structured JSON output to avoid ambiguous free-text parsing.
    """
    api_key = os.getenv("PERPLEXITY_API_KEY")
    if not api_key:
        print(f"  WARNING: PERPLEXITY_API_KEY not set - cannot look up inception year for syndicate {syndicate_num}")
        return None

    query = (
        f"What year did Lloyd's of London syndicate {syndicate_num} first begin "
        f"underwriting insurance? I need the first year they wrote any business. "
        f"Note: the syndicate NUMBER ({syndicate_num}) is NOT necessarily the same "
        f"as the year it started — many syndicates have numbers that look like years "
        f"but started in a completely different year."
    )

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "sonar",
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are a research assistant specialising in Lloyd's of London."
                ),
            },
            {"role": "user", "content": query},
        ],
        "temperature": 0.1,
        "response_format": {
            "type": "json_schema",
            "json_schema": {
                "schema": {
                    "type": "object",
                    "properties": {
                        "syndicate_number": {
                            "type": "integer",
                            "description": "The syndicate number queried",
                        },
                        "first_underwriting_year": {
                            "type": "integer",
                            "description": "The first year this syndicate wrote any insurance business",
                        },
                        "confidence": {
                            "type": "string",
                            "enum": ["high", "medium", "low"],
                            "description": "Confidence in the answer",
                        },
                        "source": {
                            "type": "string",
                            "description": "Where this information was found",
                        },
                    },
                    "required": ["syndicate_number", "first_underwriting_year", "confidence", "source"],
                },
            },
        },
    }

    try:
        response = requests.post(
            "https://api.perplexity.ai/chat/completions",
            headers=headers,
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        result = response.json()
        answer = result["choices"][0]["message"]["content"].strip()

        # Strip markdown code fences if present
        if answer.startswith("```"):
            answer = re.sub(r"^```(?:json)?\s*", "", answer)
            answer = re.sub(r"\s*```$", "", answer)

        parsed = json.loads(answer)
        year = parsed["first_underwriting_year"]
        confidence = parsed.get("confidence", "unknown")
        source = parsed.get("source", "")

        if not isinstance(year, int) or year < 1688 or year > 2030:
            print(f"  WARNING: Perplexity returned invalid year {year!r} for syndicate {syndicate_num}. "
                  f"Full response: {parsed}")
            return None

        # Reject low-confidence answers
        if confidence == "low":
            print(f"  WARNING: Perplexity low-confidence answer for syndicate {syndicate_num}: {year} ({source})")
            return None

        print(f"  Perplexity: Syndicate {syndicate_num} first underwrote in {year} "
              f"(confidence={confidence}, source={source!r})")
        return year
    except Exception as e:
        print(f"  WARNING: Perplexity lookup failed for syndicate {syndicate_num}: {e}")
        return None


def _earliest_report_year(syndicate_num: int) -> int | None:
    """Find the earliest report year we have on disk for a syndicate."""
    pdf_dir = Path("syndicate_reports/pdfs")
    if not pdf_dir.exists():
        return None
    earliest = None
    for p in pdf_dir.glob(f"syndicate_{syndicate_num}_*.pdf"):
        try:
            year = int(p.stem.split("_")[2])
            if earliest is None or year < earliest:
                earliest = year
        except (IndexError, ValueError):
            continue
    return earliest


def get_inception_year(syndicate_num: int, inception_cache: dict) -> int | None:
    """Get the first underwriting year for a syndicate.

    Checks local cache first, then queries Perplexity if missing.
    Stores result back in cache and saves to disk.

    Sanity check: if Perplexity returns a year later than our earliest
    report for this syndicate, the answer is clearly wrong — a syndicate
    can't have reports before it started underwriting. In that case,
    fall back to earliest_report_year - 2 (conservative estimate).
    """
    if syndicate_num in inception_cache:
        return inception_cache[syndicate_num]

    # Not in cache - query Perplexity
    year = _lookup_inception_year_perplexity(syndicate_num)
    if year is not None:
        earliest_report = _earliest_report_year(syndicate_num)
        if earliest_report is not None and year > earliest_report:
            fallback = earliest_report - 2
            print(f"  WARNING: Perplexity returned inception={year} for syndicate "
                  f"{syndicate_num}, but we have a report from {earliest_report}. "
                  f"Using conservative fallback: {fallback}")
            year = fallback
        inception_cache[syndicate_num] = year
        _save_inception_years(inception_cache)
    return year


def is_early_year_syndicate(syndicate_num: int, report_year: int,
                            inception_cache: dict) -> tuple[bool, int | None]:
    """Check if a report is from the first two underwriting years of a syndicate.

    Returns (should_skip, inception_year).
    A report should be skipped when report_year < inception_year + 2,
    because fewer than 3 development years are available for PYD computation.
    """
    inception_year = get_inception_year(syndicate_num, inception_cache)
    if inception_year is None:
        return False, None  # Can't determine - don't skip
    return report_year < inception_year + 2, inception_year


def _llm_cache_key(model: str, prompt_text: str, syndicate_num: int,
                   report_year: int, page_num: int = None) -> str:
    """Build a deterministic cache key from model + prompt + context.

    Hash = SHA-256 of (model_name, PROMPT_VERSION, prompt_text, syndicate,
    year[, page]).  Returns hex digest used as the cache filename stem.
    """
    parts = [model, PROMPT_VERSION, prompt_text, str(syndicate_num), str(report_year)]
    if page_num is not None:
        parts.append(str(page_num))
    blob = "|".join(parts).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()


def _llm_cache_path(cache_key: str) -> Path:
    """Return the filesystem path for a given cache key."""
    return LLM_CACHE_DIR / f"{cache_key}.json"


def _llm_cache_load(cache_key: str):
    """Load cached LLM result. Returns (data_dict, True) or (None, False)."""
    path = _llm_cache_path(cache_key)
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f), True
        except (json.JSONDecodeError, OSError):
            pass
    return None, False


def _llm_cache_save(cache_key: str, data, *, model: str = "",
                    syndicate_num: int = 0, report_year: int = 0,
                    page_num: int = None):
    """Persist LLM result to cache."""
    LLM_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    envelope = {
        "_cache_meta": {
            "model": model,
            "prompt_version": PROMPT_VERSION,
            "syndicate": syndicate_num,
            "year": report_year,
            "cached_at": datetime.now(timezone.utc).isoformat(),
        },
        "data": data,
    }
    if page_num is not None:
        envelope["_cache_meta"]["page"] = page_num
    with open(_llm_cache_path(cache_key), "w", encoding="utf-8") as f:
        json.dump(envelope, f, indent=2, ensure_ascii=True)


def convert_html_to_pdf(html_path):
    """Convert HTML file to PDF using Playwright. Returns path to cached PDF."""
    HTML_PDF_CACHE.mkdir(parents=True, exist_ok=True)
    pdf_path = HTML_PDF_CACHE / f"{html_path.stem}.pdf"
    if pdf_path.exists():
        print(f"  Using cached PDF: {pdf_path.name}")
        return pdf_path

    print(f"  Converting {html_path.name} to PDF via Playwright...")
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto(f"file:///{html_path.resolve().as_posix()}")
        page.pdf(path=str(pdf_path), format="A4", print_background=True)
        browser.close()
    print(f"  Converted: {pdf_path.name} ({pdf_path.stat().st_size / 1024:.0f} KB)")
    return pdf_path


def discover_reports():
    """Find all syndicate report files (PDF and HTML), sorted by name."""
    files = sorted(
        f for f in REPORTS_DIR.iterdir()
        if f.suffix.lower() in (".pdf", ".html", ".htm")
        and f.stem.startswith("syndicate_")
    )
    return files


def parse_report_filename(report_path):
    """Extract syndicate number and year from filename."""
    parts = report_path.stem.split("_")
    return int(parts[1]), int(parts[2])


def build_prompt(syndicate_num, report_year):
    """Build the shared extraction prompt."""
    return f"""Extract reserve data from this Lloyd's syndicate annual report and return ONLY valid JSON (no markdown fences).

Use these EXACT standard Lloyd's lines of business names where applicable:
{json.dumps(LLOYDS_LOBS)}

Use these standard causal categories:
{json.dumps(CAUSAL_CATEGORIES)}

Return this JSON structure:

{{
  "syndicate": {syndicate_num},
  "year": {report_year},
  "opening_reserves_gbp_m": <opening GROSS CLAIMS OUTSTANDING at start of year, in millions. PRIMARY SOURCE: the Balance Sheet / Statement of Financial Position — find "Claims outstanding" (or "Gross claims outstanding") in the LIABILITIES section under "Technical provisions". Use the PRIOR YEAR comparative column (e.g. in a 2019 report, use the 2018 column). This is ONLY claims reserves — do NOT include "Provision for unearned premiums". Do NOT use the reinsurers' share from the ASSETS side. SECONDARY SOURCE: Technical Reserves note "At 1 January" — but prefer Balance Sheet if both are available. null if not found>,
  "opening_reserves_page": <page number where found>,
  "opening_reserves_confidence": <0.0 to 1.0>,
  "prior_year_development_gbp_m": <GROSS amount in millions as a SIGNED number: NEGATIVE for releases, POSITIVE for strengthenings/deteriorations. MUST be the GROSS figure (insurance liabilities), NOT net of reinsurance. If the note shows Insurance liabilities / Reinsurer's share / Net columns, use the INSURANCE LIABILITIES column. Use the figure from the "Movement in prior year's provision for claims outstanding" note. CRITICAL: Do NOT use the "Claims incurred in prior underwriting years" row from the Profit and Loss Account Technical Account — this is GROSS CLAIMS INCURRED (premiums earned minus claims paid minus reserve changes), NOT the reserve movement. It is a completely different accounting concept. Do NOT use narrative text that says "net releases of £X" or "net improvement of £X" — the word "net" means after reinsurance. Do NOT use the "Movement in provision" line from the Technical Reserves reconciliation table, which includes current year movements. null if not found>,
  "prior_year_development_pct": <as percentage of opening gross claims outstanding, NEGATIVE for releases, POSITIVE for strengthenings, null if not calculable>,
  "direction": "<release|strengthening|flat|mixed — also consider year-of-account closure language: 'profit on closed year' or 'improvement on forecast' typically indicates release; 'deterioration' or 'loss on closed year' indicates strengthening; both directions across LOBs = mixed>",
  "prior_year_movement_page": <page number>,
  "prior_year_movement_confidence": <0.0 to 1.0>,
  "exact_reserve_text": "<copy VERBATIM the sentence(s) from the document that describe the prior year reserve movement. Do NOT copy rows from the Profit & Loss Technical Account (e.g. 'Claims incurred in relation to prior underwriting years 789.7 (617.6) 172.1') — these are claims incurred figures, not reserve movements>",
  "primary_causes": ["<map each cause to the closest standard causal category above>"],
  "specific_events": ["<named events e.g. 'Hurricane Ian 2022', empty list if none>"],
  "specific_years_affected": [<list of prior accident years mentioned as affected, empty list if none>],
  "prior_year_events": [
    {{
      "event_name": "<name of the event, e.g. 'Hurricane Katrina', 'Deepwater Horizon', 'Ogden rate change'>",
      "event_year": <year the event occurred>,
      "impact_description": "<how this event affected reserves, e.g. 'adverse development on 2011 Thai floods', 'release of reserves following benign claims experience'>",
      "lobs_affected": ["<standard Lloyd's LOB names affected>"],
      "amount_gbp_m": <signed impact in millions if quantified: negative=release, positive=strengthening, null if not quantified>,
      "confidence": <0.0 to 1.0>
    }}
  ],
  "named_events": [
    {{
      "event_name": "<specific name of the event, e.g. 'Hurricane Odile', 'Sewol ferry disaster', 'Japan snowfall'>",
      "event_year": <year the event occurred>,
      "event_type": "<natural_catastrophe|man_made|liability|political|other>",
      "net_loss_gbp_m": <net loss to the syndicate in millions if quantified, null if not>,
      "loss_description": "<how the report describes the impact, e.g. 'contained within catastrophe budget', 'net loss below £10m'>",
      "lobs_affected": ["<standard Lloyd's LOB names affected>"],
      "page": <page number where mentioned>,
      "confidence": <0.0 to 1.0>
    }}
  ],
  "standardized_narrative": "<one paragraph summary of the reserve movement, its causes, and affected lines>",
  "raw_causal_phrases": ["<copy exact causal phrases from the document verbatim>"],
  "lob_movements": [
    {{
      "line_of_business": "<standard Lloyd's LOB name from list above>",
      "direction": "<release|strengthening>",
      "amount_gbp_m": <signed amount in millions: negative=release, positive=strengthening, null if not quantified>,
      "percentage": <signed percentage if available, null otherwise>,
      "confidence": <0.0 to 1.0>
    }}
  ],
  "gross_premiums_written_gbp_m": <total in millions as a number>,
  "gross_premium_mix": [
    {{
      "line_of_business": "<EXACT regulatory/statutory class name as printed in the Segmental Analysis note — copy verbatim, e.g. 'Marine, aviation and transport', 'Fire and other damage to property', 'Third party liability', 'Miscellaneous', 'Reinsurance'. Do NOT rename or map these to standard Lloyd's LOBs>",
      "amount_gbp_m": <amount in millions as a number>,
      "percentage_of_total": <percentage as a number e.g. 47.6>
    }}
  ],
  "gross_premium_page": <page number>,
  "gross_premium_confidence": <0.0 to 1.0>,
  "currency": "<GBP, USD, or EUR - whichever the report's financial statements are denominated in. Report the amounts in the NATIVE currency, do NOT convert to GBP>",
  "data_quality_notes": "<any caveats about data availability or extraction uncertainty>",
  "_claims_triangle": {{
    "type": "<gross|net|loss_ratio|none — 'gross' if the report has a GROSS claims development triangle, 'net' if only NET, 'loss_ratio' if only loss ratio percentages, 'none' if no triangle>",
    "currency": "<GBP, USD, or EUR>",
    "units": "<millions|thousands|percentage — the unit used in the triangle>",
    "page": <page number where found, null if none>,
    "underwriting_years": [<list of ALL individual UW year column headers as integers, e.g. [2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018]. CRITICAL: EXCLUDE any 'X and prior' aggregate column (e.g. '2010 and prior', '2010 & prior'). These aggregate columns do NOT have proper development rows — they must be omitted entirely. Include ALL individual year columns including the most recent>],
    "development_rows": [
      [<Row 0: 'At end of underwriting year' — one number per UW year column, null if cell is blank>],
      [<Row 1: 'One year later' — one number per UW year column, null if blank>],
      [<Row 2: 'Two years later' — one number per UW year column, null if blank>],
      [<Row 3: 'Three years later' — etc.>],
      [<...continue for ALL development period rows...>]
    ]
    IMPORTANT: Include ONLY the development period rows (At end of UW year, One year later, Two years later, etc.). Do NOT include the 'Current estimate of cumulative claims incurred' summary row — that row just repeats the last value from each column and is NOT a development period. Do NOT include the 'Cumulative payments' or 'Gross outstanding claims provision' rows.
  }}
}}

Rules:
- All monetary amounts MUST be plain numbers in millions (no currency symbols, no 'm' suffix)
- Releases are NEGATIVE (e.g. a £77.4m release = -77.4)
- Strengthenings/deteriorations are POSITIVE (e.g. a £10m strengthening = 10.0)
- For confidence: 1.0 = explicitly stated, 0.8 = calculated from stated values, 0.5 = inferred, 0.0 = not found
- Map line of business names to the standard Lloyd's LOBs where possible. If a syndicate has a "Treaty" division (reinsurance accepted), map it to "Treaty" — do NOT split it into Reinsurance - Property/Casualty/Specialty unless the report explicitly breaks it down by those sub-categories.
- Map causes to the standard causal categories where possible
- For exact_reserve_text: copy verbatim, do not paraphrase
- IMPORTANT — opening_reserves_gbp_m: Use ONLY gross claims outstanding (claims reserves). Do NOT include unearned premium provisions. PRIMARY source is the Balance Sheet / Statement of Financial Position: find "Claims outstanding" under "Technical provisions" in the LIABILITIES section, and use the PRIOR YEAR comparative column. Do NOT use the "Reinsurers' share of claims outstanding" from the ASSETS side — that is the reinsurance recoverable, not the gross claims reserve. If the Balance Sheet is not available, fall back to the Technical Reserves note "At 1 January" figure.
- IMPORTANT — prior_year_development_gbp_m — WRONG SOURCES TO REJECT FIRST:
  ✗ "Claims incurred in prior underwriting years" or "Claims incurred in relation to prior underwriting years" from the Profit & Loss Account / Technical Account. Despite containing the words "prior underwriting years", this is CLAIMS INCURRED (a P&L accounting line), NOT the prior year reserve movement. Example to REJECT: "Claims incurred in relation to prior underwriting years 789.7 (617.6) 172.1" — these are gross/reinsurance/net CLAIMS INCURRED, not reserve movements. IGNORE THIS COMPLETELY.
  ✗ "Movement in provision" from Technical Reserves reconciliation — includes BOTH current + prior year combined.
  ✗ Closing reserve balances or net technical provisions — balance sheet figures, not movements.
  ✗ Changes in booked ultimates for specific named events — one event's estimate change, not total portfolio PYD.
  ✗ Year-of-account PROFIT/LOSS results — e.g. "a loss to capital providers of £17.7m" — this is the overall underwriting profit/loss, NOT a reserve movement.
- IMPORTANT — prior_year_development_gbp_m — CORRECT SOURCES (use FIRST found, must be GROSS):
  Use the GROSS figure (insurance liabilities), NOT the net figure (after reinsurer's share). When a "Movement in prior year claims" note shows columns for "Insurance liabilities", "Reinsurer's share", and "Net liabilities", use the "Insurance liabilities" column. Example: Insurance liabilities (69.6), Reinsurer's share 29.8, Net (39.8) → use -69.6, NOT -39.8. WARNING: Narrative text often quotes the NET figure (e.g. "net releases of £39.8m"). Always cross-check against the movement note — the GROSS (Insurance liabilities) column is authoritative.
  Source 1: "Movement in prior year's provision for claims outstanding" note (ONLY if it shows GROSS figure).
  Source 2: Narrative text explicitly stating GROSS amount (e.g. "released £X in respect of prior periods").
  Source 3: Year-of-account result breakdown summing non-current-year components.
  Source 4: GROSS claims development triangle — compare bottom row to previous diagonal for UW years older than 2 most recent.
  Source 5: Loss ratio development table × premiums (fallback if no absolute triangle).
  Source 6 (LAST RESORT): If NO gross figure is available from ANY of the above sources (no gross movement note, no gross triangle, no loss ratio table), but the narrative text or movement note explicitly states a NET (after reinsurance) prior year development figure, use the NET figure. Flag this clearly in data_quality_notes as "NET of reinsurance — no gross figure available". This is better than returning null when the report clearly quantifies the prior year movement, even if only net.
  IMPORTANT FALLBACK: If the movement note only shows NET figures AND a gross triangle exists, fall back to the GROSS claims development triangle (source 4). Only use Source 6 (net figure) if no gross source is available at all. Return null only if no figure (gross or net) is available.
- IMPORTANT — sign convention for "surplus/(deficit)" language: When a report says "A surplus/(deficit) run-off deviation of (X) million", the PARENTHESES around the number indicate a DEFICIT. A deficit means prior reserves were INSUFFICIENT, which is ADVERSE development = STRENGTHENING (POSITIVE sign). Example: "surplus/(deficit) of (3.0) million" means a 3.0m deficit = prior_year_development_gbp_m: +3.0, direction: "strengthening". Conversely, an unparenthesized number means a surplus = release = NEGATIVE sign.
- IMPORTANT — gross_premium_mix: Use the REGULATORY segmental analysis from the Notes to the Accounts. Copy the line of business names EXACTLY as printed (e.g. "Marine, aviation and transport", "Fire and other damage to property", "Third party liability", "Miscellaneous", "Reinsurance"). Do NOT rename them to standard Lloyd's LOB names. Do NOT split combined categories into separate entries. Do NOT use the underwriter's internal divisional breakdown. Even if the report has only ONE line of business (e.g. a monoline reinsurer writing 100% "Reinsurance"), still include it as a single entry in gross_premium_mix — do NOT return an empty array just because there is only one class. Also look for "Gross written premium income by class of business" tables in the Managing Agent's Report — these often provide a more granular divisional breakdown than the regulatory segmental analysis note.
- IMPORTANT — gross_premium_mix with "Direct insurance" and "Reinsurance acceptances" sub-tables: Some segmental analysis notes split gross premiums into "Direct insurance" and "Reinsurance acceptances" sub-tables, each with their own LOB categories (e.g. both may have "Fire and other damage to property"). In this case, list the individual Direct insurance categories with their amounts, then add a SINGLE consolidated "Reinsurance acceptances" line with the total of all reinsurance accepted premiums. Do NOT list the individual reinsurance sub-categories separately (they would create duplicate LOB names). The total should still equal gross_premiums_written_gbp_m.
- IMPORTANT — gross_premium_mix: prefer DIVISIONAL TOTALS over regulatory sub-categories. When the report contains BOTH a regulatory segmental analysis (with fine-grained statutory classes like "Marine, aviation and transport", "Fire and other damage to property") AND a divisional/business class summary (e.g. "Marine", "Property", "Specialty", "Political Lines", "Treaty"), use the DIVISIONAL summary. The divisional breakdown aggregates across direct and reinsurance business to give the TOTAL premium per business class, which is what we need. The regulatory segmental analysis often shows only the direct insurance component for each statutory class, understating the true LOB total. Each entry in gross_premium_mix should represent the TOTAL premium for that business class (direct + reinsurance combined). The amounts must still sum to gross_premiums_written_gbp_m.
- IMPORTANT — year-of-account result breakdown: Many Lloyd's reports break down the overall result by year of account, e.g. "The result is a profit of £7,833,000, of which a loss of £5,556,000 is attributable to the {report_year} year of account, a profit of £14,806,000 is attributable to the {report_year - 1} year of account and a loss of £1,417,000 is attributable to the {report_year - 2} and prior years of account." In this example, the {report_year - 1} YOA profit (£14.806m) and the {report_year - 2} & prior YOA loss (-£1.417m) are BOTH prior year development. The NET prior year development = sum of all non-current-year components = £14.806m + (-£1.417m) = £13.389m. Since this is a net profit on prior years, direction = "release", prior_year_development_gbp_m = -13.389 (negative = release). This breakdown is a PRIMARY source for prior year development — look for it in the Managing Agent's Report or Underwriter's Report. Also look for "The [YYYY] & prior years of account is closing with a collectable loss/profit of £X" which indicates the closure result for older years.
- IMPORTANT — Claims Development Table (triangle): LAST RESORT — only use this if the "Movement in prior year's provision" note, narrative text, and year-of-account result breakdown are all unavailable. Must use the GROSS claims development table, NOT the net. Most reports contain an "Analysis of claims development" or "Claims development table" showing cumulative gross claims by underwriting year across development periods. To extract the movement:
  1. Look at the BOTTOM ROW ("Current estimate of cumulative claims") — these are the latest estimates for each underwriting year.
  2. Compare each UW year's current estimate to its estimate from the PREVIOUS diagonal (one row up in the same column).
  3. EXCLUDE the two most recent underwriting years ({report_year} and {report_year - 1}) — these are still in their initial development period and movements there are not "prior year development".
  4. Sum the differences for all remaining UW years: total_pyd = SUM(current_estimate[uw_year] - previous_estimate[uw_year]) for uw_year <= {report_year - 2}.
  5. A DECREASE in the cumulative estimate = favourable development = release (NEGATIVE sign).
  6. An INCREASE = adverse development = strengthening (POSITIVE sign).
  WORKED EXAMPLE: For a report year-end 2022 with this triangle (£m):
    UW Year:            2017    2018    2019    2020    2021    2022
    At end of UW year:  380.9   376.2   177.8   134.0   260.6   489.1
    One year later:     412.5   434.0   263.0   228.8   321.4
    Two years later:    426.9   409.8   277.6   236.6
    Three years later:  433.1   394.1   280.5
    Four years later:   422.6   390.3
    Five years later:   420.8
    Current estimate:   420.8   390.3   280.5   236.6   321.4   489.1
  Exclude 2022 and 2021. For each remaining UW year, the "previous diagonal" is one row up in the same column:
    2017: current=420.8, previous=422.6 (Four years later), change = 420.8-422.6 = -1.8
    2018: current=390.3, previous=394.1 (Three years later), change = 390.3-394.1 = -3.8
    2019: current=280.5, previous=277.6 (Two years later),   change = 280.5-277.6 = +2.9
    2020: current=236.6, previous=228.8 (One year later),    change = 236.6-228.8 = +7.8
  Total PYD = -1.8 + (-3.8) + 2.9 + 7.8 = +5.1 (net strengthening, positive sign).
  CAUTION — "X and prior" aggregate column: If the triangle has a column like "2012 and prior" that aggregates multiple older UW years, EXCLUDE it from the diagonal comparison — you cannot compute a valid previous diagonal for an aggregate column. Only use individual UW year columns.
  CAUTION — RITC distortions: If the report mentions "Reinsurance to Close" (RITC) of another syndicate during the year, the claims development triangle will be distorted. The RITC adds the absorbed syndicate's reserves to existing UW years, causing apparent large increases that are NOT genuine reserve deterioration. Look in the Technical Provisions note for an RITC line item. If RITC is present, the triangle is UNRELIABLE for computing prior year development — prefer other sources (narrative text, movement notes) or return null with a data_quality_note explaining the RITC distortion.
  This is a CROSS-CHECK source — use it to verify the figure from the "Movement in prior year's provision" note or narrative. If the note is missing and no RITC distortion exists, this table may be the primary source.
- IMPORTANT — Loss Ratio Development Table (fallback if no absolute claims triangle exists): Some reports show a GROSS loss ratio development table instead of absolute claims amounts. The table shows cumulative gross loss ratios (%) by underwriting year across development periods. To compute prior year development from this:
  1. For each UW year (excluding the two most recent), find the change in loss ratio: current estimate minus previous diagonal (one row up).
  2. Multiply the change in loss ratio by the GROSS PREMIUMS WRITTEN for that UW year to get the absolute £m amount.
  3. Sum across all older UW years.
  WORKED EXAMPLE (report year-end 2014):
    UW Year:    2008  2009  2010  2011  2012
    Current:     66%   43%   94%   68%   40%
    Previous:    67%   45%   95%   71%   44%
    Change:      -1%   -2%   -1%   -3%   -4%
  If gross premiums for 2008 were £100m: 2008 contribution = -1% × 100 = -£1.0m (release).
  Sum all UW year contributions for total PYD.
  NOTE: You need the gross premiums per UW year — look in the premium development table, segmental analysis, or the premium line of the P&L Technical Account for each year. If premiums per UW year are not available, use the total gross premiums as an approximation with a data_quality_note.
- IMPORTANT — year-of-account closure language: Lloyd's syndicates close years of account after 3 years. Phrases like "profit for the closed year of account", "improvement on forecast result", "return on capacity of X%" indicate favourable prior year development (release). Phrases like "deterioration on closed year", "loss on closed year of account" indicate adverse development (strengthening). The DIFFERENCE between the final result and the prior forecast is a useful cross-check for the prior year development amount (e.g. if profit improved from £40.1m forecast to £41.4m actual, the improvement of £1.3m suggests a release). Use this as supporting evidence alongside the primary reserve movement sources.
- prior_year_events: List specific named events from years BEFORE the report year ({report_year}) that the document mentions as affecting claims or reserves. For a {report_year} report, any event from {report_year - 1} or earlier is a prior year event. Look in the Technical Reserves note for "{report_year - 1} events" subsections — these are prior year events. Also look in the Underwriter's Report for references to events from prior years (e.g. deterioration on older losses). Include both adverse and favourable impacts.
- named_events: List ALL specific named catastrophe events, large losses, and significant incidents mentioned ANYWHERE in the document — including the Technical Reserves note (which often has "{report_year-1} events" and "{report_year} events" subsections), the Underwriter's Report divisional reviews, and the Managing Agent's Report. Include events from both the current year AND prior years. Look for: hurricanes, typhoons, earthquakes, floods, snowfall, hailstorms, tornadoes, vessel losses, industrial accidents, terrorist attacks, and any other specifically named loss events. Empty list only if genuinely no named events appear in the document."""


def sanitize_ascii(obj):
    """Recursively replace non-ASCII characters with ASCII equivalents."""
    if isinstance(obj, str):
        # Common Unicode → ASCII replacements
        obj = obj.replace("\u00a3", "GBP ")  # £
        obj = obj.replace("\u20ac", "EUR ")   # €
        obj = obj.replace("\u0024", "$")      # $
        obj = obj.replace("\u2013", "-")      # en dash
        obj = obj.replace("\u2014", "-")      # em dash
        obj = obj.replace("\u2018", "'")      # left single quote
        obj = obj.replace("\u2019", "'")      # right single quote
        obj = obj.replace("\u201c", '"')      # left double quote
        obj = obj.replace("\u201d", '"')      # right double quote
        obj = obj.replace("\u2026", "...")     # ellipsis
        obj = obj.replace("\u00b1", "+/-")    # plus-minus
        obj = obj.replace("\u00d7", "x")      # multiplication sign
        # Strip any remaining non-ASCII
        obj = re.sub(r'[^\x00-\x7F]', '', obj)
        return obj
    elif isinstance(obj, dict):
        return {sanitize_ascii(k): sanitize_ascii(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_ascii(item) for item in obj]
    return obj


def add_metadata(data, model, report_path, content_hash):
    """Add pipeline metadata to extraction result."""
    data["source_type"] = "syndicate"
    data["source_file"] = str(report_path)
    data["content_hash"] = content_hash
    data["standardized_at"] = datetime.now(timezone.utc).isoformat()
    data["standardization_model"] = model
    return data


def _parse_net_pyd_from_text(reserve_text, direction=None):
    """Parse a net-of-reinsurance PYD amount from narrative reserve text.

    Returns signed float in millions (negative=release, positive=strengthening),
    or None if no clear amount can be parsed.
    """
    if not reserve_text:
        return None
    text = reserve_text.lower()
    # Match patterns like "release of GBP 1.3m", "release of £1.3m",
    # "strengthening of GBP 2.9m", etc.
    patterns = [
        r'(?:reserve\s+)?(?:release|surplus)\s+of\s+(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)\s*(?:m|million)',
        r'(?:strengthening|deterioration|deficit)\s+of\s+(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)\s*(?:m|million)',
        r'(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)\s*(?:m|million)\s+(?:net\s+)?(?:release|surplus)',
        r'(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)\s*(?:m|million)\s+(?:net\s+)?(?:strengthening|deterioration)',
        # Also match "release of GBP 1.3m" without explicit unit suffix
        r'(?:reserve\s+)?(?:release|surplus)\s+of\s+(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)',
        r'(?:strengthening|deterioration|deficit)\s+of\s+(?:gbp|£|usd|\$|eur|€)\s*([\d,.]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(1).replace(",", "")
            try:
                amount = float(amount_str)
            except ValueError:
                continue
            # Determine sign from context around the match
            context = text[max(0, match.start() - 50):match.end() + 50]
            is_release = any(w in context for w in ["release", "surplus", "favourable", "favorable"])
            is_strength = any(w in context for w in ["strengthening", "deterioration", "deficit", "adverse"])
            if is_release:
                return -amount
            elif is_strength:
                return amount
            elif direction == "release":
                return -amount
            elif direction == "strengthening":
                return amount
            # Can't determine sign — skip
            continue
    return None


def _extract_accounting_numbers(line):
    """Extract numbers from a line with accounting notation.

    Handles:  (217,691) → -217691.0,  44,434 → 44434.0.
    Skips dashes (–, -) and year-like 4-digit integers (> 1900) that
    appear WITHOUT comma separators (e.g. "2011" is a year, but "2,402"
    is a monetary value in thousands).
    """
    numbers = []
    for match in re.finditer(r'\(([\d,]+(?:\.\d+)?)\)|([\d,]+(?:\.\d+)?)', line):
        neg_val, pos_val = match.groups()
        if neg_val:
            numbers.append(-float(neg_val.replace(',', '')))
        elif pos_val:
            raw = pos_val.replace(',', '')
            val = float(raw)
            # Skip year-like numbers: exactly 4 digits, > 1900, and no comma
            # in the original (so "2011" is a year but "2,402" is monetary)
            if '.' not in raw and len(raw) == 4 and val > 1900 and ',' not in pos_val:
                continue
            numbers.append(val)
    return numbers


def _extract_pyd_from_reserves_movement(pages, report_year):
    """Extract opening reserves and PYD from a 'Movement in Underwriting
    Reserves' note.

    Some syndicates (particularly those with RITC) present reserve movements
    in this table rather than a claims development triangle.  Common for
    syndicates that assumed prior year reserves via RITC.

    Expected format::

        Movement in Underwriting Reserves
                            Reserves  Exchange  RITC
                            £'000     £'000     £'000
        2011 and prior
        Opening balance     (217,691) (10,764)  (228,455)
        Change in year       44,434    2,402     46,836

    Returns dict with 'pyd', 'opening_reserves', 'details' or None.
    """
    if not pages:
        return None

    for _page_num, text in pages:
        text_lower = text.lower()
        if not re.search(r'movement\s+in\s+(underwriting\s+)?reserves', text_lower):
            continue
        if 'opening balance' not in text_lower:
            continue
        if 'change in year' not in text_lower:
            continue

        # Detect units (£'000 → divisor = 1000)
        units_divisor = 1.0
        if re.search(r"[£$\u00a3]\s*['\u2018\u2019]?\s*000", text):
            units_divisor = 1000.0
        elif "'000" in text or "\u2019000" in text:
            units_divisor = 1000.0

        lines = text.split('\n')
        in_prior_section = False
        opening_balance = None
        change_in_year = None

        # State for collecting numbers after a label line.
        # PyMuPDF often splits tables into separate lines:
        #   "Opening balance"
        #   "(217,691)"
        #   "(10,764)"
        #   "(228,455)"
        collecting_for = None  # "opening" or "change"
        collected_nums = []

        def _flush_collected():
            nonlocal opening_balance, change_in_year, collecting_for, collected_nums
            if collecting_for == "opening" and collected_nums:
                opening_balance = collected_nums[-1]
            elif collecting_for == "change" and collected_nums:
                change_in_year = collected_nums[-1]
            collecting_for = None
            collected_nums = []

        for line in lines:
            stripped = line.strip()
            line_lower = stripped.lower()

            if re.search(r'\d{4}\s+and\s+prior', line_lower):
                _flush_collected()
                in_prior_section = True
                continue

            # Next year header (e.g. "2012 pure") ends the section
            if in_prior_section and re.match(r'\d{4}\s', line_lower):
                _flush_collected()
                break

            if not in_prior_section:
                continue

            # Check if this line is a label
            if 'opening balance' in line_lower:
                _flush_collected()
                # Numbers may be on the same line or subsequent lines
                nums = _extract_accounting_numbers(stripped)
                if nums:
                    opening_balance = nums[-1]
                else:
                    collecting_for = "opening"
                    collected_nums = []
                continue

            if 'change in year' in line_lower:
                _flush_collected()
                nums = _extract_accounting_numbers(stripped)
                if nums:
                    change_in_year = nums[-1]
                else:
                    collecting_for = "change"
                    collected_nums = []
                continue

            # If we're collecting numbers for a label, grab them
            if collecting_for and stripped:
                nums = _extract_accounting_numbers(stripped)
                if nums:
                    collected_nums.extend(nums)
                elif not re.match(r'^[\s\-\u2013\u2014\u2015]*$', stripped):
                    # Non-numeric, non-dash line → stop collecting
                    _flush_collected()

        _flush_collected()

        if opening_balance is not None and change_in_year is not None:
            opening_m = round(abs(opening_balance) / units_divisor, 3)
            # Positive change = liability decreased = release → PYD negative
            pyd_m = round(-change_in_year / units_divisor, 3)
            change_m = round(change_in_year / units_divisor, 3)
            return {
                "pyd": pyd_m,
                "opening_reserves": opening_m,
                "details": (f"from reserves movement note "
                            f"(opening={opening_m:.1f}m, "
                            f"change={change_m:+.1f}m)"),
            }

    return None


def _extract_pyd_from_loss_ratio_triangle(page_text, report_year):
    """Extract PYD from a loss ratio development triangle page.

    Some syndicates (e.g. Beazley 2623) present claims development as
    cumulative loss ratios (%) rather than absolute claims amounts.  When
    the page also contains a "Total ultimate losses ($m)" or similar row
    with absolute values per UW year, we can compute PYD as:

        pyd_j = total_ultimate_j * (current_ratio_j - prev_ratio_j)
                / current_ratio_j

    Returns (pyd_float, details_str) or (None, reason_str).
    """
    lines = page_text.split("\n")
    lines_stripped = [l.strip() for l in lines]

    # ── Step 1: Detect that this is a loss ratio triangle ──
    text_lower = page_text.lower()
    # Accept several header variants: "claims development", "gross ratios",
    # or development period labels with % symbols.
    # Do NOT accept "net ratios" alone — we need gross data for PYD.
    has_header = ("claims development" in text_lower
                  or "gross ratios" in text_lower
                  or ("12 months" in text_lower and "24 months" in text_lower))
    if not has_header:
        return None, "no claims development / loss ratio header"
    # Reject pages with net ratios but no gross ratios
    if ("net ratios" in text_lower
            and "gross ratios" not in text_lower
            and "gross claims development" not in text_lower):
        return None, "net-only loss ratio page (no gross data)"
    # Must have percentage indicators
    pct_count = text_lower.count("%")
    if pct_count < 3:
        return None, "too few % symbols for loss ratio triangle"

    # ── Step 2: Find UW years ──
    # Track "ae" (and earlier) aggregate columns that have no ratio data
    # but DO have a value in the ultimates row.
    all_years = []       # all column years (including ae aggregate)
    ae_columns = set()   # indices of aggregate columns (no ratio data)
    header_end_idx = 0

    # Strategy A: years on one line (e.g. "2010ae  2011  2012  2013")
    for i, line in enumerate(lines_stripped):
        year_matches = re.findall(r'\b((?:19|20)\d{2})\s*(ae?)?\b', line)
        if len(year_matches) >= 3:
            seen = set()
            for y_str, suffix in year_matches:
                y = int(y_str)
                if 1990 <= y <= 2030 and y not in seen:
                    seen.add(y)
                    if suffix:  # "ae" or "a" suffix
                        ae_columns.add(len(all_years))
                    all_years.append(y)
            if len(all_years) >= 3:
                header_end_idx = i
                break
            all_years.clear()
            ae_columns.clear()

    # Strategy B: years on consecutive lines (columnar PyMuPDF output)
    if not all_years:
        for i, line in enumerate(lines_stripped):
            m = re.match(r'^((?:19|20)\d{2})\s*(ae?)?$', line)
            if m:
                has_ae = bool(m.group(2))
                # Check if next non-blank line is "e" (continuation of "ae" suffix)
                if not has_ae and (i + 1 < len(lines_stripped)
                                   and lines_stripped[i + 1].lower() in ("e", "ae")):
                    has_ae = True
                year_run = [(int(m.group(1)), has_ae)]
                j = i + 1
                while j < len(lines_stripped):
                    m2 = re.match(r'^((?:19|20)\d{2})\s*(ae?)?$', lines_stripped[j])
                    if m2:
                        year_run.append((int(m2.group(1)), bool(m2.group(2))))
                        j += 1
                    elif lines_stripped[j].lower() in ("total", "total\xa0"):
                        j += 1
                        break
                    elif lines_stripped[j].lower() in ("e", "ae"):
                        # Mark previous year as aggregate
                        if year_run:
                            yr, _ = year_run[-1]
                            year_run[-1] = (yr, True)
                        j += 1
                        continue
                    else:
                        break
                if len(year_run) >= 3:
                    for idx, (yr, is_ae) in enumerate(year_run):
                        all_years.append(yr)
                        if is_ae:
                            ae_columns.add(idx)
                    header_end_idx = j - 1
                    break

    if len(all_years) < 3:
        return None, "fewer than 3 UW years found"

    all_years_sorted = sorted(all_years)
    # Ratio-bearing columns = all columns EXCEPT aggregate "ae" columns
    uw_years = [y for i, y in enumerate(all_years_sorted)
                if i not in ae_columns]
    n_cols_total = len(all_years_sorted)  # includes ae columns (for ultimates)
    n_cols = len(uw_years)                # ratio-bearing columns only

    if max(uw_years) > report_year or max(uw_years) < report_year - 5:
        return None, f"max UW year {max(uw_years)} outside range for report year {report_year}"

    # ── Step 3: Parse the loss ratio grid ──
    # Skip % header lines, dev period labels; collect numeric values
    dev_labels = [
        r"^\d+\s+months?\b", r"^(one|two|three|four|five|six|seven|eight|nine|ten)\b",
        r"year\s+later", r"years?\s+later", r"at\s+end",
    ]
    stop_labels = [
        "total ultimate", "gross claims liab",
        "less paid", "less unearned", "less non", "net claims liab",
        "net claims development",  # stop before net section on same page
        "net ratios",              # stop before net loss ratio section (e.g. Beazley 623)
        "underwriting year - net",  # stop before net loss ratio section
        "underwriting year -net",   # variant without space
        "underwriting year- net",   # variant
    ]

    all_ratios = []
    after_header = False
    for line in lines_stripped[header_end_idx + 1:]:
        line_lower = line.lower()
        if not line_lower:
            continue
        # Skip standalone % symbols
        if line_lower.strip() == "%":
            after_header = True
            continue
        # In columnar PyMuPDF output each ratio is on its own line
        # (e.g. "127%").  The first line with "%" is the first real
        # ratio value, NOT a header to skip.  Only skip if the line
        # looks like a pure header (e.g. "%" alone, already handled
        # above).  Set after_header but DO NOT skip the line.
        if not after_header and "%" in line_lower:
            after_header = True
            # fall through to extract the value from this line
        if not after_header:
            continue
        # Stop at summary rows
        if any(sl in line_lower for sl in stop_labels):
            break
        # Stop at "Gross claims" when it's NOT the header "Gross Claims Development"
        if (line_lower.strip() == "gross claims"
                or (line_lower.strip().startswith("gross claims")
                    and "development" not in line_lower)):
            break
        # Skip dev period labels
        if any(re.search(p, line_lower) for p in dev_labels):
            continue
        # Skip pure text lines
        if re.match(r'^[a-z]', line_lower) and not re.search(r'\d', line):
            continue
        # Extract numbers
        nums = re.findall(r'[\-]?[\d,]+\.?\d*', line)
        for n in nums:
            try:
                val = float(n.replace(",", ""))
                # Loss ratios should be between 0 and 200%
                if 0 < val <= 200:
                    all_ratios.append(val)
            except ValueError:
                continue

    # Group ratios into development rows.
    # Use uw_years (excludes ae columns) for grouping since ae columns
    # never have ratio data in the grid — they are aggregates that only
    # appear in the "Total ultimate losses" row.
    max_dev = report_year - min(uw_years) + 1
    dev_rows = []
    idx = 0
    for d in range(max_dev):
        expected = sum(1 for y in uw_years if report_year - y >= d)
        if expected < 1:
            break
        if idx + expected > len(all_ratios):
            break
        row = list(all_ratios[idx:idx + expected])
        row += [None] * (n_cols - len(row))
        dev_rows.append(row[:n_cols])
        idx += expected

    if len(dev_rows) < 2:
        return None, f"only {len(dev_rows)} development rows parsed from loss ratio grid"

    # ── Step 4: Find "Total ultimate losses" row ──
    # Only search within the GROSS section — stop before "Net Claims Development".
    # Collect n_cols_total values (including ae aggregate column),
    # then filter out ae columns to align with uw_years.
    gross_end_idx = len(lines_stripped)
    for i, line in enumerate(lines_stripped):
        if "net claims development" in line.lower():
            gross_end_idx = i
            break

    # The label "Total ultimate losses ($m)" may span 1-4 lines in PyMuPDF
    # output, e.g.:
    #   "Total ultimate losses ($m)  8,061.0  756.9 ..."   (all on one line)
    #   "Total"  /  "ultimate"  /  "losses"  /  "($m)"     (four lines)
    # We accumulate label tokens until we've seen "total" + "ultimate", then
    # start collecting numeric values.
    raw_ultimates = None
    label_state = 0          # 0=nothing, 1=seen "total", 2=seen "ultimate"
    collecting_ult = False
    raw_ult_vals = []
    for line in lines_stripped[:gross_end_idx]:
        line_lower = line.lower().strip()

        # Once we're collecting values, stop at next section
        if collecting_ult:
            if any(sl in line_lower for sl in
                   ["less paid", "less unearned", "gross claims liab",
                    "less non", "net claims"]):
                break
            if not line_lower:
                continue
            nums = re.findall(r'[\-]?[\d,]+\.?\d*|\([\d,]+\.?\d*\)', line)
            for n in nums:
                n_clean = n.replace(",", "").replace("(", "-").replace(")", "")
                try:
                    val = float(n_clean)
                    if abs(val) > 1:  # absolute amounts, not ratios
                        raw_ult_vals.append(val)
                except ValueError:
                    continue
            if len(raw_ult_vals) >= n_cols_total:
                break
            continue

        # Label accumulation: look for "total" then "ultimate"
        if label_state == 0 and "total" in line_lower:
            if "total ultimate" in line_lower:
                label_state = 2  # both words on same line
            else:
                label_state = 1
        elif label_state == 1 and "ultimate" in line_lower:
            label_state = 2

        # Once we've seen "total ultimate", check if this line also has values
        if label_state == 2:
            # Extract any numbers that might be on the same line as the label
            nums = re.findall(r'[\-]?[\d,]+\.?\d*|\([\d,]+\.?\d*\)', line)
            vals = []
            for n in nums:
                n_clean = n.replace(",", "").replace("(", "-").replace(")", "")
                try:
                    val = float(n_clean)
                    if abs(val) > 1:  # absolute amounts, not ratios
                        vals.append(val)
                except ValueError:
                    continue
            if vals:
                raw_ult_vals.extend(vals)
            collecting_ult = True

    if len(raw_ult_vals) >= n_cols_total:
        raw_ultimates = raw_ult_vals

    # ── Step 4b: Fallback — "Gross claims liabilities" rows ──
    # Some reports (e.g. Beazley syndicates) show total claims liabilities
    # by UW year instead of "Total ultimate losses".  Structure:
    #   "Gross claims liabilities (Beazley managed level)"  → group totals
    #   "Less non XXXX share"                               → deduction
    #   "Gross claims liabilities (XXXX share)"             → syndicate share
    # Prefer the syndicate share row; fall back to managed level.
    gcl_is_share = False
    gcl_managed = []
    gcl_share = []
    if raw_ultimates is None or len(raw_ultimates) < n_cols_total:
        gcl_state = "scan"  # scan, label, collect
        gcl_label = ""
        gcl_vals = []
        gcl_is_deduction = False

        for line in lines_stripped[:gross_end_idx]:
            ll = line.lower().strip()

            if gcl_state == "collect":
                # Stop at next section label
                if any(kw in ll for kw in
                       ["less non", "less paid", "net claims",
                        "gross claims", "net claim"]):
                    # Save collected values
                    if not gcl_is_deduction and len(gcl_vals) >= n_cols_total:
                        if "share" in gcl_label and "less" not in gcl_label:
                            gcl_share = gcl_vals[:n_cols_total]
                        else:
                            gcl_managed = gcl_vals[:n_cols_total]
                    # Start new label
                    gcl_state = "label"
                    gcl_label = ll
                    gcl_vals = []
                    gcl_is_deduction = "less" in ll
                    continue
                if not ll:
                    continue
                nums = re.findall(r'[\-]?[\d,]+\.?\d*|\([\d,]+\.?\d*\)', line)
                for n in nums:
                    n_clean = n.replace(",", "").replace("(", "-").replace(")", "")
                    try:
                        val = float(n_clean)
                        if abs(val) > 1:
                            gcl_vals.append(val)
                    except ValueError:
                        continue
                if len(gcl_vals) >= n_cols_total:
                    if not gcl_is_deduction:
                        if "share" in gcl_label and "less" not in gcl_label:
                            gcl_share = gcl_vals[:n_cols_total]
                        else:
                            gcl_managed = gcl_vals[:n_cols_total]
                    gcl_state = "scan"
                    gcl_vals = []
                continue

            if gcl_state == "label":
                # Still accumulating multi-line label text.
                # A line is numeric only if it starts with a digit (after
                # stripping parens/minus) AND does not contain alphabetic text
                # (e.g., "(3623 share)" is a label, not a number).
                is_numeric_line = (re.search(r'^\d', ll.lstrip("(-"))
                                   and not re.search(r'[a-z]', ll))
                if is_numeric_line:
                    # Numeric line — switch to collecting
                    gcl_state = "collect"
                    # fall through to collect from this line
                    nums = re.findall(r'[\-]?[\d,]+\.?\d*|\([\d,]+\.?\d*\)', line)
                    for n in nums:
                        n_clean = n.replace(",", "").replace("(", "-").replace(")", "")
                        try:
                            val = float(n_clean)
                            if abs(val) > 1:
                                gcl_vals.append(val)
                        except ValueError:
                            continue
                    continue
                gcl_label += " " + ll
                continue

            # gcl_state == "scan"
            if (("gross claims" in ll or ("claims" in ll and "liabilit" in ll))
                    and "development" not in ll):
                gcl_state = "label"
                gcl_label = ll
                gcl_vals = []
                gcl_is_deduction = False
            elif "less non" in ll:
                gcl_state = "label"
                gcl_label = ll
                gcl_vals = []
                gcl_is_deduction = True

        # Flush final section
        if gcl_state == "collect" and not gcl_is_deduction and len(gcl_vals) >= n_cols_total:
            if "share" in gcl_label and "less" not in gcl_label:
                gcl_share = gcl_vals[:n_cols_total]
            else:
                gcl_managed = gcl_vals[:n_cols_total]

        # Prefer syndicate share over managed level
        if len(gcl_share) >= n_cols_total:
            raw_ultimates = gcl_share
            raw_ult_vals = gcl_share
            gcl_is_share = True
        elif len(gcl_managed) >= n_cols_total:
            raw_ultimates = gcl_managed
            raw_ult_vals = gcl_managed

    if raw_ultimates is None or len(raw_ultimates) < n_cols_total:
        return None, (f"loss ratio triangle found ({n_cols} UW years, "
                      f"{len(dev_rows)} dev rows) but no 'Total ultimate losses' "
                      f"or 'Gross claims liabilities' row — cannot convert to "
                      f"absolute PYD")

    # Validate: reject if "ultimates" are actually year values (1990-2030).
    # This happens when "total" + "ultimately" co-occur in surrounding text
    # and year headers are mistakenly collected as ultimate amounts.
    year_like = sum(1 for v in raw_ultimates[:n_cols_total]
                    if 1990 <= v <= 2030 and v == int(v))
    if year_like >= n_cols_total * 0.5:
        return None, (f"loss ratio triangle found but 'ultimates' row contains "
                      f"{year_like}/{n_cols_total} year-like values "
                      f"({raw_ultimates[:n_cols_total]}) — false detection")

    # Keep only n_cols_total values (drop trailing "Total" column)
    raw_ultimates = raw_ultimates[:n_cols_total]
    # Filter out ae aggregate columns to align with uw_years
    ultimates = [v for i, v in enumerate(raw_ultimates) if i not in ae_columns]

    # ── Step 5: Compute PYD ──
    ult_source = ("gross claims liabilities (syndicate share)" if gcl_is_share
                  else "gross claims liabilities (managed level)" if gcl_managed
                  else "total ultimate losses")
    details = [f"Loss ratio triangle: {n_cols} UW years "
               f"({min(uw_years)}-{max(uw_years)}), {len(dev_rows)} dev rows, "
               f"ultimates from {ult_source}"]
    total_pyd = 0.0
    used_years = 0

    for col_idx, uw_year in enumerate(uw_years):
        # Exclude two most recent UW years
        if uw_year >= report_year - 1:
            continue

        # Find current (last non-null) and previous diagonal
        current_ratio = None
        current_row = None
        for row_idx in range(len(dev_rows) - 1, -1, -1):
            if dev_rows[row_idx][col_idx] is not None:
                current_ratio = dev_rows[row_idx][col_idx]
                current_row = row_idx
                break

        if current_ratio is None or current_row is None or current_row == 0:
            continue

        prev_ratio = dev_rows[current_row - 1][col_idx]
        if prev_ratio is None:
            continue

        ult = ultimates[col_idx]
        if ult <= 0 or current_ratio <= 0:
            continue

        # PYD = total_ultimate * (current_ratio - prev_ratio) / current_ratio
        pyd_j = ult * (current_ratio - prev_ratio) / current_ratio
        total_pyd += pyd_j
        used_years += 1

        details.append(f"  {uw_year}: ratio {prev_ratio:.1f}% -> {current_ratio:.1f}% "
                       f"(chg {current_ratio - prev_ratio:+.1f}pp), "
                       f"ult={ult:.1f}m, pyd={pyd_j:+.3f}m")

    if used_years == 0:
        return None, "no usable UW years in loss ratio triangle"

    total_pyd = round(total_pyd, 3)
    details.append(f"  Total PYD = {total_pyd:+.3f}m ({used_years} UW years)")
    return total_pyd, "\n".join(details)


def _extract_pyd_from_provisions_text(page_text):
    """Extract gross prior year development from a provisions movement note.

    Looks for "claims incurred in prior underwriting year" (or similar)
    followed by the gross value in $'000 or £'000.  Returns (pyd_m, details)
    or (None, reason).
    """
    lines = page_text.split("\n")
    for i, line in enumerate(lines):
        ll = line.lower().strip()
        if ("prior" in ll
                and ("underwriting" in ll or "year" in ll)
                and "incurred" in ll
                and "includes" not in ll):
            # Found the prior year claims incurred label in a provisions
            # table row.  Skip narrative text (which contains "includes")
            # — those are handled by _parse_pyd_from_pl_narrative().
            # The gross value is typically on the same line or the next line.
            # Collect the first numeric value after the label.
            candidates = []
            # Check same line first
            nums = re.findall(r'\(?([\d,]+\.?\d*)\)?', line)
            for n in nums:
                try:
                    val = float(n.replace(",", ""))
                    if val > 0:
                        candidates.append(val)
                except ValueError:
                    pass
            # Check next 3 lines
            for j in range(i + 1, min(i + 4, len(lines))):
                nline = lines[j].strip()
                if not nline:
                    continue
                # Stop at next label
                if re.match(r'^[A-Z]', nline) and not re.search(r'^\(?\d', nline):
                    break
                nums = re.findall(r'\(?([\d,]+\.?\d*)\)?', nline)
                for n in nums:
                    try:
                        val = float(n.replace(",", ""))
                        if val > 0:
                            candidates.append(val)
                    except ValueError:
                        pass
                if candidates:
                    break

            if candidates:
                raw_val = candidates[0]  # First value is gross column
                # Check sign: if line has parentheses around number, it's negative
                sign_line = line + " " + (lines[i + 1].strip() if i + 1 < len(lines) else "")
                # Detect units from surrounding text
                is_thousands = False
                for ctx_line in lines[max(0, i - 10):i + 5]:
                    ctx_lower = ctx_line.lower()
                    # Match '000, \u2019000, \u2018000, "000 and similar
                    if any(kw in ctx_lower for kw in [
                        "'000", "\u2019000", "\u2018000", "\u0027000",
                        "000s", "thousands",
                    ]):
                        is_thousands = True
                        break
                    # Also detect "$'000" / "£'000" with various quote styles
                    if re.search(r"[£$€]\s*['\u2018\u2019\u0027]?\s*000", ctx_lower):
                        is_thousands = True
                        break

                pyd_m = raw_val / 1000 if is_thousands else raw_val
                pyd_m = round(pyd_m, 3)
                return pyd_m, (f"Provisions note: 'Claims incurred in prior "
                               f"underwriting year' = {raw_val:,.0f}"
                               f"{' ($000)' if is_thousands else ''} = {pyd_m:+.3f}m")

    return None, "no 'claims incurred in prior underwriting year' found"


def _parse_pyd_from_pl_narrative(page_text):
    """Extract net PYD from P&L technical account narrative text.

    Matches patterns like:
      "includes £34,490k (2013: £51,518k) of releases in respect of prior accident years"
      "includes a net release of £5.3m relating to prior years"

    Returns (pyd_m, details_str) or (None, reason_str).
    """
    # Join lines to handle line-wrapping; work within single sentences
    text = " ".join(page_text.split())
    tl = text.lower()

    if "prior" not in tl or "includes" not in tl:
        return None, "no P&L narrative pattern"

    # Pattern A: "includes £AMOUNT[k/m] ... of {releases/strengthening} ... prior ... year"
    m = re.search(
        r'includes\b[^.]*?'
        r'[£$€]\s*([\d,]+(?:\.\d+)?)\s*([km])?\b'
        r'[^.]*?'
        r'\bof\s+(releases?|strengthening|strengthenings|adverse\s+development)\b'
        r'[^.]*?'
        r'\bprior\s+(?:accident\s+|underwriting\s+)?years?\b',
        tl,
    )
    if m:
        raw_str = m.group(1).replace(",", "")
        unit = m.group(2)
        direction_word = m.group(3)
    else:
        # Pattern B: "includes ... release/strengthening of £AMOUNT ... prior year"
        m = re.search(
            r'includes\b[^.]*?'
            r'\b(releases?|strengthening|strengthenings|adverse\s+development)\b'
            r'[^.]*?'
            r'[£$€]\s*([\d,]+(?:\.\d+)?)\s*([km])?\b'
            r'[^.]*?'
            r'\bprior\s+(?:accident\s+|underwriting\s+)?years?\b',
            tl,
        )
        if not m:
            return None, "no P&L narrative PYD pattern matched"
        direction_word = m.group(1)
        raw_str = m.group(2).replace(",", "")
        unit = m.group(3)

    raw_val = float(raw_str)
    if raw_val <= 0 or 1900 <= raw_val <= 2099:
        return None, f"P&L narrative: extracted value {raw_val} looks like a year"

    # Determine units: explicit k/m suffix, or context-based detection
    is_thousands = (unit == "k")
    if not is_thousands and unit != "m":
        if any(kw in tl for kw in [
            "'000", "\u2019000", "\u2018000", "\u0027000", "thousands",
        ]):
            is_thousands = True
        if re.search(r"[£$€]\s*['\u2018\u2019\u0027]?\s*000", tl):
            is_thousands = True

    pyd_m = raw_val / 1000 if is_thousands else raw_val
    pyd_m = round(pyd_m, 3)

    # Sign: release = negative (favourable), strengthening = positive (adverse)
    if "release" in direction_word:
        pyd_m = -abs(pyd_m)

    return pyd_m, (f"P&L narrative: '{direction_word}' of "
                   f"{raw_val:,.0f}{' (£000)' if is_thousands else ''}"
                   f" = {pyd_m:+.3f}m")


def _parse_pyd_from_yoa_narrative(page_text):
    """Extract PYD from Year of Account (YOA) style narrative text.

    Matches patterns like:
      "Prior year movements of £5.8m in 2014, comprising £5.4m adverse
       movements in claims reserving and £0.4m movement in premiums"

    Some Lloyd's syndicates (especially smaller or YOA-accounting ones)
    state PYD in this format rather than in a provisions table or
    "includes" P&L sentence.

    Returns (pyd_m, details_str) or (None, reason_str).
    """
    # Join lines to handle line-wrapping
    text = " ".join(page_text.split())
    tl = text.lower()

    if "prior year" not in tl or "movement" not in tl:
        return None, "no YOA narrative pattern"

    # Pattern: "Prior year movements of £X.Xm"
    # May also appear as "Prior year movements of £X,XXXk" or "£X,XXX"
    m = re.search(
        r'prior\s+year\s+movements?\s+of\s+'
        r'[£$€]\s*([\d,]+(?:\.\d+)?)\s*([km])?\b',
        tl,
    )
    if not m:
        return None, "no YOA narrative PYD pattern matched"

    raw_str = m.group(1).replace(",", "")
    unit = m.group(2)
    raw_val = float(raw_str)

    if raw_val <= 0 or 1900 <= raw_val <= 2099:
        return None, f"YOA narrative: extracted value {raw_val} looks like a year"

    # Determine units
    is_thousands = (unit == "k")
    if not is_thousands and unit != "m":
        if any(kw in tl for kw in [
            "'000", "\u2019000", "\u2018000", "\u0027000", "thousands",
        ]):
            is_thousands = True
        if re.search(r"[£$€]\s*['\u2018\u2019\u0027]?\s*000", tl):
            is_thousands = True

    pyd_m = raw_val / 1000 if is_thousands else raw_val
    pyd_m = round(pyd_m, 3)

    # Determine direction from context:
    # Look for "adverse" or "strengthening" near the amount → positive (adverse)
    # Look for "release" or "favourable" near the amount → negative (favourable)
    # Default to adverse (positive) if direction is unclear — "movements" without
    # qualifier typically means adverse in Lloyd's YOA reports.
    context = tl[max(0, m.start() - 50):m.end() + 200]
    if any(kw in context for kw in ["release", "favourable", "favorable", "surplus"]):
        pyd_m = -abs(pyd_m)
    elif any(kw in context for kw in ["adverse", "strengthen", "deficit"]):
        pyd_m = abs(pyd_m)
    else:
        # Ambiguous — check the broader sentence for direction cues
        pyd_m = abs(pyd_m)  # default to adverse

    direction = "adverse" if pyd_m > 0 else "favourable"
    return pyd_m, (f"YOA narrative: 'prior year movements of' "
                   f"{raw_val:,.1f}{unit or ''}"
                   f"{' (£000)' if is_thousands else ''}"
                   f" [{direction}] = {pyd_m:+.3f}m")


def _parse_pyd_from_general_narrative(page_text):
    """Extract PYD from general reserve narrative text.

    Matches patterns like:
      "a release of £4.7m of prior year reserves"
      "strengthening of £12.3m in prior year claims reserves"
      "there has been a release of £4.7m of prior year reserves"

    This is a catch-all parser for reports that don't use the "includes"
    P&L format or "prior year movements of" YOA format.

    Returns (pyd_m, details_str) or (None, reason_str).
    """
    text = " ".join(page_text.split())
    tl = text.lower()

    if "prior year" not in tl:
        return None, "no general narrative pattern"

    # Pattern A: "release/strengthening of £AMOUNT ... prior year [reserves]"
    m = re.search(
        r'\b(releases?|strengthening|strengthenings|adverse\s+development)\b'
        r'\s+of\s+'
        r'[£$€]\s*([\d,]+(?:\.\d+)?)\s*([km])?\b'
        r'.{0,40}'
        r'\bprior\s+year',
        tl,
    )
    if not m:
        # Pattern B: "£AMOUNT ... release/strengthening ... prior year"
        m = re.search(
            r'[£$€]\s*([\d,]+(?:\.\d+)?)\s*([km])?\b'
            r'.{0,40}'
            r'\b(releases?|strengthening|strengthenings|adverse\s+development)\b'
            r'.{0,40}'
            r'\bprior\s+year',
            tl,
        )
        if m:
            # Reorder groups: amount is group 1,2; direction is group 3
            raw_str = m.group(1).replace(",", "")
            unit = m.group(2)
            direction_word = m.group(3)
        else:
            return None, "no general narrative PYD pattern matched"
    else:
        direction_word = m.group(1)
        raw_str = m.group(2).replace(",", "")
        unit = m.group(3)

    raw_val = float(raw_str)
    if raw_val <= 0 or 1900 <= raw_val <= 2099:
        return None, f"general narrative: extracted value {raw_val} looks like a year"

    # Determine units
    is_thousands = (unit == "k")
    if not is_thousands and unit != "m":
        if any(kw in tl for kw in [
            "'000", "\u2019000", "\u2018000", "\u0027000", "thousands",
        ]):
            is_thousands = True
        if re.search(r"[£$€]\s*['\u2018\u2019\u0027]?\s*000", tl):
            is_thousands = True

    pyd_m = raw_val / 1000 if is_thousands else raw_val
    pyd_m = round(pyd_m, 3)

    # Sign: release = negative (favourable), strengthening = positive (adverse)
    if "release" in direction_word:
        pyd_m = -abs(pyd_m)
    else:
        pyd_m = abs(pyd_m)

    return pyd_m, (f"general narrative: '{direction_word}' of "
                   f"{raw_val:,.1f}{unit or ''}"
                   f"{' (£000)' if is_thousands else ''}"
                   f" = {pyd_m:+.3f}m")


def _normalize_top_level(obj, raw):
    """All callers expect a JSON object. LLMs occasionally wrap the object in
    a one-element array — unwrap that. Any other top-level type is treated as
    malformed so the caller's retry / fix-prompt logic kicks in."""
    if isinstance(obj, list) and len(obj) == 1 and isinstance(obj[0], dict):
        return obj[0]
    if not isinstance(obj, dict):
        raise json.JSONDecodeError(
            f"top-level JSON is {type(obj).__name__}, expected object", raw, 0)
    return obj


def parse_json_response(text):
    """Parse JSON from LLM response, stripping markdown fences if present.

    Always returns a JSON object (dict); raises json.JSONDecodeError otherwise.
    """
    raw = text.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

    # Try direct parse first
    try:
        return _normalize_top_level(json.loads(raw), raw)
    except json.JSONDecodeError:
        pass

    # Try to find a complete JSON object between first { and last }
    brace_start = raw.find("{")
    brace_end = raw.rfind("}")
    if brace_start != -1 and brace_end > brace_start:
        candidate = raw[brace_start:brace_end + 1]
        try:
            return _normalize_top_level(json.loads(candidate), candidate)
        except json.JSONDecodeError:
            pass

        # Fix common LLM JSON errors:
        # 1. Trailing commas before } or ]
        fixed = re.sub(r',\s*([}\]])', r'\1', candidate)
        # 2. Single-line // comments (outside strings)
        fixed = re.sub(r'//[^\n]*', '', fixed)
        # 3. Multi-line /* */ comments
        fixed = re.sub(r'/\*.*?\*/', '', fixed, flags=re.DOTALL)
        # 4. Unquoted property names: { key: "value" } → { "key": "value" }
        fixed = re.sub(r'(?<=[\{,])\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r' "\1":', fixed)
        # 5. Single quotes → double quotes (but not inside strings)
        # Only do this if no double-quoted strings exist at all
        if '"' not in fixed.replace('\\"', ''):
            fixed = fixed.replace("'", '"')
        try:
            return _normalize_top_level(json.loads(fixed), fixed)
        except json.JSONDecodeError:
            pass

    # Re-raise with original text for debugging
    return _normalize_top_level(json.loads(raw), raw)


def extract_with_gemini(report_path, file_bytes, content_hash, syndicate_num, report_year, model=GEMINI_MODEL):
    """Extract using Google Gemini."""
    prompt_text = build_prompt(syndicate_num, report_year)
    # Include content_hash in cache key so that changes to the slim PDF
    # (e.g. different page set after off-by-one fix) invalidate the cache.
    cache_key = _llm_cache_key(model, prompt_text + "|" + content_hash, syndicate_num, report_year)
    cached, hit = _llm_cache_load(cache_key)
    if hit:
        print(f"  [{model}] Cache hit — skipping API call")
        return cached["data"]

    print(f"  [{model}] Uploading {report_path.name}...")
    client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))
    t0 = time.time()
    uploaded_file = _gemini_call_with_retry(
        lambda: client.files.upload(file=report_path))
    print(f"  [{model}] Upload done ({time.time() - t0:.1f}s). Extracting...")

    gemini_config = GenerateContentConfig(
        temperature=0.0,
        response_mime_type="application/json",
    )
    t1 = time.time()
    response = _gemini_call_with_retry(lambda: client.models.generate_content(
        model=model,
        contents=[uploaded_file, build_prompt(syndicate_num, report_year)],
        config=gemini_config,
    ))
    print(f"  [{model}] Response received ({time.time() - t1:.1f}s)")

    # Retry up to twice on malformed JSON
    try:
        data = sanitize_ascii(parse_json_response(response.text))
    except json.JSONDecodeError as e:
        print(f"  [{model}] Malformed JSON ({e}), retrying (attempt 2)...")
        response = _gemini_call_with_retry(lambda: client.models.generate_content(
            model=model,
            contents=[uploaded_file, build_prompt(syndicate_num, report_year)],
            config=gemini_config,
        ))
        try:
            data = sanitize_ascii(parse_json_response(response.text))
        except json.JSONDecodeError as e2:
            print(f"  [{model}] Still malformed ({e2}), retrying with JSON fix prompt (attempt 3)...")
            fix_prompt = (
                "Your previous response was not valid JSON. "
                "Please re-output ONLY the JSON object, with no comments, "
                "no trailing commas, and no text outside the JSON braces. "
                "Here is what you output:\n\n" + response.text[:3000]
            )
            response = _gemini_call_with_retry(lambda: client.models.generate_content(
                model=model,
                contents=[uploaded_file, fix_prompt],
                config=gemini_config,
            ))
            data = sanitize_ascii(parse_json_response(response.text))
    data = add_metadata(data, model, report_path, content_hash)

    usage = response.usage_metadata
    input_tokens = usage.prompt_token_count
    output_tokens = usage.candidates_token_count
    total_tokens = usage.total_token_count

    prices = PRICING[model]
    cost = input_tokens * prices["input"] / 1_000_000 + output_tokens * prices["output"] / 1_000_000

    data["_extraction_meta"] = {
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "cost_usd": round(cost, 6),
    }

    print(f"  [{model}] Done. {input_tokens:,} in / {output_tokens:,} out. ${cost:.6f}")
    _llm_cache_save(cache_key, data, model=model,
                    syndicate_num=syndicate_num, report_year=report_year)
    return data


def extract_with_openai(report_path, file_bytes, content_hash, syndicate_num, report_year, model=OPENAI_MODEL):
    """Extract using OpenAI GPT with file upload."""
    prompt_text = build_prompt(syndicate_num, report_year)
    # Include content_hash in cache key so that changes to the slim PDF
    # (e.g. different page set after off-by-one fix) invalidate the cache.
    cache_key = _llm_cache_key(model, prompt_text + "|" + content_hash, syndicate_num, report_year)
    cached, hit = _llm_cache_load(cache_key)
    if hit:
        print(f"  [{model}] Cache hit — skipping API call")
        return cached["data"]

    print(f"  [{model}] Sending {report_path.name}...")
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

    file_b64 = base64.b64encode(file_bytes).decode("utf-8")
    suffix = report_path.suffix.lower()
    if suffix == ".pdf":
        mime = "application/pdf"
    else:
        mime = "text/html"

    response = client.responses.create(
        model=model,
        max_output_tokens=16384,
        input=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_file",
                        "filename": report_path.name,
                        "file_data": f"data:{mime};base64,{file_b64}",
                    },
                    {
                        "type": "input_text",
                        "text": build_prompt(syndicate_num, report_year),
                    },
                ],
            }
        ],
    )

    # Check for truncation and retry once with larger limit
    try:
        data = sanitize_ascii(parse_json_response(response.output_text))
    except json.JSONDecodeError as e:
        if "Unterminated" in str(e) or "Expecting" in str(e):
            print(f"  [{model}] Response truncated ({response.usage.output_tokens} tokens), retrying with larger limit...")
            response = client.responses.create(
                model=model,
                max_output_tokens=32768,
                input=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "input_file",
                                "filename": report_path.name,
                                "file_data": f"data:{mime};base64,{file_b64}",
                            },
                            {
                                "type": "input_text",
                                "text": build_prompt(syndicate_num, report_year),
                            },
                        ],
                    }
                ],
            )
            data = sanitize_ascii(parse_json_response(response.output_text))
        else:
            raise
    data = add_metadata(data, model, report_path, content_hash)

    input_tokens = response.usage.input_tokens
    output_tokens = response.usage.output_tokens
    total_tokens = input_tokens + output_tokens

    prices = PRICING[model]
    cost = input_tokens * prices["input"] / 1_000_000 + output_tokens * prices["output"] / 1_000_000

    data["_extraction_meta"] = {
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "cost_usd": round(cost, 6),
    }

    print(f"  [{model}] Done. {input_tokens:,} in / {output_tokens:,} out. ${cost:.6f}")
    _llm_cache_save(cache_key, data, model=model,
                    syndicate_num=syndicate_num, report_year=report_year)
    return data


# ---------------------------------------------------------------------------
# Page-level extraction — OCR + targeted page search (RAG-lite)
# ---------------------------------------------------------------------------
# Adobe PDF Extract — deterministic table extraction
# ---------------------------------------------------------------------------

ADOBE_OUTPUT_DIR = Path("pdf_extraction/adobe_output")


def _clean_adobe_cell(val):
    """Clean an Adobe-extracted cell value: strip _x000D_, commas, whitespace.

    Handles accounting-style parenthesized negatives: (26.2) → -26.2
    """
    if val is None:
        return None
    s = str(val).replace("_x000D_", "").replace("\r", "").strip()
    if not s:
        return None
    # Remove thousands commas and spaces
    s_clean = s.replace(",", "").replace(" ", "")
    # Handle accounting-style parenthesized negatives: (123.4) → -123.4
    m = re.match(r'^\(([0-9.]+)\)$', s_clean)
    if m:
        try:
            return -float(m.group(1))
        except ValueError:
            pass
    try:
        return float(s_clean)
    except ValueError:
        return s  # Return as string if not numeric


def adobe_extract_pdf(pdf_path, output_dir=ADOBE_OUTPUT_DIR):
    """Run Adobe PDF Extract API on a PDF. Returns path to output directory.

    Caches results — if output already exists, skips extraction.
    Full Adobe output (JSON, tables, figures) is retained as audit trail.
    """
    report_name = pdf_path.stem
    report_out = output_dir / report_name

    # Check cache
    if (report_out / "structuredData.json").exists():
        return report_out

    if not HAS_ADOBE:
        return None

    client_id = os.getenv("PDF_SERVICES_CLIENT_ID")
    client_secret = os.getenv("PDF_SERVICES_CLIENT_SECRET")
    if not client_id or not client_secret:
        return None

    try:
        credentials = ServicePrincipalCredentials(
            client_id=client_id, client_secret=client_secret,
        )
        pdf_services = PDFServices(credentials=credentials)

        with open(pdf_path, "rb") as f:
            input_stream = f.read()

        input_asset = pdf_services.upload(
            input_stream=input_stream, mime_type=PDFServicesMediaType.PDF,
        )

        extract_params = ExtractPDFParams(
            elements_to_extract=[ExtractElementType.TEXT, ExtractElementType.TABLES],
            elements_to_extract_renditions=[
                ExtractRenditionsElementType.TABLES,
                ExtractRenditionsElementType.FIGURES,
            ],
            table_structure_type=TableStructureType.XLSX,
        )

        extract_job = ExtractPDFJob(
            input_asset=input_asset, extract_pdf_params=extract_params,
        )
        location = pdf_services.submit(extract_job)
        response = pdf_services.get_job_result(location, ExtractPDFResult)
        result_asset: CloudAsset = response.get_result().get_resource()
        stream_asset: StreamAsset = pdf_services.get_content(result_asset)

        report_out.mkdir(parents=True, exist_ok=True)
        zip_path = report_out / "extractResult.zip"
        with open(zip_path, "wb") as f:
            f.write(stream_asset.get_input_stream())

        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(report_out)
        zip_path.unlink()

        return report_out

    except Exception as e:
        err_str = str(e)
        if "QUOTA_EXCEEDED" in err_str or "quota" in err_str.lower():
            print(f"  [Adobe] QUOTA EXCEEDED — stopping script.")
            print(f"  Re-run later to continue (cached outputs will be reused).")
            sys.exit(1)
        print(f"  [Adobe] Extract failed: {e}")
        return None


def find_triangle_in_adobe_output(adobe_dir, report_year):
    """Search Adobe output for the GROSS claims development triangle.

    Reads structuredData.json to find table elements mentioning 'claims development',
    then parses the corresponding xlsx files to find the triangle.

    Returns (triangle_dict, details_str) or (None, reason).
    """
    if not adobe_dir or not HAS_OPENPYXL:
        return None, "no Adobe output or openpyxl not installed"

    json_path = adobe_dir / "structuredData.json"
    if not json_path.exists():
        return None, "no structuredData.json"

    with open(json_path, "r") as f:
        data = json.load(f)

    elements = data.get("elements", [])

    # Find pages that mention claims development
    triangle_pages = set()
    for e in elements:
        text = (e.get("Text") or "").lower()
        if ("claims development" in text or "cumulative claims" in text
                or "analysis of claims" in text):
            page = e.get("Page")
            if page is not None:
                triangle_pages.add(page)

    if not triangle_pages:
        return None, "no claims development text found in Adobe output"

    # Find table elements on those pages with xlsx files
    candidate_xlsx = []
    for e in elements:
        path = e.get("Path", "")
        if "/Table" in path and "/TR" not in path and "/TD" not in path and "/TH" not in path:
            page = e.get("Page")
            if page in triangle_pages:
                fps = e.get("filePaths", [])
                xlsx_files = [fp for fp in fps if fp.endswith(".xlsx")]
                if xlsx_files:
                    candidate_xlsx.append((page, xlsx_files[0]))

    if not candidate_xlsx:
        return None, "no table xlsx files on claims development pages"

    # Try each candidate xlsx — look for triangle structure
    new_syndicate_details = None
    for page, xlsx_rel in candidate_xlsx:
        xlsx_path = adobe_dir / xlsx_rel
        if not xlsx_path.exists():
            continue

        result = _parse_triangle_xlsx(xlsx_path, report_year)
        if result is None:
            continue
        tri_data, details = result
        if tri_data == "new_syndicate":
            # Triangle exists but too few UW years — remember this
            new_syndicate_details = details
            continue
        return tri_data, details

    if new_syndicate_details:
        return "new_syndicate", new_syndicate_details

    return None, "no valid triangle found in candidate xlsx files"


def _parse_triangle_xlsx(xlsx_path, report_year):
    """Parse a single xlsx file and check if it's a claims development triangle.

    Returns (triangle_dict, details_str) or None if not a triangle.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_raw) < 3:
        return None

    # Parse header row — look for underwriting years
    header = [_clean_adobe_cell(c) for c in rows_raw[0]]
    uw_years = []
    uw_col_indices = []
    for i, val in enumerate(header):
        if val is None:
            continue
        # Try to parse as year
        try:
            year = int(float(val)) if isinstance(val, (int, float)) else None
            if year is None:
                # Try extracting year from string like "2011 "
                m = re.search(r'\b(19|20)\d{2}\b', str(val))
                if m:
                    year = int(m.group())
        except (ValueError, TypeError):
            m = re.search(r'\b(19|20)\d{2}\b', str(val))
            year = int(m.group()) if m else None

        if year and 1990 <= year <= 2030:
            # Skip "X and prior" columns
            val_str = str(header[i]) if not isinstance(header[i], (int, float)) else ""
            if "prior" in val_str.lower() or "&" in val_str:
                continue
            uw_years.append(year)
            uw_col_indices.append(i)

    if len(uw_years) < 3:
        # Check if this looks like a new syndicate triangle (has years but too few)
        if len(uw_years) >= 1 and max(uw_years) == report_year:
            return "new_syndicate", f"{len(uw_years)} UW year(s) ({min(uw_years)}-{report_year})"
        return None

    # Max UW year must be recent but need not equal report year (run-off syndicates)
    if max(uw_years) > report_year or max(uw_years) < report_year - 2:
        return None

    # Parse development rows — only keep rows with development period labels
    # (e.g. "At end of underwriting year", "One year later", "Two years later")
    # Skip: summary rows, payments, provisions, blank rows, year labels
    skip_labels = [
        "current estimate", "cumulative payment", "cumulative claim",
        "outstanding", "provision", "gross outstanding", "net outstanding",
    ]
    # Only keep rows whose label looks like a development period
    dev_period_patterns = [
        r"at\s+end",           # "At end of underwriting year"
        r"at\s+the\s+end",     # "At the end of underwriting year"
        r"end\s+of\s+underwriting",  # "End of underwriting year"
        r"year\s+later",       # "One year later", "Two years later"
        r"years?\s+later",     # variant
        r"^\d+\s+year",        # "1 year later"
        r"^(one|two|three|four|five|six|seven|eight|nine|ten)\b",  # word numbers
        r"after\s+\w+\s+years?",               # "After one year", "After two years"
        r"\d+\s+months?\s+later",              # "12 months later", "24 months later"
        r"estimate.*end\s+of\s+underwriting",  # "Estimate of cumulative...end of underwriting year"
    ]
    dev_rows = []
    for row_raw in rows_raw[1:]:
        label = _clean_adobe_cell(row_raw[0])
        if label is None:
            continue
        label_lower = str(label).lower().strip()

        # Skip obvious non-development rows
        if any(s in label_lower for s in skip_labels):
            continue

        # Only include rows that look like development periods
        is_dev_row = any(re.search(p, label_lower) for p in dev_period_patterns)
        if not is_dev_row:
            continue

        # Extract values for UW year columns
        values = []
        for col_idx in uw_col_indices:
            if col_idx < len(row_raw):
                val = _clean_adobe_cell(row_raw[col_idx])
                values.append(val if isinstance(val, (int, float)) else None)
            else:
                values.append(None)
        dev_rows.append(values)

    if len(dev_rows) < 2:
        return None

    # Detect currency and units from header row(s)
    # Scan ALL cells in the header row AND a possible second row (some tables
    # put units in a separate row beneath the year headers)
    scan_rows = rows_raw[:min(3, len(rows_raw))]
    all_header_text = " ".join(
        str(cell or "").replace("_x000D_", "").replace("\r", "")
        for row in scan_rows for cell in row
    )
    all_header_lower = all_header_text.lower()

    currency = "USD"
    if "gbp" in all_header_lower or chr(163) in all_header_text:
        currency = "GBP"
    elif "eur" in all_header_lower or chr(8364) in all_header_text:
        currency = "EUR"

    # Detect units: $'000, £000, $m, £'000, etc.
    # Note: reports use both plain apostrophe (') and right single quote (\u2019)
    units = "millions"
    if (re.search(r"[£$\u00a3]['\u2018\u2019]?000", all_header_text)
            or "'000" in all_header_lower or "\u2019000" in all_header_text):
        units = "thousands"
    elif re.search(r"[£$\u00a3]m\b", all_header_lower):
        units = "millions"
    # else: "millions" is default — auto-detect in compute_pyd_from_triangle
    # will catch full-integer tables from value magnitudes

    # Build triangle dict
    tri_data = {
        "type": "gross",
        "currency": currency,
        "units": units,
        "underwriting_years": uw_years,
        "development_rows": dev_rows,
    }

    return tri_data, f"Adobe xlsx triangle: {len(uw_years)} UW years, {len(dev_rows)} dev rows"


def find_lob_in_adobe_output(adobe_dir, report_year):
    """Search Adobe output for the segmental analysis / LOB breakdown table.

    Looks for tables containing "analysis of underwriting result" or
    "segmental analysis" with the current report year.  Extracts gross
    premiums written per LOB and gross claims incurred per LOB.

    Returns dict with keys:
        gross_premium_mix: list of {line_of_business, amount_gbp_m, percentage_of_total}
        gross_premiums_written_gbp_m: total GWP
        claims_incurred_by_lob: list of {line_of_business, amount_gbp_m}
        currency: detected currency
        method: "adobe"
    or None if not found.
    """
    if not adobe_dir or not HAS_OPENPYXL:
        return None

    json_path = adobe_dir / "structuredData.json"
    if not json_path.exists():
        return None

    with open(json_path, "r") as f:
        data = json.load(f)

    elements = data.get("elements", [])

    # Find pages mentioning segmental analysis or underwriting result
    lob_pages = set()
    for e in elements:
        text = (e.get("Text") or "").lower()
        if any(kw in text for kw in [
            "analysis of underwriting result",
            "segmental analysis",
            "class of business",
            "analysis of net premiums",
        ]):
            page = e.get("Page")
            if page is not None:
                lob_pages.add(page)

    if not lob_pages:
        return None

    # Find table xlsx files on those pages
    candidate_xlsx = []
    for e in elements:
        path = e.get("Path", "")
        if "/Table" in path and "/TR" not in path and "/TD" not in path and "/TH" not in path:
            page = e.get("Page")
            if page in lob_pages:
                fps = e.get("filePaths", [])
                xlsx_files = [fp for fp in fps if fp.endswith(".xlsx")]
                if xlsx_files:
                    candidate_xlsx.append((page, xlsx_files[0]))

    if not candidate_xlsx:
        return None

    # Try each candidate — pick the one with the most LOB rows (prefer full segmental)
    best_result = None
    best_lob_count = 0
    for page, xlsx_rel in candidate_xlsx:
        xlsx_path = adobe_dir / xlsx_rel
        if not xlsx_path.exists():
            continue

        result = _parse_lob_xlsx(xlsx_path, report_year)
        if result is not None:
            n_lobs = len(result["gross_premium_mix"])
            if n_lobs > best_lob_count:
                best_lob_count = n_lobs
                best_result = result

    return best_result


def _parse_lob_xlsx(xlsx_path, report_year):
    """Parse a segmental analysis xlsx for LOB breakdown.

    Handles format variations across syndicates:
    - Different column headers: "Gross written premiums", "Gross Premiums Written",
      "Ceded Balance" vs "Reinsurance Balance", etc.
    - Different currencies: GBP £000, USD $000, USD full integers, EUR
    - Different units: thousands ($000, £000, £'000) vs millions ($m) vs full integers
    - Different LOB names: statutory classes with syndicate-specific sub-divisions
    - Section headers like "Direct Insurance:" with no numeric values
    - Comparative year tables on the same page

    Returns dict or None.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_raw) < 3:
        return None

    # --- Step 1: Detect currency and units from header/title text ---
    header_text = " ".join(
        str(c or "").replace("_x000D_", "").replace("\r", "")
        for row in rows_raw[:2] for c in row
    ).lower()

    currency = "GBP"  # default
    if "us$" in header_text or "usd" in header_text or "$" in header_text:
        currency = "USD"
    elif "eur" in header_text or chr(8364) in header_text:
        currency = "EUR"

    # Detect units: thousands, millions, or full integers
    # Look for explicit unit markers: £'000, $000, £000, $m, £m, etc.
    # Note: Adobe xlsx uses right single quote (\u2019) not plain apostrophe
    units_divisor = 1.0  # default: assume millions
    if (re.search(r"[£$\u00a3]['\u2018\u2019]?000", header_text)
            or "'000" in header_text or "\u2019000" in header_text):
        units_divisor = 1000.0  # values are in thousands → divide by 1000 for millions
    elif "$m" in header_text or "£m" in header_text:
        units_divisor = 1.0  # already in millions
    # else: could be full integers — detect later from magnitude

    # --- Step 2: Fuzzy column matching from header row ---
    header = [_clean_adobe_cell(c) for c in rows_raw[0]]

    # Match columns by keyword patterns (order of preference)
    gwp_patterns = [
        lambda s: "written" in s and "premium" in s,       # "Gross written premiums"
        lambda s: "written" in s and "gross" in s,          # "Gross Written Premiums"
        lambda s: "gross" in s and "premium" in s and "earned" not in s,  # "Gross Premiums" but not earned
    ]
    earned_patterns = [
        lambda s: "earned" in s and "premium" in s,         # "Gross premiums earned"
    ]
    claims_patterns = [
        lambda s: "claim" in s and ("gross" in s or "incurred" in s),  # "Gross claims incurred"
        lambda s: "claim" in s,                              # "Claims incurred"
    ]

    def _find_col(patterns, header_vals):
        """Find column index matching first successful pattern."""
        for pat in patterns:
            for i, val in enumerate(header_vals):
                if val is not None and pat(str(val).lower()):
                    return i
        return None

    gwp_col = _find_col(gwp_patterns, header)
    earned_col = _find_col(earned_patterns, header)
    claims_col = _find_col(claims_patterns, header)

    # Prefer written premiums; fall back to earned
    if gwp_col is None:
        gwp_col = earned_col

    # --- Step 3: Fallback — if no header keywords matched, try positional ---
    if gwp_col is None and len(header) >= 5:
        lob_keywords = [
            "accident", "motor", "marine", "fire", "property", "liability",
            "miscellaneous", "reinsurance", "energy", "casualty", "aviation",
            "pecuniary", "credit", "suretyship", "specialty", "transport",
        ]
        has_lob_rows = any(
            any(kw in str(row[0] or "").lower() for kw in lob_keywords)
            for row in rows_raw[1:8]
        )
        if has_lob_rows:
            gwp_col = 1
            claims_col = 3 if len(header) >= 4 else None

    if gwp_col is None:
        return None

    # --- Step 4: Check this table is for the report year (not a prior-year comparative) ---
    title_text = " ".join(str(c or "") for c in rows_raw[0])
    # If the title mentions a DIFFERENT year, skip this table
    for yr in range(report_year - 10, report_year + 2):
        if str(yr) in title_text:
            if yr != report_year:
                return None  # This is a comparative table for a different year
            break  # Confirmed it's the report year

    # --- Step 5: Parse LOB rows ---
    # Labels to skip: section headers (no values) and subtotals/totals
    section_headers_lower = {
        "direct insurance", "direct insurance:", "direct",
        "reinsurance acceptances", "reinsurance acceptances:",
    }
    total_labels_lower = {"total", "sub-total", "subtotal", "grand total"}

    lob_entries = []
    claims_entries = []
    total_gwp = 0.0

    for row in rows_raw[1:]:
        label = _clean_adobe_cell(row[0])
        if label is None:
            continue
        label_str = str(label).strip()
        label_lower = label_str.lower().rstrip(":")

        # Skip section headers (rows with a label but no numeric values)
        if label_lower in section_headers_lower:
            continue

        # Detect subtotal rows ("Direct Insurance" subtotal with values)
        is_subtotal = label_lower in total_labels_lower or (
            label_lower.startswith("direct insurance") and not label_lower.startswith("direct insurance:")
        )

        if is_subtotal:
            # Don't add as LOB, but capture total GWP from "Total" row
            if "total" in label_lower and gwp_col < len(row):
                val = _clean_adobe_cell(row[gwp_col])
                if isinstance(val, (int, float)):
                    total_gwp = abs(val)
            continue

        # Get GWP value
        gwp_val = None
        if gwp_col < len(row):
            val = _clean_adobe_cell(row[gwp_col])
            if isinstance(val, (int, float)):
                gwp_val = abs(val)

        # Get claims value
        claims_val = None
        if claims_col is not None and claims_col < len(row):
            val = _clean_adobe_cell(row[claims_col])
            if isinstance(val, (int, float)):
                claims_val = val

        if gwp_val is not None and gwp_val > 0:
            lob_entries.append({
                "line_of_business": label_str,
                "amount_raw": gwp_val,
            })
        elif claims_val is not None:
            # LOBs with near-zero GWP but active claims (e.g. Motor run-off)
            lob_entries.append({
                "line_of_business": label_str,
                "amount_raw": 0.0,
            })

        if claims_val is not None:
            claims_entries.append({
                "line_of_business": label_str,
                "amount_raw": claims_val,
            })

    if not lob_entries:
        return None

    # --- Step 6: Auto-detect units if not determined from header ---
    if total_gwp == 0 and lob_entries:
        total_gwp = sum(e["amount_raw"] for e in lob_entries)

    # No Lloyd's syndicate has GWP > £10bn, so magnitude tells us the units:
    #   > 10,000,000 → full currency units (divide by 1M)
    #   > 10,000     → thousands (divide by 1K)
    #   ≤ 10,000     → already millions
    if units_divisor == 1.0 and total_gwp > 10_000_000:
        units_divisor = 1_000_000.0  # full integers → millions
    elif units_divisor == 1.0 and total_gwp > 10_000:
        units_divisor = 1_000.0  # thousands → millions

    # Apply units conversion and compute percentages
    total_gwp_m = round(total_gwp / units_divisor, 1) if units_divisor != 1.0 else round(total_gwp, 1)

    for e in lob_entries:
        raw = e.pop("amount_raw")
        e["amount_gbp_m"] = round(raw / units_divisor, 1) if units_divisor != 1.0 else round(raw, 1)
        e["percentage_of_total"] = round(e["amount_gbp_m"] / total_gwp_m * 100, 1) if total_gwp_m > 0 else 0

    for e in claims_entries:
        raw = e.pop("amount_raw")
        e["amount_gbp_m"] = round(raw / units_divisor, 1) if units_divisor != 1.0 else round(raw, 1)

    return {
        "gross_premium_mix": lob_entries,
        "gross_premiums_written_gbp_m": total_gwp_m,
        "claims_incurred_by_lob": claims_entries if claims_entries else None,
        "currency": currency,
        "method": "adobe",
    }


def find_provisions_movement_in_adobe(adobe_dir, report_year):
    """Search Adobe output for the 'movement in claims provisions' table.

    Extracts gross prior year development from the note that shows:
    - Claims: prior underwriting years (gross / RI share / net)

    Returns dict with keys:
        gross_prior_year_claims: float (gross claims on prior UW years)
        ri_share_prior_year: float (reinsurer's share)
        net_prior_year_claims: float (net = gross - RI)
    or None if not found.
    """
    if not adobe_dir or not HAS_OPENPYXL:
        return None

    json_path = adobe_dir / "structuredData.json"
    if not json_path.exists():
        return None

    with open(json_path, "r") as f:
        data = json.load(f)

    elements = data.get("elements", [])

    # Find pages mentioning claims provisions movement
    provision_pages = set()
    for e in elements:
        text = (e.get("Text") or "").lower()
        if any(kw in text for kw in [
            "movement in claims",
            "movement in provision",
            "claims outstanding",
            "provision for claims",
        ]):
            page = e.get("Page")
            if page is not None:
                provision_pages.add(page)

    if not provision_pages:
        return None

    # Find table xlsx files on those pages
    candidate_xlsx = []
    for e in elements:
        path = e.get("Path", "")
        if "/Table" in path and "/TR" not in path and "/TD" not in path and "/TH" not in path:
            page = e.get("Page")
            if page in provision_pages:
                fps = e.get("filePaths", [])
                xlsx_files = [fp for fp in fps if fp.endswith(".xlsx")]
                if xlsx_files:
                    candidate_xlsx.append((page, xlsx_files[0]))

    for page, xlsx_rel in candidate_xlsx:
        xlsx_path = adobe_dir / xlsx_rel
        if not xlsx_path.exists():
            continue

        result = _parse_provisions_xlsx(xlsx_path, report_year)
        if result is not None:
            return result

    return None


def _parse_provisions_xlsx(xlsx_path, report_year):
    """Parse a claims provisions movement xlsx.

    Looks for a row containing 'prior' (prior underwriting years claims).
    Handles varying column headers and unit conventions across syndicates.
    Returns dict or None.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_raw) < 3:
        return None

    # Detect units from header text (scan first 2 rows — some tables split title/units)
    header_text = " ".join(
        str(c or "").replace("_x000D_", "").replace("\r", "")
        for row in rows_raw[:min(2, len(rows_raw))] for c in row
    )
    units_divisor = 1.0
    if (re.search(r"[£$\u00a3]['\u2018\u2019]?000", header_text)
            or "'000" in header_text or "\u2019000" in header_text):
        units_divisor = 1000.0
    # else assume millions or detect from magnitude later

    # Detect column layout from header using fuzzy matching
    header = rows_raw[0] if rows_raw else []
    gross_col = ri_col = net_col = None
    for i, cell in enumerate(header):
        h = str(cell or "").lower()
        if "gross" in h or "insurance liabilities" in h:
            gross_col = i
        elif "reinsur" in h or "share" in h or "ceded" in h:
            ri_col = i
        elif "net" in h:
            net_col = i

    # Fallback: assume positional (col1=gross, col2=RI, col3=net)
    if gross_col is None and len(header) >= 4:
        gross_col, ri_col, net_col = 1, 2, 3

    # Look for rows mentioning "prior" year claims
    for row in rows_raw:
        label = str(row[0] or "").lower()
        if "prior" in label and ("claim" in label or "underwriting" in label or "year" in label):
            result = {}
            raw_vals = []
            for key, col in [
                ("gross_prior_year_claims", gross_col),
                ("ri_share_prior_year", ri_col),
                ("net_prior_year_claims", net_col),
            ]:
                if col is not None and col < len(row):
                    val = _clean_adobe_cell(row[col])
                    if isinstance(val, (int, float)):
                        raw_vals.append(abs(val))
                        result[key] = val

            if not result:
                continue

            # Auto-detect units from value magnitudes if header was ambiguous
            effective_divisor = units_divisor
            if effective_divisor == 1.0 and raw_vals:
                max_val = max(raw_vals)
                if max_val > 1_000_000:
                    effective_divisor = 1_000_000.0
                elif max_val > 10_000:
                    effective_divisor = 1_000.0

            for key in result:
                result[key] = round(result[key] / effective_divisor, 1) if effective_divisor != 1.0 else result[key]

            return result

    return None


# ---------------------------------------------------------------------------
# Page-level extraction — OCR + targeted page search (RAG-lite)
# ---------------------------------------------------------------------------

# Patterns that indicate a claims development triangle page
TRIANGLE_PATTERNS = [
    r"claims?\s+development",
    r"analysis\s+of\s+claims",
    r"cumulative\s+claims?\s+incurred",
    r"underwriting\s+year.*later",
    r"end\s+of\s+underwriting\s+year",
    r"at\s+end\s+of\s+.*year",
    r"one\s+year\s+later",
    r"two\s+years?\s+later",
    r"three\s+years?\s+later",
    # Loss ratio triangle format (e.g. Beazley 2623)
    r"\b12\s+months\b",
    r"\b24\s+months\b",
    r"total\s+ultimate\s+loss",
    r"gross\s+claims\s+liabilities",
]

# Patterns that indicate prior year reserve movement text
RESERVE_MOVEMENT_PATTERNS = [
    r"prior\s+year[\u2019\u2018']?s?\s+(reserve|claim|development|movement|provision|business)",
    r"movement\s+in\s+prior\s+year",
    r"prior\s+underwriting\s+year",
    r"prior\s+years?\s+of\s+account",              # Lloyd's YOA: "2012 & prior years of account"
    r"reserve\s+(release|strengthen|deteriorat|surplus|deficit)",
    r"run.?off\s+(surplus|deficit|deviation|result|improvement|deterioration|release|strengthening)",
    r"(favourable|favorable|profitable).*development",
    r"adverse.*development",
    r"prior\s+year.*release",
    r"prior\s+year.*strengthen",
    r"released?\s+prior\s+year\s+reserve",   # "released prior year reserves"
    r"release\s+of\s+.{0,30}prior\s+year",    # "release of [£4.7m of] prior year reserves"
    r"claims\s+incurred\s+in\s+prior",       # provisions movement note
    r"movement\s+in\s+(claims|provision)",    # provisions movement heading
    r"relating\s+to\s+prior\s+(year|underwriting)",  # "relating to prior year's business"
    r"(improvement|deterioration)\s+(for|in|of|on|relating)",  # run-off improvement/deterioration
    r"(better|worse)\s+than\s+expected\s+(claims|loss|experience)",  # expectation-based language
    r"\d{4}.{0,20}year.{0,5}of.{0,5}account.{0,80}(profit|loss|surplus|deficit|develop)",  # "2012 year of account closed with a profit"
    r"(closed|closing)\s+year.{0,50}(profit|loss|surplus|deficit|result|of\s+account)",     # "closed year result"
    r"prior\s+accident\s+year",                                                             # "prior accident years" (common in older reports)
    r"in\s+respect\s+of\s+prior\s+(?:accident\s+|underwriting\s+)?year",                    # "in respect of prior [accident/underwriting] years"
    r"releases?\s+in\s+respect\s+of",                                                       # "releases in respect of ..." (prior year context)
]


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF, page by page. Returns list of (page_num, text).

    Tries PyMuPDF first, then pdfplumber, then OCR via Tesseract.
    """
    pages = []

    # Try PyMuPDF
    if fitz:
        try:
            doc = fitz.open(str(pdf_path))
            for i, page in enumerate(doc):
                text = page.get_text()
                pages.append((i + 1, text.strip()))
            doc.close()
            # Check if we got meaningful text — need at least 100 chars/page average
            total_chars = sum(len(t) for _, t in pages)
            if len(pages) > 0 and total_chars / len(pages) > 100:
                return pages, "pymupdf"
        except Exception as e:
            logger.debug(f"PyMuPDF extraction failed for {pdf_path.name}: {e}")

    # Try pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages = []
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                pages.append((i + 1, text.strip()))
            total_chars = sum(len(t) for _, t in pages)
            if len(pages) > 0 and total_chars / len(pages) > 100:
                return pages, "pdfplumber"
    except Exception as e:
        logger.debug(f"pdfplumber extraction failed for {pdf_path.name}: {e}")

    # OCR with Tesseract — cache results to avoid re-running
    if HAS_PDF2IMAGE and HAS_TESSERACT:
        ocr_cache_path = Path("pdf_extraction/ocr_page_cache")
        ocr_cache_path.mkdir(parents=True, exist_ok=True)
        cache_file = ocr_cache_path / f"{pdf_path.stem}.json"

        if cache_file.exists():
            try:
                with open(cache_file, "r") as f:
                    cached = json.load(f)
                pages = [(p["page"], p["text"]) for p in cached]
                return pages, "ocr_cache"
            except Exception as e:
                logger.debug(f"OCR cache read failed for {pdf_path.name}: {e}")

        try:
            images = convert_from_path(str(pdf_path), dpi=200)
            pages = []
            for i, img in enumerate(images):
                text = pytesseract.image_to_string(img)
                pages.append((i + 1, text.strip()))
            # Cache the OCR results
            with open(cache_file, "w") as f:
                json.dump([{"page": p, "text": t} for p, t in pages], f)
            return pages, "ocr"
        except Exception as e:
            print(f"  [OCR] Failed: {e}")

    return pages, "none"


def find_relevant_pages(pages, report_year):
    """Search page text for triangle tables and reserve movement narratives.

    Returns dict with keys:
        'triangle_pages': list of (page_num, text) for pages with triangle data
        'reserve_pages': list of (page_num, text) for pages with reserve movement text
    """
    triangle_pages = []
    reserve_pages = []

    for page_num, text in pages:
        if not text:
            continue
        text_lower = text.lower()

        # Check for triangle
        triangle_score = sum(1 for p in TRIANGLE_PATTERNS
                           if re.search(p, text_lower))
        if triangle_score >= 2:
            triangle_pages.append((page_num, text))

        # Check for reserve movement narrative
        reserve_score = sum(1 for p in RESERVE_MOVEMENT_PATTERNS
                          if re.search(p, text_lower))
        if reserve_score >= 2:
            reserve_pages.append((page_num, text))

    return {
        "triangle_pages": triangle_pages,
        "reserve_pages": reserve_pages,
    }


def render_page_as_image_b64(pdf_path, page_num):
    """Render a single PDF page as a base64-encoded PNG image."""
    if not HAS_PDF2IMAGE:
        return None
    try:
        images = convert_from_path(str(pdf_path), dpi=250,
                                   first_page=page_num, last_page=page_num)
        if images:
            buf = io.BytesIO()
            images[0].save(buf, format="PNG")
            return base64.b64encode(buf.getvalue()).decode("utf-8")
    except Exception as e:
        print(f"  [render] Failed to render page {page_num}: {e}")
    return None


TRIANGLE_EXTRACT_PROMPT = """Extract the claims development triangle from this page.

This is a table showing cumulative gross incurred claims by underwriting year.
Return ONLY valid JSON with this structure:

{{
  "type": "<gross|net|loss_ratio|none>",
  "currency": "<GBP|USD|EUR>",
  "units": "<millions|thousands|percentage>",
  "underwriting_years": [<list of individual UW year integers, e.g. [2011, 2012, 2013, ...]. EXCLUDE any 'X and prior' aggregate column>],
  "development_rows": [
    [<Row 0: values for each UW year, null if cell is blank>],
    [<Row 1: next development period>],
    ...
  ]
}}

CRITICAL RULES:
- EXCLUDE any 'X and prior' aggregate column (e.g. '2010 and prior')
- EXCLUDE 'Current estimate of cumulative claims' summary row
- EXCLUDE 'Cumulative payments' and 'Outstanding claims provision' rows
- Include ONLY development period rows (At end of UW year, One year later, Two years later, etc.)
- Use null for empty cells
- The most recent UW year column should be {report_year}
- The oldest UW year should have the most filled development rows
"""


def extract_triangle_from_page(pdf_path, page_num, report_year, model="gemini-2.5-flash"):
    """Send a single page image to an LLM to extract the triangle table.

    Returns (triangle_dict, cost) or (None, 0) on failure.
    """
    prompt = TRIANGLE_EXTRACT_PROMPT.replace("{report_year}", str(report_year))

    # Parse syndicate number from filename for cache key
    _parts = Path(pdf_path).stem.split("_")
    _syn = int(_parts[1]) if len(_parts) >= 3 else 0
    cache_key = _llm_cache_key(model, prompt, _syn, report_year, page_num=page_num)
    cached, hit = _llm_cache_load(cache_key)
    if hit:
        return cached["data"], 0  # cost=0 for cached results

    img_b64 = render_page_as_image_b64(pdf_path, page_num)
    if not img_b64:
        return None, 0

    if model.startswith("gemini"):
        client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))
        config = GenerateContentConfig(
            temperature=0.0,
            response_mime_type="application/json",
        )
        # Send image as inline data
        image_bytes = base64.b64decode(img_b64)
        response = _gemini_call_with_retry(lambda: client.models.generate_content(
            model=model,
            contents=[
                {"inline_data": {"mime_type": "image/png",
                                 "data": image_bytes}},
                prompt,
            ],
            config=config,
        ))
        try:
            data = parse_json_response(response.text)
            usage = response.usage_metadata
            cost = (usage.prompt_token_count * PRICING[model]["input"] / 1_000_000 +
                    usage.candidates_token_count * PRICING[model]["output"] / 1_000_000)
            _llm_cache_save(cache_key, data, model=model,
                            syndicate_num=_syn, report_year=report_year,
                            page_num=page_num)
            return data, cost
        except Exception as e:
            logger.warning(f"Failed to parse Gemini vision response: {e}")
            return None, 0
    else:
        # OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        response = client.responses.create(
            model=model,
            max_output_tokens=8192,
            input=[{
                "role": "user",
                "content": [
                    {"type": "input_image",
                     "image_url": f"data:image/png;base64,{img_b64}"},
                    {"type": "input_text", "text": prompt},
                ],
            }],
        )
        try:
            data = parse_json_response(response.output_text)
            cost = (response.usage.input_tokens * PRICING[model]["input"] / 1_000_000 +
                    response.usage.output_tokens * PRICING[model]["output"] / 1_000_000)
            _llm_cache_save(cache_key, data, model=model,
                            syndicate_num=_syn, report_year=report_year,
                            page_num=page_num)
            return data, cost
        except Exception as e:
            logger.warning(f"Failed to parse GPT vision response: {e}")
            return None, 0


def extract_triangle_with_retry(pdf_path, page_num, report_year, max_retries=3):
    """Extract triangle from a page, validating and retrying up to max_retries times.

    Uses Gemini for extraction (cheapest vision model). Validates structure after each attempt.

    Returns (triangle_dict, computed_pyd, details, total_cost) or (None, None, None, total_cost).
    """
    total_cost = 0
    model = "gemini-2.5-flash"

    last_details = None
    for attempt in range(1, max_retries + 1):
        tri_data, cost = extract_triangle_from_page(pdf_path, page_num, report_year, model)
        total_cost += cost

        if not tri_data:
            last_details = "no triangle data on this page"
            continue

        # Validate and compute PYD
        pyd, details = compute_pyd_from_triangle(tri_data, report_year)
        if pyd is not None:
            print(f"    [page {page_num}] PYD={pyd:+.3f}m OK (attempt {attempt})")
            return tri_data, pyd, details, total_cost

        # New/young syndicate — triangle is structurally correct but too small for PYD
        if details and ("fewer than 2 development rows" in details
                        or "no usable UW years" in details):
            return tri_data, None, "new_syndicate", total_cost

        last_details = details

    # Only print one summary line for pages that didn't yield a triangle
    print(f"    [page {page_num}] No usable triangle ({last_details})")
    return None, None, None, total_cost


def extract_pyd_from_relevant_pages(pdf_path, report_year):
    """RAG-lite: find relevant pages, extract triangle or reserve narrative.

    Priority order:
    1. Adobe PDF Extract (deterministic xlsx tables) — best for scanned PDFs
    2. OCR + LLM vision on individual pages — fallback if Adobe unavailable
    3. Reserve movement text — for reports without triangles

    Returns dict:
        'triangle': computed triangle data (dict) or None
        'pyd': computed PYD from triangle (float) or None
        'pyd_details': details string
        'reserve_text': concatenated reserve movement text from relevant pages
        'method': how triangle was extracted (adobe/ocr_vision/none)
        'cost': total cost of page-level extractions
    """
    result = {
        "triangle": None, "pyd": None, "pyd_details": None,
        "reserve_text": "", "method": "none", "cost": 0,
        "adobe_lob": None, "adobe_provisions": None,
        "first_year_syndicate": False,
        "no_triangle_data": False,
        "relevant_pages": [],
        "rotated_pages": set(),
    }

    # Step 1: Table extraction (deterministic, best quality)
    backend_name = TABLE_BACKEND.value.capitalize()
    extraction = extract_tables(pdf_path, report_year, backend=TABLE_BACKEND,
                                azure_paid=AZURE_PAID)
    result["relevant_pages"] = extraction.relevant_pages
    result["rotated_pages"] = extraction.rotated_pages

    if extraction.triangle:
        tri_data = extraction.triangle.to_dict()
        uw_years = tri_data.get("underwriting_years", [])
        usable_years = [y for y in uw_years if int(y) <= report_year - 2]
        if len(usable_years) == 0:
            oldest = min(uw_years) if uw_years else "?"
            print(f"  [{backend_name}] NEW SYNDICATE: triangle spans {oldest}-{report_year} "
                  f"({len(uw_years)} UW years, {len(usable_years)} usable) "
                  f"-- no prior year development possible")
            result["first_year_syndicate"] = True
        else:
            pyd, pyd_details = compute_pyd_from_triangle(tri_data, report_year)
            if pyd is not None:
                print(f"  [{backend_name}] Triangle PYD: {pyd:+.3f}m ({extraction.triangle_details})")
                result["triangle"] = tri_data
                result["pyd"] = pyd
                result["pyd_details"] = pyd_details
                result["method"] = extraction.method
            elif pyd_details and "no usable UW years" in pyd_details:
                # Triangle is structurally valid but all data-bearing columns
                # are too recent for PYD (e.g. syndicate started 2015 but
                # has nominal UW year columns going back to 2013 with dashes).
                oldest = min(uw_years) if uw_years else "?"
                print(f"  [{backend_name}] NEW SYNDICATE: triangle found "
                      f"({extraction.triangle_details}) but no UW years "
                      f"old enough for PYD -- treating as first-year syndicate")
                result["first_year_syndicate"] = True
                result["triangle"] = tri_data
    elif extraction.first_year_syndicate:
        # Table extraction found a table with < 3 UW years but couldn't find a
        # full triangle.  Don't trust this — the parser may have found a partial
        # or net triangle while the full gross triangle exists elsewhere in the PDF.
        # Fall through to LLM extraction which can read PYD from the narrative.
        print(f"  [{backend_name}] Possible new syndicate ({extraction.triangle_details}) "
              f"-- table extraction inconclusive, will try LLM extraction")
    elif extraction.triangle_details:
        print(f"  [{backend_name}] No triangle found: {extraction.triangle_details}")

    # Step 1b: Extract LOB breakdown from segmental analysis
    if extraction.lob:
        lob_data = extraction.lob.to_dict()
        n_lobs = len(lob_data["gross_premium_mix"])
        total = lob_data["gross_premiums_written_gbp_m"]
        print(f"  [{backend_name}] LOB breakdown: {n_lobs} classes, GWP={total}m ({lob_data['currency']})")
        result["adobe_lob"] = lob_data  # key kept for backwards compatibility

    # Step 1c: Extract claims provisions movement (gross PYD from note)
    if extraction.provisions:
        prov_data = extraction.provisions.to_dict()
        gross = prov_data.get("gross_prior_year_claims")
        net = prov_data.get("net_prior_year_claims")
        opening = prov_data.get("opening_gross_claims_outstanding")
        parts = []
        if gross is not None:
            parts.append(f"gross={gross:+.1f}m")
        if net is not None:
            parts.append(f"net={net:+.1f}m")
        if opening is not None:
            parts.append(f"opening claims={opening:.3f}m")
        if parts:
            print(f"  [{backend_name}] Provisions movement: {', '.join(parts)}")
        result["adobe_provisions"] = prov_data  # key kept for backwards compatibility

    # Cross-validate triangle PYD vs provisions PYD.
    # Triangle diagonal differences measure total development (including normal
    # emergence for immature UW years), while the provisions note measures the
    # actual balance-sheet reserve movement.  When both are available and they
    # disagree in sign, the provisions figure is more reliable for "prior year
    # development" as used in actuarial reserve analysis.
    if result["pyd"] is not None and result["method"] in ("azure", "nutrient", "adobe"):
        prov = result.get("adobe_provisions") or {}
        prov_gross = prov.get("gross_prior_year_claims")
        if prov_gross is not None and prov_gross != 0.0:
            tri_pyd = result["pyd"]
            # Disagree in sign — provisions is authoritative for reserve movement
            if (tri_pyd > 0 and prov_gross < 0) or (tri_pyd < 0 and prov_gross > 0):
                print(f"  [{backend_name}] Triangle PYD ({tri_pyd:+.3f}m) disagrees in sign "
                      f"with provisions ({prov_gross:+.1f}m) — using provisions")
                result["pyd"] = prov_gross
                result["pyd_details"] = (f"provisions overrides triangle (sign disagreement: "
                                         f"triangle={tri_pyd:+.3f}m, provisions={prov_gross:+.1f}m)")
                result["method"] = "provisions"

    # If confirmed first-year syndicate, check for reserves movement fallback
    # before giving up.  Some syndicates (e.g. those with RITC) show a
    # "Movement in Underwriting Reserves" note with opening balance and
    # change-in-year even when the claims triangle has too few UW years.
    # If the reserves movement note isn't found, fall through to the normal
    # text extraction flow — the report may state PYD in narrative text
    # (e.g. "Prior year movements of £5.8m") even when the triangle is small.
    pages = None
    text_method = None
    if result["first_year_syndicate"]:
        pages, text_method = extract_text_from_pdf(pdf_path)
        override = _extract_pyd_from_reserves_movement(pages, report_year) if pages else None
        if override:
            result["first_year_syndicate"] = False
            result["pyd"] = override["pyd"]
            result["pyd_details"] = override["details"]
            result["method"] = "reserves_movement"
            result["opening_reserves_from_movement"] = override["opening_reserves"]
            print(f"  [RAG] Reserves movement note overrides first-year: "
                  f"PYD={override['pyd']:+.3f}m, opening={override['opening_reserves']:.1f}m")

    # Step 2: Extract text for page search (needed for reserve text and LLM fallback)
    if pages is None:
        pages, text_method = extract_text_from_pdf(pdf_path)

    if pages:
        total_chars = sum(len(t) for _, t in pages)
        print(f"  [RAG] {len(pages)} pages via {text_method} ({total_chars:,} chars)")

        relevant = find_relevant_pages(pages, report_year)
        tri_pages = relevant["triangle_pages"]
        res_pages = relevant["reserve_pages"]

        # Merge text-based page numbers into the relevant_pages set.
        # find_relevant_pages returns 1-indexed page numbers (from
        # extract_text_from_pdf), but table_extraction uses 0-indexed
        # page numbers throughout.  Convert to 0-indexed before merging.
        text_page_nums = set(pn - 1 for pn, _ in tri_pages + res_pages)
        existing = set(result["relevant_pages"])
        result["relevant_pages"] = sorted(existing | text_page_nums)

        # Step 3: If Adobe didn't find a triangle, try LLM vision on triangle pages
        if result["pyd"] is None and not result["first_year_syndicate"] and tri_pages and HAS_PDF2IMAGE:
            print(f"  [RAG] {len(tri_pages)} triangle page(s) found, trying LLM vision...")
            for page_num, page_text in tri_pages[:2]:
                if "net" in page_text.lower() and "gross" not in page_text.lower():
                    continue
                tri_data, pyd, details, cost = extract_triangle_with_retry(
                    pdf_path, page_num, report_year
                )
                result["cost"] += cost
                if details == "new_syndicate":
                    uw_years = tri_data.get("underwriting_years", []) if tri_data else []
                    n_uw = len(uw_years)
                    oldest = min(uw_years) if uw_years else report_year
                    print(f"  [RAG] NEW SYNDICATE: triangle has {n_uw} UW year(s) "
                          f"({oldest}-{report_year}) — no prior year development possible")
                    result["first_year_syndicate"] = True
                    result["triangle"] = tri_data
                    break
                if pyd is not None:
                    print(f"  [RAG] LLM vision triangle PYD: {pyd:+.3f}m")
                    result["triangle"] = tri_data
                    result["pyd"] = pyd
                    result["pyd_details"] = details
                    result["method"] = "ocr_vision"
                    break

        # Step 3b: If still no PYD, try loss ratio triangle on triangle pages.
        # The loss ratio grid and the "Total ultimate losses" row may span
        # consecutive pages, so concatenate all triangle page texts.
        # Exclude pages that are purely "Net Claims Development" (no gross).
        if result["pyd"] is None and not result["first_year_syndicate"] and tri_pages:
            gross_tri_pages = []
            for pn, txt in tri_pages:
                txt_lower = txt.lower()
                # Skip pages that have net claims development but no gross
                if ("net claims development" in txt_lower
                        and "gross claims development" not in txt_lower):
                    continue
                # Skip pages that have net ratios but no gross ratios
                if ("net ratios" in txt_lower
                        and "gross ratios" not in txt_lower
                        and "gross claims development" not in txt_lower):
                    continue
                gross_tri_pages.append((pn, txt))
            combined_tri_text = "\n".join(text for _, text in gross_tri_pages)
            lr_pyd, lr_details = _extract_pyd_from_loss_ratio_triangle(
                combined_tri_text, report_year
            )
            if lr_pyd is not None:
                print(f"  [RAG] Loss ratio triangle PYD: {lr_pyd:+.3f}m")
                result["pyd"] = lr_pyd
                result["pyd_details"] = lr_details
                result["method"] = "loss_ratio_triangle"
            elif lr_details and "loss ratio triangle found" in lr_details:
                # A valid loss ratio grid structure was detected but PYD
                # could not be computed (no "Total ultimate losses" row).
                # Still mark that the report has a triangle so that we
                # don't exclude it — LLMs may extract PYD from narrative.
                result["has_loss_ratio_triangle"] = True
                print(f"  [RAG] Loss ratio triangle found but no absolute PYD: {lr_details}")

        # Step 4: Collect reserve movement text
        if res_pages:
            result["reserve_text"] = "\n---\n".join(
                f"[Page {pn}]\n{text}" for pn, text in res_pages
            )
            if result["pyd"] is None:
                print(f"  [RAG] No triangle, but found {len(res_pages)} reserve text page(s)")
            # If this was flagged as first-year but we found reserve movement
            # text, the syndicate likely has prior year movements (e.g. from
            # RITC) even though the claims triangle is small.  Clear the flag
            # so the narrative PYD parsers can attempt extraction.
            if result["first_year_syndicate"] and result["pyd"] is None:
                result["first_year_syndicate"] = False
                print(f"  [RAG] Reserve text found — clearing first-year flag, will try narrative parsers")

    # Step 5: If still no PYD but we have provisions data with gross PYD,
    # use it as a fallback.  The provisions movement note (e.g. "Claims
    # incurred in relation to prior underwriting years") is a reliable
    # source of prior year development when no claims development triangle
    # is available in the report.
    if result["pyd"] is None and not result["first_year_syndicate"]:
        prov = result.get("adobe_provisions") or {}
        prov_gross = prov.get("gross_prior_year_claims")
        if prov_gross is not None:
            result["pyd"] = prov_gross
            result["pyd_details"] = "from provisions movement note (table extraction)"
            result["method"] = "provisions"
            print(f"  [RAG] Using provisions PYD as fallback: {prov_gross:+.3f}m")

    # Step 5b: If still no PYD, try text-based provisions extraction
    # from reserve movement pages.  This catches provisions notes that
    # Azure's table detection missed (e.g. columnar text layouts).
    if result["pyd"] is None and not result["first_year_syndicate"] and result["reserve_text"]:
        for page_num, page_text in (res_pages if pages else []):
            prov_pyd, prov_details = _extract_pyd_from_provisions_text(page_text)
            if prov_pyd is not None:
                result["pyd"] = prov_pyd
                result["pyd_details"] = prov_details
                result["method"] = "provisions_text"
                print(f"  [RAG] Provisions text PYD: {prov_pyd:+.3f}m ({prov_details})")
                break

    # Step 5c: If still no PYD, try P&L technical account narrative
    # extraction.  Some reports state PYD in prose rather than in a
    # provisions movement table, e.g. "includes £34,490k of releases
    # in respect of prior accident years".
    if result["pyd"] is None and not result["first_year_syndicate"] and result["reserve_text"]:
        for page_num, page_text in (res_pages if pages else []):
            pl_pyd, pl_details = _parse_pyd_from_pl_narrative(page_text)
            if pl_pyd is not None:
                result["pyd"] = pl_pyd
                result["pyd_details"] = pl_details
                result["method"] = "pl_narrative"
                print(f"  [RAG] P&L narrative PYD: {pl_pyd:+.3f}m ({pl_details})")
                break

    # Step 5d: If still no PYD, try YOA-style narrative extraction.
    # Some Lloyd's syndicates state PYD as "Prior year movements of £X.Xm"
    # in their year of account review, rather than in provisions notes or
    # "includes" P&L sentences.
    if result["pyd"] is None and not result["first_year_syndicate"] and result["reserve_text"]:
        for page_num, page_text in (res_pages if pages else []):
            yoa_pyd, yoa_details = _parse_pyd_from_yoa_narrative(page_text)
            if yoa_pyd is not None:
                result["pyd"] = yoa_pyd
                result["pyd_details"] = yoa_details
                result["method"] = "yoa_narrative"
                print(f"  [RAG] YOA narrative PYD: {yoa_pyd:+.3f}m ({yoa_details})")
                break

    # Step 5e: If still no PYD, try general narrative extraction.
    # Catches patterns like "a release of £4.7m of prior year reserves"
    # that don't use the "includes" P&L format or "movements of" YOA format.
    if result["pyd"] is None and not result["first_year_syndicate"] and result["reserve_text"]:
        for page_num, page_text in (res_pages if pages else []):
            gen_pyd, gen_details = _parse_pyd_from_general_narrative(page_text)
            if gen_pyd is not None:
                result["pyd"] = gen_pyd
                result["pyd_details"] = gen_details
                result["method"] = "general_narrative"
                print(f"  [RAG] General narrative PYD: {gen_pyd:+.3f}m ({gen_details})")
                break

    # If after all attempts we have no PYD, no triangle, no reserve text,
    # and it's not a first-year syndicate, flag as "no triangle data" —
    # this report has no usable reserve development information and should
    # be excluded from downstream analysis.
    # Exception: if a loss ratio triangle was found, the report has useful
    # data that LLMs can use to extract PYD from narrative text.
    if (result["pyd"] is None
            and not result["first_year_syndicate"]
            and not result["reserve_text"]
            and not result.get("has_loss_ratio_triangle")):
        result["no_triangle_data"] = True

    return result


# ---------------------------------------------------------------------------
# Triangle verification — compute PYD from raw triangle data in Python
# ---------------------------------------------------------------------------


def _validate_triangle_structure(uw_years, rows, report_year):
    """Check if the triangle has the expected staircase structure.

    In a proper triangle, UW year columns should have development periods
    that form a staircase: the oldest UW year has the most rows filled,
    and each newer year has one fewer row. Returns a score 0.0-1.0.
    """
    n_cols = len(uw_years)
    n_rows = len(rows)
    if n_cols < 2 or n_rows < 2:
        return 0.0

    # Expected pattern: column 0 (oldest) should have most non-nulls,
    # each subsequent column should have one fewer.
    # For run-off syndicates (max UW year < report year), extra development
    # rows shift all columns up by extra_dev_years.
    try:
        max_uw = max(int(y) for y in uw_years)
    except (ValueError, TypeError):
        max_uw = report_year
    extra_dev_years = max(0, report_year - max_uw)

    matches = 0
    checks = 0
    for col_idx in range(n_cols):
        expected_filled = min(n_rows, n_cols - col_idx + extra_dev_years)
        actual_filled = sum(1 for r in range(n_rows) if rows[r][col_idx] is not None)
        checks += 1
        # Allow ±1 tolerance (for summary rows or slight variations)
        if abs(actual_filled - expected_filled) <= 1:
            matches += 1

    return matches / checks if checks > 0 else 0.0


def compute_pyd_from_triangle(triangle_data, report_year):
    """Compute prior year development from extracted triangle data.

    The LLM extracts the raw triangle table (all development rows).
    Python finds the diagonals: for each UW year column, the current estimate
    is the last non-null value, and the previous diagonal is one row above that.

    Args:
        triangle_data: dict with keys: type, units, underwriting_years,
                       development_rows (list of lists)
        report_year: int

    Returns:
        (pyd_value, details_str) or (None, reason_str)
    """
    if not triangle_data or not isinstance(triangle_data, dict):
        return None, "no triangle data"

    tri_type = triangle_data.get("type", "none")
    if tri_type == "none" or tri_type is None:
        return None, "no triangle in report"

    uw_years = triangle_data.get("underwriting_years")
    rows = triangle_data.get("development_rows")

    if not uw_years or not rows:
        return None, "missing triangle arrays"

    # Validate: the most recent UW year should match the report year
    try:
        max_uw = max(int(y) for y in uw_years)
    except (ValueError, TypeError):
        return None, "cannot parse UW years"
    if max_uw > report_year or max_uw < report_year - 5:
        return None, (f"triangle max UW year ({max_uw}) outside range for report year ({report_year}) "
                     f"— likely misaligned extraction")

    n_cols = len(uw_years)

    # Strip trailing all-null rows (development periods with no data yet,
    # e.g. "After five years" when triangle only covers 4 UW years)
    while rows and all(v is None for v in rows[-1]):
        rows.pop()

    n_rows = len(rows)

    # Validate: number of rows should be consistent with the development span.
    # The oldest UW year column can have up to (report_year - min_uw_year + 1)
    # development rows.  For triangles with year gaps (e.g. [2013..2018, 2022])
    # the row count depends on the span, not the column count.
    try:
        min_uw = min(int(y) for y in uw_years)
    except (ValueError, TypeError):
        min_uw = max_uw
    expected_max_rows = report_year - min_uw + 2  # +2 for tolerance
    if n_rows > expected_max_rows:
        return None, (f"triangle has {n_rows} rows but span is only "
                     f"{min_uw}-{report_year} — likely includes summary rows or is misaligned")

    # Normalise row lengths early — pad short rows with None so that all
    # subsequent column-indexed access (validation checks, PYD computation)
    # is safe.  LLM-extracted triangles sometimes return rows shorter than
    # n_cols, which would cause IndexError without this padding.
    for i in range(n_rows):
        if not isinstance(rows[i], list):
            rows[i] = []
        while len(rows[i]) < n_cols:
            rows[i].append(None)

    # Validate: at least one column should have filled rows.
    # In a proper NxN triangle, column 0 (oldest) should have ~N non-null
    # values, but syndicates that started recently may have legitimately
    # empty oldest columns (dashes = no claims activity in early years).
    # Only reject if ALL columns are empty (truly broken extraction).
    if n_rows >= 2 and n_cols >= 2:
        any_col_filled = False
        first_filled_col = None
        for c in range(n_cols):
            col_filled = sum(1 for r in range(n_rows) if rows[r][c] is not None)
            if col_filled > 0:
                any_col_filled = True
                if first_filled_col is None:
                    first_filled_col = c
        if not any_col_filled:
            return None, (f"all columns have 0 filled rows "
                         f"— likely shifted/misaligned")

    if n_rows < 2:
        return None, "triangle has fewer than 2 development rows"

    # Validate: detect year-like values contaminating data rows.
    # When a segmental analysis table is misidentified as a triangle, the
    # development_rows contain calendar year numbers (e.g. 2001, 2013, 2014)
    # alongside actual claims amounts.  Real claims triangles never have
    # values in the 1980-2030 range mixed with small amounts.
    year_like_count = 0
    total_vals = 0
    for r in range(n_rows):
        for c in range(n_cols):
            val = rows[r][c]
            if val is not None:
                try:
                    fv = float(val)
                    total_vals += 1
                    if 1980 <= fv <= 2030 and fv == int(fv):
                        year_like_count += 1
                except (ValueError, TypeError):
                    pass
    if total_vals > 0 and year_like_count / total_vals > 0.15:
        return None, (f"triangle has {year_like_count}/{total_vals} values that look like "
                     f"calendar years (1980-2030) — likely a misidentified segmental table")

    # Validate: detect loss ratio triangles misidentified as claims triangles.
    # Loss ratios are percentages (0-200%), while claims amounts are typically
    # in the hundreds/thousands/millions.  If all non-null values are < 200,
    # this is almost certainly a loss ratio triangle, not a claims triangle.
    if total_vals > 3:
        pct_like_count = sum(
            1 for r in range(n_rows) for c in range(n_cols)
            if rows[r][c] is not None
            and isinstance(rows[r][c], (int, float))
            and 0 < float(rows[r][c]) <= 200
        )
        if pct_like_count == total_vals:
            max_val = max(
                float(rows[r][c])
                for r in range(n_rows) for c in range(n_cols)
                if rows[r][c] is not None and isinstance(rows[r][c], (int, float))
            )
            if max_val <= 200:
                return None, (f"triangle has all {total_vals} values in 0-200 range "
                             f"(max={max_val:.1f}) — likely a loss ratio triangle, "
                             f"not a claims development triangle")

    # Detect and strip "Current estimate" summary row if LLM included it.
    # The summary row repeats the last non-null value from each column's
    # development rows.  It may have nulls in recent columns (where the
    # triangle itself has nulls), so we check non-null entries only.
    # We work on a COPY to avoid mutating the original triangle data.
    rows = [list(r) for r in rows]  # deep copy
    if n_rows >= 3:
        last_row = rows[-1]
        filled_cols = [col for col in range(n_cols) if last_row[col] is not None]
        if len(filled_cols) >= 2:
            # Check if non-null values in last row match the last non-null
            # in each corresponding column above
            matches = 0
            mismatched_cols = []
            for col in filled_cols:
                for r in range(n_rows - 2, -1, -1):
                    if rows[r][col] is not None:
                        try:
                            if abs(float(rows[r][col]) - float(last_row[col])) < 0.01:
                                matches += 1
                            else:
                                mismatched_cols.append(col)
                        except (ValueError, TypeError):
                            pass
                        break
            if matches >= len(filled_cols) * 0.7:
                if 0 in mismatched_cols:
                    # Col 0 has a DIFFERENT value from rows above — this last row
                    # contains real development data for the oldest UW year merged
                    # with the summary row. NULL out only the columns that matched
                    # (those are summary duplicates), keep the mismatched ones.
                    for col in filled_cols:
                        if col not in mismatched_cols:
                            rows[-1][col] = None
                else:
                    # Pure summary row — strip it entirely
                    rows = rows[:-1]
                    n_rows -= 1

    # Validate structure — proper triangles have a staircase pattern
    structure_score = _validate_triangle_structure(uw_years, rows, report_year)

    # For each column, find the current estimate (last non-null)
    # and the previous diagonal (one row above the current)
    details = []
    total_pyd = 0.0
    used_years = 0

    for col_idx, uw_year in enumerate(uw_years):
        try:
            uw_year = int(uw_year)
        except (ValueError, TypeError):
            continue

        # Exclude two most recent UW years
        if uw_year >= report_year - 1:
            continue

        # Find last non-null value in this column (= current estimate)
        current_row_idx = None
        current_val = None
        for row_idx in range(n_rows - 1, -1, -1):
            val = rows[row_idx][col_idx]
            if val is not None:
                try:
                    current_val = float(val)
                    current_row_idx = row_idx
                    break
                except (ValueError, TypeError):
                    continue

        if current_val is None or current_row_idx is None:
            # All-None column for an old enough UW year means zero claims
            # activity (dashes in the triangle = no claims, not missing data).
            # PYD contribution is 0 — count the year as used.
            expected_dev_periods = report_year - uw_year + 1
            if expected_dev_periods >= 2:
                details.append(f"  {uw_year}: all-zero column (no claims activity), PYD=0")
                used_years += 1
            else:
                details.append(f"  {uw_year}: skipped (no current estimate)")
            continue

        # Previous diagonal = one row above in same column
        if current_row_idx == 0:
            details.append(f"  {uw_year}: skipped (only 1 development period)")
            continue

        prev_val = None
        prev_row_idx = current_row_idx - 1
        val = rows[prev_row_idx][col_idx]
        if val is not None:
            try:
                prev_val = float(val)
            except (ValueError, TypeError):
                pass

        if prev_val is None:
            details.append(f"  {uw_year}: skipped (no previous diagonal at row {prev_row_idx})")
            continue

        change = round(current_val - prev_val, 3)
        total_pyd += change
        used_years += 1
        row_labels = ["End of UW yr", "1yr later", "2yr later", "3yr later",
                       "4yr later", "5yr later", "6yr later", "7yr later",
                       "8yr later", "9yr later", "10yr later"]
        cur_label = row_labels[current_row_idx] if current_row_idx < len(row_labels) else f"row{current_row_idx}"
        prev_label = row_labels[prev_row_idx] if prev_row_idx < len(row_labels) else f"row{prev_row_idx}"
        details.append(f"  {uw_year}: {current_val} ({cur_label}) - {prev_val} ({prev_label}) = {change:+.3f}")

    if used_years == 0:
        return None, "no usable UW years in triangle"

    # Convert units
    units = triangle_data.get("units", "millions")

    # Auto-detect units from value magnitudes when header detection was ambiguous.
    # Lloyd's syndicate cumulative claims typically range:
    #   Full integers: 1,000,000 - 500,000,000  (individual pounds)
    #   Thousands:     1,000 - 500,000           (£000)
    #   Millions:      1 - 500                   (£m)
    if units not in ("thousands", "full", "percentage"):
        # Auto-detect units from value magnitudes when header detection was ambiguous
        # or returned an unknown unit string (e.g. "units").
        sample_vals = [rows[r][c] for r in range(min(2, n_rows)) for c in range(n_cols)
                       if rows[r][c] is not None and isinstance(rows[r][c], (int, float))]
        if sample_vals:
            max_val = max(abs(v) for v in sample_vals)
            if max_val > 1_000_000:
                units = "full"
                details.append(f"  (auto-detected: values up to {max_val:,.0f} — treating as full currency units)")
            elif max_val > 10_000:
                units = "thousands"
                details.append(f"  (auto-detected: values up to {max_val:,.0f} — treating as thousands)")
            else:
                units = "millions"

    if units == "full":
        total_pyd = round(total_pyd / 1_000_000, 3)
    elif units == "thousands":
        total_pyd = round(total_pyd / 1000, 3)
        details.append(f"  (converted from thousands to millions)")
    elif units == "percentage":
        return None, "loss ratio triangle -- needs premium data for conversion"

    total_pyd = round(total_pyd, 3)
    details.append(f"  Structure score: {structure_score:.2f}")
    details_str = "\n".join(details) + f"\n  Total PYD = {total_pyd:+.3f}m ({used_years} UW years)"
    return total_pyd, details_str


def _apply_triangle_pyd(result, computed_pyd, model_name, details, reason):
    """Apply a code-computed PYD value to a result dict, updating direction and notes."""
    result = dict(result)
    old_pyd = result.get("prior_year_development_gbp_m")
    # Sanity check BEFORE applying: releases can't exceed 100% of opening
    # reserves (reserves can't go negative).  Strengthenings CAN exceed 100%
    # in extreme scenarios, so only reject negative values below -100%.
    opening = result.get("opening_reserves_gbp_m")
    if computed_pyd != 0 and opening and opening > 0:
        pyd_pct = round(computed_pyd / opening * 100, 2)
        if pyd_pct < -100:
            msg = (f"  [{model_name}] Triangle PYD ratio {pyd_pct:.1f}% < -100% — "
                   f"likely unit mismatch or misaligned triangle. Discarding.")
            return result, msg
    elif computed_pyd != 0 and (not opening or opening == 0):
        # Non-zero PYD from zero opening reserves is nonsensical
        msg = (f"  [{model_name}] Triangle PYD {computed_pyd:+.3f}m but opening reserves = 0 — "
               f"likely misidentified table. Discarding.")
        return result, msg
    result["prior_year_development_gbp_m"] = computed_pyd
    if computed_pyd == 0:
        pyd_pct = 0.0
    elif opening and opening > 0:
        pyd_pct = round(computed_pyd / opening * 100, 2)
    else:
        pyd_pct = None
    result["prior_year_development_pct"] = pyd_pct
    if computed_pyd < 0:
        result["direction"] = "release"
    elif computed_pyd > 0:
        result["direction"] = "strengthening"
    else:
        result["direction"] = "flat"
    if old_pyd is not None:
        old_notes = result.get("data_quality_notes", "")
        result["data_quality_notes"] = (
            f"{old_notes} [CODE OVERRIDE: Model said PYD={old_pyd}, "
            f"but code computed {computed_pyd} from triangle ({reason}). Using code value.]"
        )
    msg = (f"  [{model_name}] Triangle verification: {reason} "
           f"(model={old_pyd}, code={computed_pyd})\n{details}")
    return result, msg


def verify_triangles(result_gemini, result_openai, gemini_name, openai_name, report_year):
    """Cross-validate triangle extractions between models.

    When both models extract triangles, compare the computed PYD values.
    When they agree, use the value with high confidence.
    When they disagree, trust the one with better structural validity.
    When only one has a triangle, use it only if it passes sanity checks.

    Returns updated (result_gemini, result_openai, messages_list).
    """
    tri_g = result_gemini.get("_claims_triangle")
    tri_o = result_openai.get("_claims_triangle")
    has_g = tri_g and tri_g.get("type") not in ("none", None)
    has_o = tri_o and tri_o.get("type") not in ("none", None)

    messages = []

    if not has_g and not has_o:
        return result_gemini, result_openai, messages

    # Compute PYD from each available triangle
    pyd_g, details_g = (None, "no triangle")
    pyd_o, details_o = (None, "no triangle")
    struct_g, struct_o = 0.0, 0.0

    if has_g:
        pyd_g, details_g = compute_pyd_from_triangle(tri_g, report_year)
        uw_g = tri_g.get("underwriting_years", [])
        rows_g = tri_g.get("development_rows", [])
        if uw_g and rows_g:
            struct_g = _validate_triangle_structure(uw_g, rows_g, report_year)

    if has_o:
        pyd_o, details_o = compute_pyd_from_triangle(tri_o, report_year)
        uw_o = tri_o.get("underwriting_years", [])
        rows_o = tri_o.get("development_rows", [])
        if uw_o and rows_o:
            struct_o = _validate_triangle_structure(uw_o, rows_o, report_year)

    # Sanity check: PYD from triangle should be reasonable relative to reserves
    opening_g = result_gemini.get("opening_reserves_gbp_m")
    opening_o = result_openai.get("opening_reserves_gbp_m")
    opening = None
    if opening_g and opening_o:
        try:
            opening = (float(opening_g) + float(opening_o)) / 2
        except (ValueError, TypeError):
            pass
    elif opening_g:
        try:
            opening = float(opening_g)
        except (ValueError, TypeError):
            pass
    elif opening_o:
        try:
            opening = float(opening_o)
        except (ValueError, TypeError):
            pass

    def _passes_sanity(pyd_val, struct_score):
        """Releases must not exceed 100% of opening reserves; structure must be OK."""
        if pyd_val is None:
            return False
        if struct_score < 0.5:
            return False
        if pyd_val == 0:
            return True  # Zero PYD is always sane
        if opening and opening > 0:
            # Releases can't exceed 100% (reserves can't go negative).
            # Strengthenings CAN exceed 100% in extreme scenarios.
            pyd_pct = pyd_val / opening * 100
            if pyd_pct < -100:
                return False
        elif not opening or opening == 0:
            # Non-zero PYD from zero opening reserves is nonsensical
            return False
        return True

    sane_g = _passes_sanity(pyd_g, struct_g)
    sane_o = _passes_sanity(pyd_o, struct_o)

    # Case 1: Both have valid triangles
    if pyd_g is not None and pyd_o is not None:
        if abs(pyd_g - pyd_o) < 1.0:
            # Both agree — high confidence, use the value
            agreed_pyd = round((pyd_g + pyd_o) / 2, 3)
            messages.append(f"  Triangle cross-check: BOTH AGREE (Gemini={pyd_g}, GPT={pyd_o}, "
                           f"struct_g={struct_g:.2f}, struct_o={struct_o:.2f})")
            # Apply to both models
            for result, model_name, details in [
                (result_gemini, gemini_name, details_g),
                (result_openai, openai_name, details_o),
            ]:
                model_pyd = result.get("prior_year_development_gbp_m")
                if model_pyd is None:
                    r, msg = _apply_triangle_pyd(result, agreed_pyd, model_name, details, "FILL from agreed triangles")
                    if model_name == gemini_name:
                        result_gemini = r
                    else:
                        result_openai = r
                    messages.append(msg)
                else:
                    try:
                        if abs(float(model_pyd) - agreed_pyd) < 0.5:
                            messages.append(f"  [{model_name}] Triangle verification: CONFIRMED "
                                          f"(model={model_pyd}, code={agreed_pyd})")
                        else:
                            r, msg = _apply_triangle_pyd(result, agreed_pyd, model_name, details,
                                                        "OVERRIDE from agreed triangles")
                            if model_name == gemini_name:
                                result_gemini = r
                            else:
                                result_openai = r
                            messages.append(msg)
                    except (ValueError, TypeError):
                        pass
            return result_gemini, result_openai, messages

        # Triangles disagree — use the one with better structure
        messages.append(f"  Triangle cross-check: DISAGREE (Gemini={pyd_g} struct={struct_g:.2f}, "
                       f"GPT={pyd_o} struct={struct_o:.2f})")
        if sane_g and not sane_o:
            best_pyd, best_details, best_name = pyd_g, details_g, "Gemini triangle"
        elif sane_o and not sane_g:
            best_pyd, best_details, best_name = pyd_o, details_o, "GPT triangle"
        elif struct_g > struct_o + 0.1:
            best_pyd, best_details, best_name = pyd_g, details_g, "Gemini triangle (better structure)"
        elif struct_o > struct_g + 0.1:
            best_pyd, best_details, best_name = pyd_o, details_o, "GPT triangle (better structure)"
        else:
            # Both have similar structure but disagree — don't trust either
            messages.append("  Triangle cross-check: SKIPPED — triangles disagree and "
                          "neither has clearly better structure. Keeping model values.")
            return result_gemini, result_openai, messages

        messages.append(f"  Using {best_name}: PYD={best_pyd}")
        for result, model_name, details in [
            (result_gemini, gemini_name, best_details),
            (result_openai, openai_name, best_details),
        ]:
            model_pyd = result.get("prior_year_development_gbp_m")
            if model_pyd is None:
                r, msg = _apply_triangle_pyd(result, best_pyd, model_name, details,
                                            f"FILL from {best_name}")
                if model_name == gemini_name:
                    result_gemini = r
                else:
                    result_openai = r
                messages.append(msg)
            else:
                try:
                    if abs(float(model_pyd) - best_pyd) >= 0.5:
                        r, msg = _apply_triangle_pyd(result, best_pyd, model_name, details,
                                                    f"OVERRIDE from {best_name}")
                        if model_name == gemini_name:
                            result_gemini = r
                        else:
                            result_openai = r
                        messages.append(msg)
                except (ValueError, TypeError):
                    pass
        return result_gemini, result_openai, messages

    # Case 2: Only one model has a triangle
    single_pyd = pyd_g if pyd_g is not None else pyd_o
    single_details = details_g if pyd_g is not None else details_o
    single_struct = struct_g if pyd_g is not None else struct_o
    single_sane = sane_g if pyd_g is not None else sane_o
    source_name = gemini_name if pyd_g is not None else openai_name

    if not single_sane:
        messages.append(f"  [{source_name}] Triangle verification: REJECTED — "
                       f"computed PYD={single_pyd} fails sanity check "
                       f"(struct={single_struct:.2f}, "
                       f"opening={opening})\n{single_details}")
        return result_gemini, result_openai, messages

    # Single triangle passes sanity — apply to both models
    messages.append(f"  [{source_name}] Triangle: PYD={single_pyd}, struct={single_struct:.2f}")
    for result, model_name in [
        (result_gemini, gemini_name),
        (result_openai, openai_name),
    ]:
        model_pyd = result.get("prior_year_development_gbp_m")
        if model_pyd is None:
            r, msg = _apply_triangle_pyd(result, single_pyd, model_name, single_details,
                                        f"FILL from {source_name} triangle")
            if model_name == gemini_name:
                result_gemini = r
            else:
                result_openai = r
            messages.append(msg)
        else:
            try:
                if abs(float(model_pyd) - single_pyd) < 0.5:
                    messages.append(f"  [{model_name}] Triangle verification: CONFIRMED "
                                  f"(model={model_pyd}, code={single_pyd})")
                elif abs(float(model_pyd) - single_pyd) >= 0.5:
                    r, msg = _apply_triangle_pyd(result, single_pyd, model_name, single_details,
                                                f"OVERRIDE from {source_name} triangle")
                    if model_name == gemini_name:
                        result_gemini = r
                    else:
                        result_openai = r
                    messages.append(msg)
            except (ValueError, TypeError):
                pass

    return result_gemini, result_openai, messages


# ---------------------------------------------------------------------------
# Currency field normalization
# ---------------------------------------------------------------------------

# The LLM prompt uses _gbp_m as the canonical field suffix for ALL currencies,
# with a separate "currency" field to indicate the actual currency.  Some LLMs
# (notably Gemini) rename the fields to _usd_m or _eur_m for non-GBP syndicates.
# This function remaps those back to the canonical _gbp_m names so that
# downstream comparison / auto-resolution logic works correctly.

_CURRENCY_SUFFIXES = ("_usd_m", "_eur_m")
_CANONICAL_SUFFIX = "_gbp_m"


def _normalize_currency_fields(result):
    """Rename _usd_m / _eur_m fields to _gbp_m in an LLM extraction result.

    Operates in-place on the top-level dict AND on dicts inside list fields
    (lob_movements, named_events, prior_year_events, gross_premium_mix).

    Returns True if any renaming was performed.
    """
    renamed = False

    # --- Top-level fields ---
    for key in list(result.keys()):
        for suffix in _CURRENCY_SUFFIXES:
            if key.endswith(suffix):
                canonical = key[: -len(suffix)] + _CANONICAL_SUFFIX
                # Only rename if the canonical slot is absent or None
                if result.get(canonical) is None:
                    result[canonical] = result.pop(key)
                    renamed = True
                else:
                    # Canonical already has a value — just remove the variant
                    del result[key]
                    renamed = True
                break  # done with this key

    # --- Nested list-of-dicts fields ---
    for list_field in ("lob_movements", "named_events", "prior_year_events",
                       "gross_premium_mix"):
        items = result.get(list_field)
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            for key in list(item.keys()):
                for suffix in _CURRENCY_SUFFIXES:
                    if key.endswith(suffix):
                        canonical = key[: -len(suffix)] + _CANONICAL_SUFFIX
                        if item.get(canonical) is None:
                            item[canonical] = item.pop(key)
                            renamed = True
                        else:
                            del item[key]
                            renamed = True
                        break

    return renamed


# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------

SKIP_FIELDS = {
    "source_type", "source_file", "content_hash",
    "standardized_at", "standardization_model", "_extraction_meta",
    "_claims_triangle",
    "_rag_triangle",
    "_adobe_lob",
    "_adobe_provisions",
}


def compare_results(a, b, model_a, model_b):
    """Compare two extraction results and return discrepancies."""
    discrepancies = []
    all_keys = sorted(set(list(a.keys()) + list(b.keys())) - SKIP_FIELDS)

    for key in all_keys:
        val_a = a.get(key, "<MISSING>")
        val_b = b.get(key, "<MISSING>")

        if key in ("lob_movements", "gross_premium_mix"):
            _compare_list_of_dicts(key, val_a, val_b, model_a, model_b, discrepancies)
        elif key in ("primary_causes", "specific_events", "specific_years_affected", "raw_causal_phrases"):
            _compare_lists(key, val_a, val_b, model_a, model_b, discrepancies)
        elif key in ("exact_reserve_text", "standardized_narrative", "data_quality_notes"):
            if str(val_a).strip() != str(val_b).strip():
                discrepancies.append({
                    "field": key, "type": "text_difference",
                    model_a: _truncate(str(val_a), 120),
                    model_b: _truncate(str(val_b), 120),
                })
        else:
            if val_a != val_b:
                discrepancies.append({
                    "field": key, "type": "value_mismatch",
                    model_a: val_a, model_b: val_b,
                })

    return discrepancies


def check_tolerance(discrepancies, model_a, model_b):
    """Classify discrepancies as tolerated or hard failures. Returns (passed, tolerated, hard_failures)."""
    hard_failures = []
    tolerated = []
    for d in discrepancies:
        field = d["field"]
        is_tolerated = (
            d["type"] == "text_difference"
            or any(t in field for t in ("page", "confidence", "Page", "Confidence"))
            or any(field.startswith(p) for p in (
                "named_events", "prior_year_events", "raw_causal_phrases",
                "specific_events", "specific_years_affected",
                "lob_movements", "primary_causes",
            ))
            or _is_numeric_near(d.get(model_a), d.get(model_b), rel_tol=0.005)
        )
        # For gross_premium_mix percentages, when total GPW is negative the
        # sign convention is ambiguous (e.g. 88% vs -87.43%).  Tolerate if
        # the absolute values are close.
        if (not is_tolerated
                and "percentage_of_total" in field
                and "gross_premium_mix" in field):
            try:
                va = abs(float(d.get(model_a)))
                vb = abs(float(d.get(model_b)))
                if _is_numeric_near(va, vb, rel_tol=0.05):
                    is_tolerated = True
            except (TypeError, ValueError):
                pass
        if is_tolerated:
            tolerated.append(d)
        else:
            hard_failures.append(d)

    return len(hard_failures) == 0, tolerated, hard_failures


def resolve_computed_fields(hard_failures, result_a, result_b, model_a, model_b,
                            syndicate_num=None, report_year=None):
    """Auto-resolve derived fields when the inputs they depend on agree.

    prior_year_development_pct = prior_year_development_gbp_m / opening_reserves_gbp_m * 100

    If both models agree on the numerator and denominator, compute the correct
    percentage and mark whichever model is closer as tolerated.

    Returns (remaining_hard_failures, auto_resolved).
    """
    remaining = []
    resolved = []

    # Known-truth fields: syndicate number and report year come from the
    # filename, so when one LLM returns None we can auto-resolve.
    known_truth = {}
    if syndicate_num is not None:
        known_truth["syndicate"] = syndicate_num
        known_truth["syndicate_number"] = syndicate_num
    if report_year is not None:
        known_truth["report_year"] = report_year

    for d in hard_failures:
        # Auto-resolve fields with known ground truth (from filename)
        if d["field"] in known_truth:
            truth = known_truth[d["field"]]
            val_a = d.get(model_a)
            val_b = d.get(model_b)
            # Accept if one model got it right or returned None
            if val_a == truth or val_b == truth or val_a is None or val_b is None:
                # Fill in the correct value on whichever model was wrong/null
                result_a[d["field"]] = truth
                result_b[d["field"]] = truth
                closer = model_a if val_a == truth else model_b
                print(f"  Auto-resolved {d['field']}: {model_a}={val_a}, "
                      f"{model_b}={val_b}, truth={truth}")
                d["auto_resolved"] = True
                d["closer_model"] = closer
                resolved.append(d)
                continue
        if d["field"] == "prior_year_development_pct":
            pyd_a = result_a.get("prior_year_development_gbp_m")
            pyd_b = result_b.get("prior_year_development_gbp_m")
            open_a = result_a.get("opening_reserves_gbp_m")
            open_b = result_b.get("opening_reserves_gbp_m")

            # Both models must agree on the inputs (within tolerance)
            if (pyd_a is not None and pyd_b is not None
                    and open_a is not None and open_b is not None
                    and _is_numeric_near(pyd_a, pyd_b)
                    and _is_numeric_near(open_a, open_b)
                    and float(open_a) != 0):
                computed = round(float(pyd_a) / float(open_a) * 100, 2)
                print(f"  Auto-resolved prior_year_development_pct: "
                      f"{pyd_a} / {open_a} * 100 = {computed}%")

                # Check which model is closer
                val_a = d.get(model_a)
                val_b = d.get(model_b)
                try:
                    err_a = abs(float(val_a) - computed) if val_a is not None else 999
                    err_b = abs(float(val_b) - computed) if val_b is not None else 999
                except (TypeError, ValueError):
                    err_a = err_b = 999

                closer = model_a if err_a <= err_b else model_b
                closer_val = val_a if err_a <= err_b else val_b
                print(f"    Computed: {computed}, {model_a}: {val_a}, {model_b}: {val_b}")
                print(f"    {closer} is closer (error: {min(err_a, err_b):.4f})")

                d["auto_resolved"] = True
                d["computed_value"] = computed
                d["closer_model"] = closer
                resolved.append(d)
                continue

        if d["field"] == "direction":
            dir_a = d.get(model_a)
            dir_b = d.get(model_b)
            pyd_a = result_a.get("prior_year_development_gbp_m")
            pyd_b = result_b.get("prior_year_development_gbp_m")
            # Auto-resolve when one model returned null (couldn't extract)
            # and the other has a valid direction consistent with its PYD
            if dir_a is None and dir_b is not None:
                # Accept model_b's direction if its PYD is consistent
                if (pyd_b == 0 and dir_b == "flat") or \
                   (pyd_b is not None and pyd_b < 0 and dir_b == "release") or \
                   (pyd_b is not None and pyd_b > 0 and dir_b == "strengthening") or \
                   pyd_a is None:  # model_a has no PYD info at all
                    print(f"  Auto-resolved direction: {model_a}=null, "
                          f"accepting {model_b}={dir_b}")
                    result_a["direction"] = dir_b
                    d["auto_resolved"] = True
                    d["closer_model"] = model_b
                    resolved.append(d)
                    continue
            elif dir_b is None and dir_a is not None:
                if (pyd_a == 0 and dir_a == "flat") or \
                   (pyd_a is not None and pyd_a < 0 and dir_a == "release") or \
                   (pyd_a is not None and pyd_a > 0 and dir_a == "strengthening") or \
                   pyd_b is None:
                    print(f"  Auto-resolved direction: {model_b}=null, "
                          f"accepting {model_a}={dir_a}")
                    result_b["direction"] = dir_a
                    d["auto_resolved"] = True
                    d["closer_model"] = model_a
                    resolved.append(d)
                    continue

        if d["field"] == "opening_reserves_gbp_m":
            # Auto-resolve using RAG opening claims outstanding from provisions table
            prov_a = result_a.get("_adobe_provisions") or {}
            rag_opening = prov_a.get("opening_gross_claims_outstanding")
            if rag_opening is not None:
                val_a = d.get(model_a)
                val_b = d.get(model_b)
                # Use RAG value and accept whichever LLM is closer
                try:
                    err_a = abs(float(val_a) - rag_opening) if val_a is not None else 999
                    err_b = abs(float(val_b) - rag_opening) if val_b is not None else 999
                except (TypeError, ValueError):
                    err_a = err_b = 999
                closer = model_a if err_a <= err_b else model_b
                print(f"  Auto-resolved opening_reserves_gbp_m using RAG provisions table: "
                      f"{rag_opening:.3f}m")
                print(f"    {model_a}: {val_a}, {model_b}: {val_b}, RAG: {rag_opening:.3f}")
                # Override both models with RAG value
                result_a["opening_reserves_gbp_m"] = rag_opening
                result_b["opening_reserves_gbp_m"] = rag_opening
                # Recompute PYD percentage with corrected opening reserves
                for r in [result_a, result_b]:
                    pyd = r.get("prior_year_development_gbp_m")
                    if pyd is not None and rag_opening > 0:
                        r["prior_year_development_pct"] = round(
                            float(pyd) / rag_opening * 100, 2
                        )
                    elif rag_opening == 0 or pyd == 0:
                        r["prior_year_development_pct"] = 0.0
                d["auto_resolved"] = True
                d["rag_value"] = rag_opening
                d["closer_model"] = closer
                resolved.append(d)
                continue

        remaining.append(d)

    return remaining, resolved


def _get_field_context(field, gem_data, gpt_data):
    """Build supporting context lines for a disputed field.

    Returns a list of formatted strings showing related fields from both
    models (e.g. opening_reserves and prior_year_development_gbp_m when
    the disputed field is prior_year_development_pct).
    """
    context_map = {
        "prior_year_development_pct": [
            "opening_reserves_gbp_m",
            "prior_year_development_gbp_m",
        ],
        "prior_year_development_gbp_m": [
            "opening_reserves_gbp_m",
            "direction",
            "exact_reserve_text",
        ],
        "direction": [
            "prior_year_development_gbp_m",
            "exact_reserve_text",
        ],
    }

    # For gross_premium_mix sub-fields, show the total premiums
    field_base = field.split("[")[0]
    if field_base == "gross_premium_mix":
        context_map[field] = ["gross_premiums_written_gbp_m", "currency"]

    related = context_map.get(field) or context_map.get(field_base)
    if not related:
        return []

    lines = []
    for rf in related:
        g_val = gem_data.get(rf)
        o_val = gpt_data.get(rf)
        agree = ""
        if g_val is not None and o_val is not None:
            g_lower = str(g_val).strip().lower()
            o_lower = str(o_val).strip().lower()
            if _is_numeric_near(g_val, o_val):
                agree = " (AGREE)"
            elif g_lower == o_lower:
                agree = " (AGREE)"
            elif g_lower in o_lower or o_lower in g_lower:
                agree = " (AGREE)"
            else:
                agree = " (DISAGREE)"
        g_str = str(g_val) if g_val is not None else "null"
        o_str = str(o_val) if o_val is not None else "null"
        lines.append(f"{rf}{agree}:")
        lines.append(f"  Gemini: {g_str}")
        lines.append(f"  GPT:    {o_str}")
    return lines


def print_discrepancies(discrepancies, model_a, model_b):
    """Print discrepancies to console."""
    if not discrepancies:
        print("    No discrepancies.")
        return
    for i, d in enumerate(discrepancies, 1):
        print(f"    {i}. [{d['type']}] {d['field']}")
        if d["type"] == "list_extra":
            print(f"       Only in {d['only_in']}: {_truncate(str(d['value']), 100)}")
        elif d["type"] == "text_difference":
            print(f"       {model_a}: {d[model_a]}")
            print(f"       {model_b}: {d[model_b]}")
        else:
            print(f"       {model_a}: {d.get(model_a, '')}")
            print(f"       {model_b}: {d.get(model_b, '')}")


def _compare_lists(field, a, b, model_a, model_b, discrepancies):
    """Compare two lists of strings (set-based for unordered lists)."""
    if not isinstance(a, list):
        a = []
    if not isinstance(b, list):
        b = []

    set_a = set(str(x) for x in a)
    set_b = set(str(x) for x in b)

    for item in sorted(set_a - set_b):
        discrepancies.append({
            "field": field, "type": "list_extra",
            "only_in": model_a, "value": item,
        })
    for item in sorted(set_b - set_a):
        discrepancies.append({
            "field": field, "type": "list_extra",
            "only_in": model_b, "value": item,
        })


def _is_zero_entry(item):
    """Check if a premium mix / LOB movement entry has zero or null amounts."""
    amt = item.get("amount_gbp_m", 0) or 0
    pct = item.get("percentage_of_total", 0) or 0
    return abs(float(amt)) < 0.01 and abs(float(pct)) < 0.01


def _is_subtotal_entry(item):
    """Check if an entry is a subtotal row (e.g. 'Total direct insurance')."""
    lob = str(item.get("line_of_business", "")).lower()
    return lob.startswith("total") or lob.startswith("sub-total") or lob.startswith("subtotal")


def _compare_list_of_dicts(field, a, b, model_a, model_b, discrepancies):
    """Compare two lists of dicts by matching on line_of_business key."""
    if not isinstance(a, list):
        a = []
    if not isinstance(b, list):
        b = []

    key_field = "line_of_business"
    index_a = {item.get(key_field, f"__idx{i}"): item for i, item in enumerate(a)}
    index_b = {item.get(key_field, f"__idx{i}"): item for i, item in enumerate(b)}

    all_keys = sorted(set(list(index_a.keys()) + list(index_b.keys())))
    for k in all_keys:
        if k not in index_a:
            # Tolerate if the extra entry has zero amount or is a subtotal row
            if _is_zero_entry(index_b[k]) or _is_subtotal_entry(index_b[k]):
                continue
            discrepancies.append({
                "field": f"{field}[{k}]", "type": "list_extra",
                "only_in": model_b, "value": json.dumps(index_b[k]),
            })
        elif k not in index_b:
            if _is_zero_entry(index_a[k]) or _is_subtotal_entry(index_a[k]):
                continue
            discrepancies.append({
                "field": f"{field}[{k}]", "type": "list_extra",
                "only_in": model_a, "value": json.dumps(index_a[k]),
            })
        else:
            item_a, item_b = index_a[k], index_b[k]
            for sub_key in sorted(set(list(item_a.keys()) + list(item_b.keys()))):
                va = item_a.get(sub_key, "<MISSING>")
                vb = item_b.get(sub_key, "<MISSING>")
                if va != vb:
                    discrepancies.append({
                        "field": f"{field}[{k}].{sub_key}",
                        "type": "value_mismatch",
                        model_a: va, model_b: vb,
                    })


def _truncate(s, max_len):
    """Truncate string for display."""
    return s[:max_len] + "..." if len(s) > max_len else s


def _is_numeric_near(a, b, rel_tol=0.005, abs_tol=0.05):
    """Check if two values are both numeric and within tolerance.

    Uses relative tolerance for large values, absolute tolerance for near-zero.
    """
    try:
        # Treat None/null vs 0.0 as equivalent (both mean "not found / zero")
        a_is_none = a is None or a == "None"
        b_is_none = b is None or b == "None"
        if a_is_none and b_is_none:
            return True
        if a_is_none:
            a = 0.0
        if b_is_none:
            b = 0.0
        fa, fb = float(a), float(b)
        if fa == fb:
            return True
        diff = abs(fa - fb)
        if diff <= abs_tol:
            return True
        denom = max(abs(fa), abs(fb), 1e-9)
        return diff / denom <= rel_tol
    except (TypeError, ValueError):
        return False


def append_to_disagreement_log(report_stem, hard_failures):
    """Append hard failure entries to the disagreement log for later manual adjudication."""
    log_path = AUDIT_DIR / "disagreement_log.json"
    with open(log_path, "r") as f:
        log = json.load(f)

    existing_ids = {e["id"] for e in log["entries"]}
    next_num = max(
        (int(e["id"].split("-")[1]) for e in log["entries"] if e["id"].startswith("RUN-")),
        default=0,
    ) + 1

    for d in hard_failures:
        entry_id = f"RUN-{next_num:04d}"
        if any(
            e["report"] == report_stem and e["field"] == d["field"]
            and e["status"] != "pending_review"
            for e in log["entries"]
        ):
            continue  # already adjudicated in a prior dev entry
        # Skip if already logged as pending for this report+field
        if any(
            e["report"] == report_stem and e["field"] == d["field"]
            for e in log["entries"]
        ):
            continue

        entry = {
            "id": entry_id,
            "report": report_stem,
            "field": d["field"],
            "gemini_value": d.get(GEMINI_MODEL, d.get("only_in", "")),
            "gpt_value": d.get(OPENAI_MODEL, d.get("value", "")),
            "correct_model": None,
            "manual_verification": None,
            "action": None,
            "prompt_version_before": PROMPT_VERSION,
            "prompt_version_after": None,
            "status": "pending_review",
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        }
        log["entries"].append(entry)
        next_num += 1

    with open(log_path, "w") as f:
        json.dump(sanitize_json_ascii(log), f, indent=2, ensure_ascii=True)


def write_run_manifest(run_stats):
    """Append this run's stats to the cumulative run manifest log."""
    entry = {
        "run_id": datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ"),
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "spec": {
            "prompt_version": PROMPT_VERSION,
            "field_definitions_version": FIELD_DEFINITIONS_VERSION,
            "tolerance_rules_version": TOLERANCE_RULES_VERSION,
        },
        "models": {
            "model_a": GEMINI_MODEL,
            "model_b": OPENAI_MODEL,
        },
        "corpus": {
            "reports_dir": str(REPORTS_DIR),
            "total_reports_found": run_stats["total_found"],
            "reports_processed_this_run": run_stats["processed"],
            "reports_passed": run_stats["passed"],
            "reports_failed": run_stats["failed"],
            "reports_errored": run_stats["errored"],
            "reports_skipped_already_done": run_stats["already_done"],
        },
        "costs": {
            "total_tokens": run_stats["total_tokens"],
            "total_cost_usd": round(run_stats["total_cost"], 4),
        },
        "output_dir": str(OUTPUT_DIR),
        "stopped_early": run_stats.get("stopped_early", False),
    }

    manifest_path = AUDIT_DIR / "run_manifest.json"

    # Load existing runs or start fresh
    if manifest_path.exists():
        with open(manifest_path, "r") as f:
            manifest = json.load(f)
    else:
        manifest = {"runs": []}

    # Migrate from old single-run format to runs list
    if "runs" not in manifest:
        manifest = {"runs": [manifest]}

    manifest["runs"].append(entry)

    with open(manifest_path, "w") as f:
        json.dump(sanitize_json_ascii(manifest), f, indent=2, ensure_ascii=True)
    print(f"  Run manifest updated: {manifest_path} ({len(manifest['runs'])} runs total)")


def _pyd_direction(value):
    """Return the reserve-development direction implied by a numeric PYD."""
    return "release" if value < 0 else "strengthening" if value > 0 else "flat"


def _apply_loss_ratio_fallback(result, rag_pyd, model_name):
    """Apply the conditional hierarchy for a managed-level loss-ratio PYD.

    A deterministic loss-ratio result fills a narrative blank.  A syndicate-specific
    narrative value is retained when its direction agrees, but a sign contradiction
    is resolved in favour of the deterministic result.  The return value names the
    exercised branch so focused behavioural tests can verify the hierarchy.
    """
    model_pyd = result.get("prior_year_development_gbp_m")
    rag_dir = _pyd_direction(rag_pyd)
    if model_pyd is None:
        result["prior_year_development_gbp_m"] = rag_pyd
        opening = result.get("opening_reserves_gbp_m")
        if rag_pyd == 0:
            result["prior_year_development_pct"] = 0.0
        elif opening and opening > 0:
            result["prior_year_development_pct"] = round(rag_pyd / opening * 100, 2)
        result["direction"] = rag_dir
        old_notes = result.get("data_quality_notes", "") or ""
        result["data_quality_notes"] = (
            f"{old_notes} [MANAGED LEVEL: PYD from loss ratio triangle is at "
            f"managed/group level, not syndicate share. May differ from "
            f"syndicate-level figure.]"
        ).strip()
        print(f"  [{model_name}] PYD filled from RAG triangle (managed level): "
              f"{rag_pyd:+.3f}m")
        return "filled_blank"

    try:
        old_pyd = float(model_pyd)
    except (ValueError, TypeError):
        return "invalid_narrative"
    llm_dir = _pyd_direction(old_pyd)
    if llm_dir != rag_dir and rag_dir != "flat":
        result["prior_year_development_gbp_m"] = rag_pyd
        opening = result.get("opening_reserves_gbp_m")
        if rag_pyd == 0:
            result["prior_year_development_pct"] = 0.0
        elif opening and opening > 0:
            result["prior_year_development_pct"] = round(rag_pyd / opening * 100, 2)
        result["direction"] = rag_dir
        old_notes = result.get("data_quality_notes", "") or ""
        result["data_quality_notes"] = (
            f"{old_notes} [RAG DIRECTION OVERRIDE: LLM said PYD={old_pyd} "
            f"({llm_dir}), loss ratio triangle computed {rag_pyd:+.3f}m "
            f"({rag_dir}). LLM direction contradicts deterministic RAG; "
            f"overriding with RAG value.]"
        ).strip()
        print(f"  [{model_name}] RAG DIRECTION OVERRIDE: LLM PYD={old_pyd} "
              f"({llm_dir}) contradicts RAG {rag_pyd:+.3f}m ({rag_dir})")
        return "overrode_contradiction"

    print(f"  [{model_name}] Keeping LLM PYD={old_pyd} "
          f"(loss ratio triangle gives {rag_pyd:+.3f}m at managed/group level)")
    return "kept_agreement"


def process_one_report(report_path, inception_cache=None):
    """Process a single report with both models.

    Returns (output_data, passed, discrepancies, hard_failures) or
    None if the report should be skipped (e.g. first-year syndicate).
    """
    if inception_cache is None:
        inception_cache = _load_inception_years()

    syndicate_num, report_year = parse_report_filename(report_path)

    # Check inception cache — but don't skip yet.  The cache may be wrong
    # (e.g. Perplexity returned the wrong year).  We'll validate against
    # the actual triangle below and correct the cache if needed.
    inception_skip, inception_year = is_early_year_syndicate(
        syndicate_num, report_year, inception_cache
    )

    # Convert HTML to PDF if needed
    actual_path = report_path
    if report_path.suffix.lower() in (".html", ".htm"):
        actual_path = convert_html_to_pdf(report_path)

    with open(actual_path, "rb") as f:
        file_bytes = f.read()
    content_hash = hashlib.sha256(file_bytes).hexdigest()

    # Early check: run RAG-lite first to detect first-year syndicates
    # before spending money on LLM calls
    rag_result = extract_pyd_from_relevant_pages(actual_path, report_year)

    # Correct inception cache if the triangle contradicts it.
    # The triangle is ground truth — if the RAG found a valid triangle
    # with usable UW years, the inception cache was wrong.
    # BUT: never overwrite manually-verified inception years.
    manual_overrides = _load_manual_overrides()
    if inception_skip and not rag_result.get("first_year_syndicate"):
        tri = rag_result.get("triangle")
        uw_years = tri.get("underwriting_years", []) if tri else []
        if uw_years:
            tri_inception = min(int(y) for y in uw_years)
            old_inception = inception_cache.get(syndicate_num)
            if syndicate_num in manual_overrides:
                print(f"  [Inception] Triangle shows UW years back to {tri_inception}, "
                      f"but syndicate {syndicate_num} has manual override ({old_inception}) — keeping manual value, not skipping")
                inception_skip = False
            else:
                print(f"  [Inception] Cache said inception={old_inception} (would skip), "
                      f"but triangle shows UW years back to {tri_inception} — correcting cache")
                inception_cache[syndicate_num] = tri_inception
                _save_inception_years(inception_cache)
                inception_skip = False

    if rag_result.get("first_year_syndicate"):
        # Triangle has <=2 UW years — update inception cache if this is new info
        if syndicate_num not in inception_cache and syndicate_num not in manual_overrides:
            # Conservative estimate: first UW year = report_year - 1
            inception_cache[syndicate_num] = report_year - 1
            _save_inception_years(inception_cache)
            print(f"  Inception year for syndicate {syndicate_num} estimated as {report_year - 1} (from triangle)")
        # Build minimal audit-trail JSON for first-year syndicates
        first_year_output = {
            "extraction_timestamp": datetime.now(timezone.utc).isoformat(),
            "spec": {
                "prompt_version": PROMPT_VERSION,
                "field_definitions_version": FIELD_DEFINITIONS_VERSION,
                "tolerance_rules_version": TOLERANCE_RULES_VERSION,
            },
            "source_file": str(report_path),
            "first_year_syndicate": True,
            "reason": "Syndicate too new — insufficient underwriting years for prior year development analysis",
            "syndicate": syndicate_num,
            "year": report_year,
        }
        adobe_lob = rag_result.get("adobe_lob")
        if adobe_lob:
            first_year_output["gross_premium_mix"] = adobe_lob["gross_premium_mix"]
            first_year_output["gross_premiums_written_gbp_m"] = adobe_lob["gross_premiums_written_gbp_m"]
            first_year_output["currency"] = adobe_lob.get("currency", "GBP")
        return "first_year", first_year_output

    if rag_result.get("no_triangle_data"):
        # No triangle, no reserve text — report has no usable reserve
        # development data.  Skip expensive LLM calls and recommend exclusion.
        no_data_output = {
            "extraction_timestamp": datetime.now(timezone.utc).isoformat(),
            "spec": {
                "prompt_version": PROMPT_VERSION,
                "field_definitions_version": FIELD_DEFINITIONS_VERSION,
                "tolerance_rules_version": TOLERANCE_RULES_VERSION,
            },
            "source_file": str(report_path),
            "no_triangle_data": True,
            "excluded": True,
            "exclusion_reason": "No claims development triangle or reserve movement text found in report — recommend non-inclusion in analysis",
            "syndicate": syndicate_num,
            "year": report_year,
        }
        adobe_lob = rag_result.get("adobe_lob")
        if adobe_lob:
            no_data_output["gross_premium_mix"] = adobe_lob["gross_premium_mix"]
            no_data_output["gross_premiums_written_gbp_m"] = adobe_lob["gross_premiums_written_gbp_m"]
            no_data_output["currency"] = adobe_lob.get("currency", "GBP")
        return "no_triangle_data", no_data_output

    # Build slim PDF with only relevant pages for LLM extraction
    relevant_pages = rag_result.get("relevant_pages", [])
    rotated_pages = rag_result.get("rotated_pages", set())
    if relevant_pages:
        slim_dir = OUTPUT_DIR / "llm_slim"
        slim_dir.mkdir(parents=True, exist_ok=True)
        llm_pdf = slim_dir / f"{actual_path.stem}_llm.pdf"
        _extract_pages_to_pdf(actual_path, relevant_pages, llm_pdf, rotated_pages)
        slim_size = llm_pdf.stat().st_size / 1024
        print(f"  [LLM] Slim PDF: {len(relevant_pages)} pages, {slim_size:.0f} KB "
              f"(full report: {len(file_bytes) / 1024:.0f} KB)")
        with open(llm_pdf, "rb") as f:
            llm_bytes = f.read()
        llm_hash = hashlib.sha256(llm_bytes).hexdigest()
    else:
        # Fallback: send full PDF if no relevant pages identified
        llm_pdf = actual_path
        llm_bytes = file_bytes
        llm_hash = content_hash

    # Extract with both models (slim PDF with relevant pages only)
    result_gemini = extract_with_gemini(
        llm_pdf, llm_bytes, llm_hash, syndicate_num, report_year
    )
    result_openai = extract_with_openai(
        llm_pdf, llm_bytes, llm_hash, syndicate_num, report_year
    )

    # Normalize currency field names: some LLMs rename _gbp_m → _usd_m/_eur_m
    # for non-GBP syndicates.  Remap to canonical _gbp_m before comparison.
    for _r, _name in [(result_gemini, GEMINI_MODEL), (result_openai, OPENAI_MODEL)]:
        if _normalize_currency_fields(_r):
            print(f"  [{_name}] Normalized currency field names → _gbp_m")

    # RAG-lite was already run above (before LLM calls) for early first-year detection.
    # Use the cached result — no need to re-run.
    rag_method = rag_result.get("method", "")
    # Loss ratio triangle PYD is often at managed/group level rather than
    # syndicate level (e.g. Beazley 2623). It fills a narrative blank and otherwise
    # retains the syndicate-level value unless their directions contradict.
    # Net triangles are treated as authoritative — the deterministic PYD
    # computation is more reliable than LLM extraction, even for net
    # triangles (e.g. syndicate 386/2018 where LLMs wildly disagree).
    rag_tri_type = (rag_result.get("triangle") or {}).get("type", "none")
    rag_is_fallback_only = (
        rag_method == "loss_ratio_triangle"
    )

    if rag_result["pyd"] is not None:
        # RAG found a valid triangle PYD — use it as ground truth
        rag_pyd = rag_result["pyd"]
        rag_details = rag_result["pyd_details"]

        # Sanity check: RAG PYD should be reasonable relative to opening reserves.
        # If |PYD%| > 100%, the triangle is likely garbled (e.g. segmental analysis
        # table misidentified as a claims triangle with year values in data rows).
        rag_sane = True
        avg_opening = None
        for r in [result_gemini, result_openai]:
            op = r.get("opening_reserves_gbp_m")
            if op and op > 0:
                avg_opening = op
                break
        if avg_opening and avg_opening > 0:
            rag_pyd_pct = rag_pyd / avg_opening * 100
            # Releases can't exceed 100% (reserves can't go negative).
            if rag_pyd_pct < -100:
                rag_sane = False
                print(f"  [RAG] Triangle PYD {rag_pyd:+.3f}m ({rag_pyd_pct:.1f}% of opening) "
                      f"< -100% — likely misidentified table. Discarding RAG PYD.")
            # Strengthenings > 200% are implausible — likely a garbled
            # triangle or bogus "ultimates" values.
            elif rag_pyd_pct > 200:
                rag_sane = False
                print(f"  [RAG] Triangle PYD {rag_pyd:+.3f}m ({rag_pyd_pct:.1f}% of opening) "
                      f"> +200% — likely misidentified table. Discarding RAG PYD.")
        elif abs(rag_pyd) > 0.1:
            # Non-zero PYD but zero opening reserves -- can't have reserve
            # development when there are no reserves.  Likely a net triangle
            # or misidentified table.
            rag_sane = False
            print(f"  [RAG] Triangle PYD {rag_pyd:+.3f}m but opening reserves = 0 "
                  f"— likely misidentified table or net triangle. Discarding RAG PYD.")

        if rag_sane:
            print(f"  [RAG] Triangle PYD: {rag_pyd:+.3f}m")

            # Apply to both models
            for result, model_name in [
                (result_gemini, GEMINI_MODEL),
                (result_openai, OPENAI_MODEL),
            ]:
                if rag_is_fallback_only:
                    _apply_loss_ratio_fallback(result, rag_pyd, model_name)
                    continue
                model_pyd = result.get("prior_year_development_gbp_m")
                if model_pyd is None:
                    result["prior_year_development_gbp_m"] = rag_pyd
                    _op = result.get("opening_reserves_gbp_m")
                    if rag_pyd == 0:
                        result["prior_year_development_pct"] = 0.0
                    elif _op and _op > 0:
                        result["prior_year_development_pct"] = round(
                            rag_pyd / _op * 100, 2
                        )
                    result["direction"] = "release" if rag_pyd < 0 else "strengthening" if rag_pyd > 0 else "flat"
                    level = " (managed level)" if rag_method == "loss_ratio_triangle" else ""
                    net_note = " (net triangle)" if rag_tri_type == "net" else ""
                    print(f"  [{model_name}] PYD filled from RAG triangle{level}{net_note}: {rag_pyd:+.3f}m")
                    if rag_tri_type == "net":
                        old_notes = result.get("data_quality_notes", "") or ""
                        result["data_quality_notes"] = (
                            f"{old_notes} [NET TRIANGLE: PYD from net-of-reinsurance "
                            f"claims development triangle.]"
                        ).strip()
                else:
                    try:
                        old_pyd = float(model_pyd)
                        diff = abs(old_pyd - rag_pyd)

                        # Standard claims triangle — deterministic PYD is
                        # authoritative over LLM-extracted values (which may
                        # come from wrong sources like net movements or P&L
                        # gross change in provision).
                        result["prior_year_development_gbp_m"] = rag_pyd
                        _op = result.get("opening_reserves_gbp_m")
                        if rag_pyd == 0:
                            result["prior_year_development_pct"] = 0.0
                        elif _op and _op > 0:
                            result["prior_year_development_pct"] = round(
                                rag_pyd / _op * 100, 2
                            )
                        result["direction"] = "release" if rag_pyd < 0 else "strengthening" if rag_pyd > 0 else "flat"
                        if diff >= 0.5:
                            old_notes = result.get("data_quality_notes", "") or ""
                            result["data_quality_notes"] = (
                                f"{old_notes} [RAG OVERRIDE: Model said PYD={old_pyd}, "
                                f"RAG triangle computed {rag_pyd}. Using RAG value.]"
                            )
                            print(f"  [{model_name}] PYD overridden by RAG triangle: {old_pyd} -> {rag_pyd:+.3f}m")
                        else:
                            print(f"  [{model_name}] PYD confirmed by RAG triangle: {model_pyd} -> {rag_pyd:+.3f}m")
                    except (ValueError, TypeError):
                        pass

        # Store RAG triangle in both results for reference (even if PYD was rejected)
        if rag_result["triangle"]:
            result_gemini["_rag_triangle"] = rag_result["triangle"]
            result_openai["_rag_triangle"] = rag_result["triangle"]

            # Learn inception year from triangle's earliest UW year
            # Never overwrite manually-verified inception years
            tri_uw_years = rag_result["triangle"].get("underwriting_years", [])
            if tri_uw_years and syndicate_num not in manual_overrides:
                tri_inception = min(int(y) for y in tri_uw_years)
                cached = inception_cache.get(syndicate_num)
                if cached is None or tri_inception < cached:
                    inception_cache[syndicate_num] = tri_inception
                    _save_inception_years(inception_cache)

        if not rag_sane:
            # RAG PYD was insane — fall back to LLM-extracted triangles
            result_gemini, result_openai, tri_messages = verify_triangles(
                result_gemini, result_openai, GEMINI_MODEL, OPENAI_MODEL, report_year
            )
            for msg in tri_messages:
                print(msg)
    else:
        # No triangle from RAG — fall back to LLM-extracted triangles
        result_gemini, result_openai, tri_messages = verify_triangles(
            result_gemini, result_openai, GEMINI_MODEL, OPENAI_MODEL, report_year
        )
        for msg in tri_messages:
            print(msg)

    # Apply Adobe LOB breakdown (deterministic override of LLM-extracted LOBs)
    adobe_lob = rag_result.get("adobe_lob")
    if adobe_lob:
        for result, model_name in [
            (result_gemini, GEMINI_MODEL),
            (result_openai, OPENAI_MODEL),
        ]:
            result["gross_premium_mix"] = adobe_lob["gross_premium_mix"]
            result["gross_premiums_written_gbp_m"] = adobe_lob["gross_premiums_written_gbp_m"]
            result["gross_premium_confidence"] = 1.0
            result["_adobe_lob"] = adobe_lob
        print(f"  [Adobe] LOB breakdown applied to both models")

    # Apply Adobe provisions movement as cross-check
    adobe_prov = rag_result.get("adobe_provisions")
    if adobe_prov:
        for result, model_name in [
            (result_gemini, GEMINI_MODEL),
            (result_openai, OPENAI_MODEL),
        ]:
            result["_adobe_provisions"] = adobe_prov

    # If opening reserves came from a reserves movement note (RITC syndicates),
    # inject it into the provisions dict so the auto-resolution mechanism
    # can use it for opening_reserves_gbp_m discrepancy resolution.
    movement_opening = rag_result.get("opening_reserves_from_movement")
    if movement_opening is not None:
        prov_dict = adobe_prov or {}
        if prov_dict.get("opening_gross_claims_outstanding") is None:
            prov_dict["opening_gross_claims_outstanding"] = movement_opening
            for result in [result_gemini, result_openai]:
                result["_adobe_provisions"] = prov_dict
            print(f"  [RAG] Reserves movement opening ({movement_opening:.1f}m) "
                  f"available for opening reserves resolution")

    # Proactive RAG opening reserves override: if the balance sheet / provisions
    # extraction found gross claims outstanding, use it authoritatively for both
    # models — similar to how RAG triangle PYD overrides LLM-extracted PYD.
    rag_opening = (adobe_prov or {}).get("opening_gross_claims_outstanding")
    if rag_opening is not None and rag_opening > 0:
        for result, model_name in [
            (result_gemini, GEMINI_MODEL),
            (result_openai, OPENAI_MODEL),
        ]:
            llm_opening = result.get("opening_reserves_gbp_m")
            result["opening_reserves_gbp_m"] = rag_opening
            # Recompute PYD percentage with corrected opening reserves
            pyd = result.get("prior_year_development_gbp_m")
            if pyd is not None and rag_opening > 0:
                result["prior_year_development_pct"] = round(
                    float(pyd) / rag_opening * 100, 2
                )
            elif pyd == 0 or rag_opening == 0:
                result["prior_year_development_pct"] = 0.0
            if llm_opening is not None:
                try:
                    diff = abs(float(llm_opening) - rag_opening)
                    if diff >= 0.5:
                        old_notes = result.get("data_quality_notes", "") or ""
                        result["data_quality_notes"] = (
                            f"{old_notes} [RAG OVERRIDE: Model said opening_reserves="
                            f"{llm_opening}, balance sheet gives {rag_opening:.3f}m. "
                            f"Using RAG value.]"
                        ).strip()
                        print(f"  [{model_name}] Opening reserves overridden by RAG balance sheet: "
                              f"{llm_opening} -> {rag_opening:.3f}m")
                    else:
                        print(f"  [{model_name}] Opening reserves confirmed by RAG balance sheet: "
                              f"{llm_opening} -> {rag_opening:.3f}m")
                except (TypeError, ValueError):
                    print(f"  [{model_name}] Opening reserves set from RAG balance sheet: "
                          f"{rag_opening:.3f}m")
            else:
                print(f"  [{model_name}] Opening reserves filled from RAG balance sheet: "
                      f"{rag_opening:.3f}m")

    # Net-of-reinsurance PYD fallback: if both LLMs returned null PYD but both
    # found narrative text with a quantified net reserve movement, use the net
    # figure as a last resort.  This is better than leaving PYD as null when
    # the report clearly states a net amount.
    for result, model_name in [
        (result_gemini, GEMINI_MODEL),
        (result_openai, OPENAI_MODEL),
    ]:
        if result.get("prior_year_development_gbp_m") is not None:
            continue  # already has a value
        reserve_text = result.get("exact_reserve_text") or ""
        if not reserve_text:
            continue
        # Try to parse a net figure from the reserve text
        net_pyd = _parse_net_pyd_from_text(reserve_text, result.get("direction"))
        if net_pyd is not None:
            result["prior_year_development_gbp_m"] = net_pyd
            if net_pyd == 0:
                result["prior_year_development_pct"] = 0.0
            elif result.get("opening_reserves_gbp_m") and result["opening_reserves_gbp_m"] > 0:
                result["prior_year_development_pct"] = round(
                    net_pyd / result["opening_reserves_gbp_m"] * 100, 2
                )
            result["direction"] = "release" if net_pyd < 0 else "strengthening" if net_pyd > 0 else "flat"
            old_notes = result.get("data_quality_notes", "") or ""
            result["data_quality_notes"] = (
                f"{old_notes} [NET FALLBACK: No gross PYD available. "
                f"Using net-of-reinsurance figure ({net_pyd:+.3f}m) from narrative text.]"
            )
            print(f"  [{model_name}] PYD filled from narrative net figure: {net_pyd:+.3f}m (net of reinsurance)")

    # Zero-opening override: if opening reserves = 0, PYD must be 0 and
    # direction must be flat — there are no prior year reserves to develop
    opening_g = result_gemini.get("opening_reserves_gbp_m")
    opening_o = result_openai.get("opening_reserves_gbp_m")
    if opening_g == 0 and opening_o == 0:
        for model_name, result in [(GEMINI_MODEL, result_gemini), (OPENAI_MODEL, result_openai)]:
            pyd_val = result.get("prior_year_development_gbp_m")
            if pyd_val is not None and pyd_val != 0:
                print(f"  [{model_name}] Opening reserves = 0 -- PYD forced: "
                      f"{pyd_val:+.3f}m -> 0.0 (no prior reserves to develop)")
            result["prior_year_development_gbp_m"] = 0.0
            result["prior_year_development_pct"] = 0.0
            result["direction"] = "flat"

    # Force direction from resolved PYD — triangle PYD dominates direction
    for model_name, result in [(GEMINI_MODEL, result_gemini), (OPENAI_MODEL, result_openai)]:
        pyd_val = result.get("prior_year_development_gbp_m")
        if pyd_val is not None:
            expected_dir = "flat" if pyd_val == 0 else ("release" if pyd_val < 0 else "strengthening")
            current_dir = result.get("direction")
            if current_dir != expected_dir:
                result["direction"] = expected_dir
                if current_dir is not None:
                    print(f"  [{model_name}] Direction forced: {current_dir} -> {expected_dir} (PYD={pyd_val:+.3f}m)")

    # Compare
    discrepancies = compare_results(result_gemini, result_openai, GEMINI_MODEL, OPENAI_MODEL)
    passed, tolerated, hard_failures = check_tolerance(discrepancies, GEMINI_MODEL, OPENAI_MODEL)

    # Auto-resolve computed fields: prior_year_development_pct can be derived
    # from opening_reserves_gbp_m and prior_year_development_gbp_m when those agree.
    hard_failures, auto_resolved = resolve_computed_fields(
        hard_failures, result_gemini, result_openai, GEMINI_MODEL, OPENAI_MODEL,
        syndicate_num=syndicate_num, report_year=report_year
    )
    tolerated.extend(auto_resolved)
    passed = len(hard_failures) == 0

    # Cost
    meta_g = result_gemini.get("_extraction_meta", {})
    meta_o = result_openai.get("_extraction_meta", {})
    total_cost = meta_g.get("cost_usd", 0) + meta_o.get("cost_usd", 0)
    total_tokens = meta_g.get("total_tokens", 0) + meta_o.get("total_tokens", 0)

    output_data = {
        "extraction_timestamp": datetime.now(timezone.utc).isoformat(),
        "spec": {
            "prompt_version": PROMPT_VERSION,
            "field_definitions_version": FIELD_DEFINITIONS_VERSION,
            "tolerance_rules_version": TOLERANCE_RULES_VERSION,
        },
        "source_file": str(report_path),
        "models": {
            GEMINI_MODEL: result_gemini,
            OPENAI_MODEL: result_openai,
        },
        "total_cost_usd": round(total_cost, 6),
        "total_tokens": total_tokens,
        "validation": {
            "passed": passed,
            "total_discrepancies": len(discrepancies),
            "within_tolerance": len(tolerated),
            "hard_failures": len(hard_failures),
            "hard_failure_details": [
                {
                    "field": d["field"],
                    "type": d["type"],
                    GEMINI_MODEL: d.get(GEMINI_MODEL, d.get("only_in", "")),
                    OPENAI_MODEL: d.get(OPENAI_MODEL, d.get("value", "")),
                }
                if d["type"] != "list_extra"
                else {
                    "field": d["field"],
                    "type": d["type"],
                    "only_in": d["only_in"],
                    "value": d["value"],
                }
                for d in hard_failures
            ],
        },
    }

    return output_data, passed, discrepancies, hard_failures


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Guard against non-existent flags that would silently run against ALL reports
    for bad_flag in ("--syndicates", "--years"):
        if bad_flag in sys.argv:
            print(f"ERROR: {bad_flag} is not a valid flag for test_gemini.py.")
            print(f"  Use --single syndicate_NNNN_YYYY to process a single report.")
            print(f"  Example: python test_gemini.py --single syndicate_2357_2016 --batch")
            sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    # Print spec versions at startup
    print(f"Prompt version: {PROMPT_VERSION}")
    print(f"Field definitions: {FIELD_DEFINITIONS_VERSION}")
    print(f"Tolerance rules: {TOLERANCE_RULES_VERSION}")
    print(f"Models: {GEMINI_MODEL} vs {OPENAI_MODEL}")
    print()

    reports = discover_reports()
    print(f"Found {len(reports)} reports in {REPORTS_DIR}")

    # --table-backend flag: select table extraction backend (nutrient/adobe/azure)
    for idx, arg in enumerate(sys.argv):
        if arg == "--table-backend" and idx + 1 < len(sys.argv):
            backend_name = sys.argv[idx + 1].lower()
            try:
                TABLE_BACKEND = TableBackend(backend_name)
            except ValueError:
                print(f"Unknown table backend: {backend_name}. "
                      f"Use: nutrient, adobe, or azure")
                sys.exit(1)
            break
    print(f"Table extraction backend: {TABLE_BACKEND.value}")

    # --azure-free flag: use free F0 tier (one page per request instead of batched)
    if "--azure-free" in sys.argv:
        AZURE_PAID = False
        print("Azure mode: free (F0 tier) — one page per request")

    # --single flag: process only the named report (e.g. --single syndicate_1856_2024)
    single_arg = None
    for idx, arg in enumerate(sys.argv):
        if arg == "--single" and idx + 1 < len(sys.argv):
            single_arg = sys.argv[idx + 1]
            break

    # --clean flag: delete all existing outputs to force re-run under current spec
    if "--clean" in sys.argv:
        existing = list(OUTPUT_DIR.glob("syndicate_*.json"))
        if existing:
            print(f"  --clean: deleting {len(existing)} existing output files...")
            for f in existing:
                f.unlink()
            print(f"  Deleted. All reports will be re-processed.")

    if single_arg:
        # Filter to just the named report, and remove existing output to force re-run
        to_process = [r for r in reports if r.stem == single_arg]
        if not to_process:
            print(f"Report '{single_arg}' not found in {REPORTS_DIR}")
            sys.exit(1)
        # Delete existing output so it will be re-processed
        existing_output = OUTPUT_DIR / f"{single_arg}.json"
        if existing_output.exists():
            existing_output.unlink()
            print(f"  Deleted existing output: {existing_output}")
        already_done = set()
        print(f"Single report mode: {single_arg}")
    else:
        # Skip already-processed reports (verify JSON is valid — truncated files from interrupted runs are re-processed)
        already_done = set()
        for f in OUTPUT_DIR.glob("syndicate_*.json"):
            if f.stem == "syndicate_inception_years":
                continue
            try:
                with open(f) as fh:
                    json.load(fh)
                already_done.add(f.stem)
            except (json.JSONDecodeError, OSError):
                print(f"  Removing corrupt output: {f.name}")
                f.unlink()
        to_process = [r for r in reports if r.stem not in already_done]
        print(f"Already processed: {len(already_done)}, remaining: {len(to_process)}")

    if not to_process:
        print("Nothing to do.")
        sys.exit(0)

    # Running totals
    run_total_cost = 0.0
    run_total_tokens = 0
    run_processed = 0
    run_passed = 0
    run_failed = 0
    run_errored = 0
    run_skipped_first_year = 0
    run_skipped_no_data = 0

    # Enable Ctrl+C to interrupt blocking API calls on Windows
    signal.signal(signal.SIGINT, signal.default_int_handler)

    # Pre-load logs so Ctrl+C handler can always save them
    dis_log = load_disagreement_log()
    rej_log = load_rejection_log()

    # Load syndicate inception years cache (shared across all reports)
    inception_cache = _load_inception_years()
    print(f"Loaded inception years for {len(inception_cache)} syndicates")

    for i, report_path in enumerate(to_process, 1):
        syndicate_num, report_year = parse_report_filename(report_path)
        print(f"\n{'=' * 70}")
        print(f"[{i}/{len(to_process)}] Syndicate {syndicate_num} / {report_year}  ({report_path.name})")
        print(f"{'=' * 70}")

        try:
            result = process_one_report(report_path, inception_cache)
            if isinstance(result, tuple) and len(result) == 2 and result[0] == "first_year":
                # First-year syndicate — write audit-trail JSON and skip further processing
                first_year_data = result[1]
                output_file = OUTPUT_DIR / f"{report_path.stem}.json"
                with open(output_file, "w") as f:
                    json.dump(sanitize_json_ascii(first_year_data), f, indent=2, ensure_ascii=True)
                print(f"  SKIP: First/second-year syndicate - no prior year development possible")
                print(f"  Audit JSON written: {output_file.name}")
                print(f"  >> RESULT: Report not used (insufficient underwriting history)")
                run_skipped_first_year += 1
                continue
            if isinstance(result, tuple) and len(result) == 2 and result[0] == "no_triangle_data":
                # No triangle or reserve text — write audit JSON and skip LLM calls
                no_data = result[1]
                output_file = OUTPUT_DIR / f"{report_path.stem}.json"
                with open(output_file, "w") as f:
                    json.dump(sanitize_json_ascii(no_data), f, indent=2, ensure_ascii=True)
                print(f"  SKIP: No claims development triangle or reserve text in report")
                print(f"  Audit JSON written: {output_file.name}")
                print(f"  >> RESULT: Recommend non-inclusion in analysis (no usable reserve data)")
                run_skipped_no_data += 1
                continue
            output_data, passed, discrepancies, hard_failures = result
        except KeyboardInterrupt:
            print(f"\n\n  Ctrl+C — stopping after {run_processed} reports.")
            print(f"  Saving logs and manifest...")
            save_disagreement_log(dis_log)
            save_rejection_log(rej_log)
            write_run_manifest({
                "total_found": len(reports),
                "already_done": len(already_done),
                "processed": run_processed,
                "passed": run_passed,
                "failed": run_failed,
                "errored": run_errored,
                "skipped_first_year": run_skipped_first_year,
                "skipped_no_data": run_skipped_no_data,
                "total_tokens": run_total_tokens,
                "total_cost": run_total_cost,
                "stopped_early": True,
                "stop_reason": "keyboard_interrupt",
            })
            print(f"  Done. Re-run to continue from where you left off.")
            sys.exit(130)
        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()
            print(f"  Skipping {report_path.name} and continuing...")
            run_errored += 1
            continue

        run_processed += 1
        run_total_cost += output_data["total_cost_usd"]
        run_total_tokens += output_data["total_tokens"]

        # Write output JSON
        output_file = OUTPUT_DIR / f"{report_path.stem}.json"
        with open(output_file, "w") as f:
            json.dump(sanitize_json_ascii(output_data), f, indent=2, ensure_ascii=True)

        status = "PASSED" if passed else "FAILED"
        print(f"  Validation: {status}  ({len(discrepancies)} discrepancies, {len(hard_failures)} hard failures)")
        print(f"  Written: {output_file}")

        # Console summary: PYD amount + % of reserves
        # Resolve across all models (first non-null), matching dashboard logic
        _any_model = output_data.get("models", {})
        _pyd_gbp = None
        _pyd_pct = None
        _opening = None
        _currency = "GBP"
        for _m in _any_model.values():
            if _pyd_gbp is None and _m.get("prior_year_development_gbp_m") is not None:
                _pyd_gbp = _m["prior_year_development_gbp_m"]
            if _pyd_pct is None and _m.get("prior_year_development_pct") is not None:
                _pyd_pct = _m["prior_year_development_pct"]
            if _opening is None and _m.get("opening_reserves_gbp_m") is not None:
                _opening = _m["opening_reserves_gbp_m"]
            if _m.get("currency"):
                _currency = _m["currency"]
        if _pyd_gbp is not None and _pyd_pct is not None:
            _dir = "release" if _pyd_gbp < 0 else "strengthening" if _pyd_gbp > 0 else "flat"
            if _opening is not None:
                print(f"  >> RESULT: PYD = {_pyd_gbp:+.1f}m {_currency} ({_pyd_pct:+.1f}% of "
                      f"{_opening:.0f}m reserves) [{_dir}]")
            else:
                print(f"  >> RESULT: PYD = {_pyd_gbp:+.1f}m {_currency} ({_pyd_pct:+.1f}%) [{_dir}]")
        elif _pyd_gbp is not None:
            _dir = "release" if _pyd_gbp < 0 else "strengthening" if _pyd_gbp > 0 else "flat"
            print(f"  >> RESULT: PYD = {_pyd_gbp:+.1f}m [{_dir}]")
        else:
            print(f"  >> RESULT: No prior year development extracted")

        # Check completeness: warn about missing fields that affect dashboard reliability
        _missing = []
        _has_premium_mix = any(
            m.get("gross_premium_mix") for m in _any_model.values()
        )
        if not _has_premium_mix:
            _missing.append("premium mix (no LOB breakdown found)")
        if _pyd_pct is None:
            _missing.append("PYD %")
        if _missing:
            print(f"  >> INCOMPLETE: missing {'; '.join(_missing)}")

        if not passed:
            run_failed += 1
            print(f"\n  HARD FAILURES:")
            print_discrepancies(hard_failures, GEMINI_MODEL, OPENAI_MODEL)

            batch_mode = "--batch" in sys.argv

            if batch_mode:
                # Old behaviour: log and continue (no interactive adjudication)
                append_to_disagreement_log(
                    report_path.stem,
                    output_data["validation"]["hard_failure_details"],
                )
                if "--stop-on-failure" in sys.argv:
                    print(f"\n  STOPPING -- hard failure on {report_path.name}")
                    break
                else:
                    print(f"  (batch mode -- continuing)")
            else:
                # --- Inline adjudication (human-in-the-loop) ---
                dis_log = load_disagreement_log()
                rej_log = load_rejection_log()

                # Find source PDF for adjudicator
                actual_pdf = report_path
                if report_path.suffix.lower() in (".html", ".htm"):
                    cached = HTML_PDF_CACHE / f"{report_path.stem}.pdf"
                    if cached.exists():
                        actual_pdf = cached

                report_results = []
                report_excluded = False
                stopped = False

                for d in output_data["validation"]["hard_failure_details"]:
                    field = d["field"]
                    gemini_val = d.get(GEMINI_MODEL, d.get("only_in", ""))
                    gpt_val = d.get(OPENAI_MODEL, d.get("value", ""))

                    # Skip if already adjudicated in a prior run
                    already = any(
                        e["report"] == report_path.stem and e["field"] == field
                        and e["status"] not in ("pending_review",)
                        for e in dis_log["entries"]
                    )
                    if already:
                        print(f"\n  [{field}] Already adjudicated, skipping")
                        continue

                    # Auto-accept immaterial differences
                    # Core reserve fields: monetary < 0.2m, pct < 2pp
                    # Gross premium mix fields: amount_gbp_m < 2m, percentage_of_total < 2pp
                    core_monetary = (
                        "opening_reserves_gbp_m",
                        "prior_year_development_gbp_m",
                        "gross_premiums_written_gbp_m",
                    )
                    is_core_monetary = field in core_monetary
                    is_mix_monetary = field.startswith("gross_premium_mix") and ".amount_gbp_m" in field
                    is_monetary = is_core_monetary or is_mix_monetary
                    is_pct = (
                        field == "prior_year_development_pct"
                        or (field.startswith("gross_premium_mix") and ".percentage_of_total" in field)
                    )
                    if is_monetary or is_pct:
                        if is_core_monetary:
                            threshold = 0.2
                        elif is_mix_monetary:
                            threshold = 2.0
                        else:
                            threshold = 2.0  # percentage fields
                        unit = "m" if is_monetary else "pp"
                        try:
                            g_num = float(gemini_val) if gemini_val not in (None, "None", "") else None
                            o_num = float(gpt_val) if gpt_val not in (None, "None", "") else None
                            if g_num is not None and o_num is not None and abs(g_num - o_num) < threshold:
                                # Priority: adjudicator (if available) > Gemini > GPT
                                picked_val = g_num
                                picked_model = GEMINI_MODEL
                                print(f"\n  [{field}]")
                                print(f"    Gemini: {gemini_val}")
                                print(f"    GPT:    {gpt_val}")
                                print(f"    Auto-accepted: immaterial difference "
                                      f"({abs(g_num - o_num):.3f}{unit} < {threshold}{unit}), using {picked_model}: {picked_val}")
                                report_results.append({
                                    "field": field,
                                    "decision_type": "auto_accept",
                                    "final_model": picked_model,
                                    "final_value": picked_val,
                                    "adjudicator_value": picked_val,
                                    "adjudicator_confidence": 1.0,
                                    "evidence": f"Immaterial difference: {g_num} vs {o_num}, diff={abs(g_num - o_num):.3f}{unit}",
                                })
                                now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                                next_adj_num = max(
                                    (int(e["id"].split("-")[1])
                                     for e in dis_log["entries"]
                                     if e["id"].startswith("ADJ-")),
                                    default=0,
                                ) + 1
                                dis_log["entries"].append({
                                    "id": f"ADJ-{next_adj_num:04d}",
                                    "report": report_path.stem,
                                    "field": field,
                                    "gemini_value": sanitize_for_json(str(gemini_val)),
                                    "gpt_value": sanitize_for_json(str(gpt_val)),
                                    "adjudicator_value": picked_val,
                                    "adjudicator_confidence": 1.0,
                                    "adjudicator_evidence": f"Immaterial difference: {g_num} vs {o_num}",
                                    "final_model": picked_model,
                                    "final_value": picked_val,
                                    "human_decision": f"Auto-accepted: immaterial difference ({abs(g_num - o_num):.3f}{unit} < {threshold}{unit})",
                                    "status": "resolved_auto_accept",
                                    "prompt_version": PROMPT_VERSION,
                                    "date": now,
                                })
                                save_disagreement_log(dis_log)
                                continue
                        except (ValueError, TypeError):
                            pass

                    print(f"\n  [{field}]")
                    print(f"    Gemini: {gemini_val}")
                    print(f"    GPT:    {gpt_val}")

                    # Build supporting context from both models
                    gem_data = output_data["models"][GEMINI_MODEL]
                    gpt_data = output_data["models"][OPENAI_MODEL]
                    context_lines = _get_field_context(field, gem_data, gpt_data)

                    # Auto-compute prior_year_development_pct from amount / opening reserves
                    if field == "prior_year_development_pct":
                        # Get best available values (from resolved results or agreed model values)
                        pyd_val = None
                        reserves_val = None

                        # Check if prior_year_development_gbp_m was already resolved
                        for r in report_results:
                            if r["field"] == "prior_year_development_gbp_m":
                                pyd_val = r.get("final_value")
                                break
                        if pyd_val is None:
                            # Models agreed — use either
                            pyd_val = gem_data.get("prior_year_development_gbp_m")

                        # Check if opening_reserves_gbp_m was already resolved
                        for r in report_results:
                            if r["field"] == "opening_reserves_gbp_m":
                                reserves_val = r.get("final_value")
                                break
                        if reserves_val is None:
                            reserves_val = gem_data.get("opening_reserves_gbp_m")

                        # Convert to float
                        try:
                            pyd_float = float(pyd_val) if pyd_val is not None else None
                            res_float = float(reserves_val) if reserves_val is not None else None
                        except (ValueError, TypeError):
                            pyd_float = None
                            res_float = None

                        if pyd_float is not None and res_float and res_float != 0:
                            computed_pct = round(pyd_float / res_float * 100, 2)
                            print(f"    Auto-computed: {pyd_float} / {res_float} * 100 = {computed_pct}%")
                            report_results.append({
                                "field": field,
                                "decision_type": "auto_accept",
                                "final_model": "computed",
                                "final_value": computed_pct,
                                "adjudicator_value": computed_pct,
                                "adjudicator_confidence": 1.0,
                                "evidence": f"Computed: {pyd_float} / {res_float} * 100 = {computed_pct}",
                            })
                            now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                            next_adj_num = max(
                                (int(e["id"].split("-")[1])
                                 for e in dis_log["entries"]
                                 if e["id"].startswith("ADJ-")),
                                default=0,
                            ) + 1
                            dis_log["entries"].append({
                                "id": f"ADJ-{next_adj_num:04d}",
                                "report": report_path.stem,
                                "field": field,
                                "gemini_value": sanitize_for_json(str(gemini_val)),
                                "gpt_value": sanitize_for_json(str(gpt_val)),
                                "adjudicator_value": computed_pct,
                                "adjudicator_confidence": 1.0,
                                "adjudicator_evidence": f"Computed: {pyd_float} / {res_float} * 100 = {computed_pct}",
                                "final_model": "computed",
                                "final_value": computed_pct,
                                "human_decision": "Auto-computed from prior_year_development_gbp_m / opening_reserves_gbp_m",
                                "status": "resolved_auto_accept",
                                "prompt_version": PROMPT_VERSION,
                                "date": now,
                            })
                            save_disagreement_log(dis_log)
                            continue

                    prompt = build_verification_prompt(
                        field, gemini_val, gpt_val, syndicate_num, report_year
                    )
                    print(f"    Sending to {ADJUDICATOR_MODEL}...")

                    # Collect page hints from both models for trimming large PDFs
                    page_hints = []
                    for page_field in (
                        "opening_reserves_page", "prior_year_movement_page",
                        "gross_premium_page",
                    ):
                        for model_data in (gem_data, gpt_data):
                            pg = model_data.get(page_field)
                            if isinstance(pg, (int, float)) and pg > 0:
                                page_hints.append(int(pg))

                    try:
                        result, tokens_in, tokens_out = call_adjudicator(
                            actual_pdf, prompt, page_hints=page_hints or None,
                            syndicate_num=syndicate_num, report_year=report_year,
                            field=field,
                        )
                    except Exception as e:
                        print(f"    ERROR calling adjudicator: {e}")
                        print(f"    You can still make a manual decision.")
                        print(f"      [g] Use Gemini value")
                        print(f"      [o] Use GPT value")
                        print(f"      [v] Enter a custom value")
                        print(f"      [x] Exclude this report entirely")
                        print(f"      [s] Stop script")
                        choice = ask_human("    Your decision: ", ["g", "o", "v", "x", "s"])
                        if choice == "s":
                            save_disagreement_log(dis_log)
                            save_rejection_log(rej_log)
                            print(f"\n  Logs saved. Stopping.")
                            write_run_manifest({
                                "total_found": len(reports),
                                "already_done": len(already_done),
                                "processed": run_processed,
                                "passed": run_passed,
                                "failed": run_failed,
                                "errored": run_errored,
                                "skipped_first_year": run_skipped_first_year,
                "skipped_no_data": run_skipped_no_data,
                                "total_tokens": run_total_tokens,
                                "total_cost": run_total_cost,
                                "stopped_early": True,
                            })
                            sys.exit(0)
                        if choice == "x":
                            reason = ask_human("    Reason for exclusion: ")
                            now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                            rej_log["entries"] = [
                                e for e in rej_log["entries"]
                                if e["report"] != report_path.stem
                            ]
                            rej_log["entries"].append({
                                "report": report_path.stem,
                                "syndicate": syndicate_num,
                                "year": report_year,
                                "rejected": True,
                                "reason": reason,
                                "date": now,
                            })
                            save_rejection_log(rej_log)
                            save_disagreement_log(dis_log)
                            output_data["excluded"] = True
                            output_data["exclusion_reason"] = reason
                            output_data["exclusion_date"] = now
                            output_file = OUTPUT_DIR / f"{report_path.stem}.json"
                            with open(output_file, "w") as f:
                                json.dump(sanitize_json_ascii(output_data), f, indent=2, ensure_ascii=True)
                            print(f"    Report excluded: {reason}")
                            report_excluded = True
                            break
                        if choice == "v":
                            custom = ask_human("    Enter custom value: ")
                            final_model = "human"
                            final_value = custom
                        else:
                            final_model = GEMINI_MODEL if choice == "g" else OPENAI_MODEL
                            final_value = gemini_val if choice == "g" else gpt_val
                        report_results.append({
                            "field": field,
                            "decision_type": "override",
                            "final_model": final_model,
                            "final_value": final_value,
                        })
                        continue

                    adj_cost = tokens_in * 3.0 / 1_000_000 + tokens_out * 15.0 / 1_000_000
                    run_total_cost += adj_cost
                    run_total_tokens += tokens_in + tokens_out

                    adj_value = result.get("correct_value")
                    evidence = sanitize_for_json(str(result.get("evidence", "")))
                    confidence = result.get("confidence", 0.0)

                    correct_model, reason = determine_correct_model(
                        field, adj_value, gemini_val, gpt_val
                    )

                    # Auto-accept rule 1: one model null, adjudicator returns a value.
                    # Trust the adjudicator whether it agrees with the non-null model
                    # or computes its own value from the PDF.
                    gemini_is_null = gemini_val is None or gemini_val == "None" or gemini_val == ""
                    gpt_is_null = gpt_val is None or gpt_val == "None" or gpt_val == ""
                    one_null_one_value = (gemini_is_null != gpt_is_null)
                    adj_resolves_null = (
                        one_null_one_value
                        and adj_value is not None
                    )
                    if adj_resolves_null:
                        final = correct_model if correct_model not in ("neither", None) else "adjudicator"
                        if correct_model not in ("neither", None):
                            reason_text = f"Auto-accepted: one model null, adjudicator confirms {final}"
                        else:
                            reason_text = f"Auto-accepted: one model null, adjudicator computed own value"
                        print(f"    {reason_text} (confidence {confidence})")
                        print(f"    Value: {adj_value}")
                        print(f"    Evidence: {evidence}")
                        report_results.append({
                            "field": field,
                            "decision_type": "auto_accept",
                            "final_model": final,
                            "final_value": adj_value,
                            "adjudicator_value": adj_value,
                            "adjudicator_confidence": confidence,
                            "evidence": evidence,
                        })
                        now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                        next_adj_num = max(
                            (int(e["id"].split("-")[1])
                             for e in dis_log["entries"]
                             if e["id"].startswith("ADJ-")),
                            default=0,
                        ) + 1
                        dis_log["entries"].append({
                            "id": f"ADJ-{next_adj_num:04d}",
                            "report": report_path.stem,
                            "field": field,
                            "gemini_value": sanitize_for_json(str(gemini_val)),
                            "gpt_value": sanitize_for_json(str(gpt_val)),
                            "adjudicator_value": sanitize_for_json(str(adj_value)),
                            "adjudicator_confidence": confidence,
                            "adjudicator_evidence": evidence,
                            "final_model": final,
                            "final_value": sanitize_for_json(str(adj_value)),
                            "human_decision": reason_text,
                            "status": "resolved_auto_accept",
                            "prompt_version": PROMPT_VERSION,
                            "date": now,
                        })
                        save_disagreement_log(dis_log)
                        continue

                    # Auto-accept rule 2 (general): adjudicator agrees with either model
                    adj_picks_side = (
                        correct_model not in ("neither", None)
                        and adj_value is not None
                    )
                    if adj_picks_side:
                        print(f"    Auto-accepted: adjudicator agrees with "
                              f"{correct_model} (confidence {confidence})")
                        print(f"    Value: {adj_value}")
                        print(f"    Evidence: {evidence}")
                        report_results.append({
                            "field": field,
                            "decision_type": "auto_accept",
                            "final_model": correct_model,
                            "final_value": adj_value,
                            "adjudicator_value": adj_value,
                            "adjudicator_confidence": confidence,
                            "evidence": evidence,
                        })
                        now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                        next_adj_num = max(
                            (int(e["id"].split("-")[1])
                             for e in dis_log["entries"]
                             if e["id"].startswith("ADJ-")),
                            default=0,
                        ) + 1
                        dis_log["entries"].append({
                            "id": f"ADJ-{next_adj_num:04d}",
                            "report": report_path.stem,
                            "field": field,
                            "gemini_value": sanitize_for_json(str(gemini_val)),
                            "gpt_value": sanitize_for_json(str(gpt_val)),
                            "adjudicator_value": sanitize_for_json(str(adj_value)),
                            "adjudicator_confidence": confidence,
                            "adjudicator_evidence": evidence,
                            "final_model": correct_model,
                            "final_value": sanitize_for_json(str(adj_value)),
                            "human_decision": f"Auto-accepted: adjudicator agrees with {correct_model}",
                            "status": "resolved_auto_accept",
                            "prompt_version": PROMPT_VERSION,
                            "date": now,
                        })
                        save_disagreement_log(dis_log)
                        continue

                    decision_type, decision_data = present_adjudication(
                        report_path.stem, field, gemini_val, gpt_val,
                        adj_value, confidence, evidence, correct_model, reason,
                        context_lines=context_lines,
                    )

                    if decision_type == "stop":
                        save_disagreement_log(dis_log)
                        save_rejection_log(rej_log)
                        print(f"\n  Logs saved. Stopping.")
                        write_run_manifest({
                            "total_found": len(reports),
                            "already_done": len(already_done),
                            "processed": run_processed,
                            "passed": run_passed,
                            "failed": run_failed,
                            "errored": run_errored,
                            "skipped_first_year": run_skipped_first_year,
                "skipped_no_data": run_skipped_no_data,
                            "total_tokens": run_total_tokens,
                            "total_cost": run_total_cost,
                            "stopped_early": True,
                        })
                        sys.exit(0)

                    if decision_type == "exclude":
                        # Record exclusion immediately and skip remaining fields
                        now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                        rej_log["entries"] = [
                            e for e in rej_log["entries"]
                            if e["report"] != report_path.stem
                        ]
                        rej_log["entries"].append({
                            "report": report_path.stem,
                            "syndicate": syndicate_num,
                            "year": report_year,
                            "rejected": True,
                            "reason": decision_data,
                            "date": now,
                        })
                        save_rejection_log(rej_log)
                        save_disagreement_log(dis_log)
                        # Update the output JSON with exclusion flag
                        output_data["excluded"] = True
                        output_data["exclusion_reason"] = decision_data
                        output_data["exclusion_date"] = now
                        output_file = OUTPUT_DIR / f"{report_path.stem}.json"
                        with open(output_file, "w") as f:
                            json.dump(sanitize_json_ascii(output_data), f, indent=2, ensure_ascii=True)
                        print(f"    Report excluded: {decision_data}")
                        report_excluded = True
                        break

                    # Determine final model/value
                    if decision_type == "approve":
                        final_model = decision_data
                        final_value = adj_value
                        human_reason = "Approved adjudicator recommendation"
                        status = "resolved_model_error"
                    elif decision_type == "override":
                        final_model = decision_data
                        final_value = gemini_val if decision_data == GEMINI_MODEL else gpt_val
                        human_reason = f"Human override: chose {final_model}"
                        status = "resolved_human_override"
                    elif decision_type == "override_value":
                        final_model = "human"
                        final_value = decision_data
                        human_reason = f"Human provided custom value: {decision_data}"
                        status = "resolved_human_override"

                    report_results.append({
                        "field": field,
                        "decision_type": decision_type,
                        "final_model": final_model,
                        "final_value": final_value,
                        "adjudicator_value": adj_value,
                        "adjudicator_confidence": confidence,
                        "evidence": evidence,
                    })

                    # Update disagreement log immediately
                    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    existing_entry = None
                    for e in dis_log["entries"]:
                        if e["report"] == report_path.stem and e["field"] == field:
                            existing_entry = e
                            break

                    if existing_entry:
                        existing_entry["correct_model"] = final_model if final_model != "neither" else None
                        existing_entry["manual_verification"] = (
                            f"Adjudicated by {ADJUDICATOR_MODEL}: {evidence}"
                        )
                        existing_entry["adjudicator_value"] = adj_value
                        existing_entry["adjudicator_confidence"] = confidence
                        existing_entry["human_decision"] = human_reason
                        existing_entry["final_value"] = sanitize_for_json(str(final_value))
                        existing_entry["action"] = human_reason
                        existing_entry["status"] = status
                        existing_entry["date"] = now
                    else:
                        next_adj_num = max(
                            (int(e["id"].split("-")[1])
                             for e in dis_log["entries"]
                             if e["id"].startswith("ADJ-")),
                            default=0,
                        ) + 1
                        new_entry = {
                            "id": f"ADJ-{next_adj_num:04d}",
                            "report": report_path.stem,
                            "field": field,
                            "gemini_value": sanitize_for_json(str(gemini_val)),
                            "gpt_value": sanitize_for_json(str(gpt_val)),
                            "correct_model": final_model if final_model != "neither" else None,
                            "adjudicator_value": adj_value,
                            "adjudicator_confidence": confidence,
                            "final_value": sanitize_for_json(str(final_value)),
                            "manual_verification": f"Adjudicated by {ADJUDICATOR_MODEL}: {evidence}",
                            "human_decision": human_reason,
                            "action": human_reason,
                            "prompt_version_before": PROMPT_VERSION,
                            "prompt_version_after": None,
                            "status": status,
                            "date": now,
                        }
                        dis_log["entries"].append(new_entry)

                    save_disagreement_log(dis_log)
                    print(f"    Recorded: {human_reason}")

                # Report-level decision (if any fields were adjudicated and not already excluded)
                if report_results and not report_excluded:
                    # Auto-include if every disagreement was auto-accepted
                    all_auto = all(
                        r.get("decision_type") == "auto_accept"
                        for r in report_results
                    )
                    if all_auto:
                        report_decision, report_data = "include", None
                        print(f"    Auto-included: all {len(report_results)} disagreement(s) were auto-accepted")
                    else:
                        report_decision, report_data = present_report_decision(
                            report_path.stem, syndicate_num, report_year, report_results
                        )

                    if report_decision == "stop":
                        save_disagreement_log(dis_log)
                        save_rejection_log(rej_log)
                        print(f"\n  Logs saved. Stopping.")
                        write_run_manifest({
                            "total_found": len(reports),
                            "already_done": len(already_done),
                            "processed": run_processed,
                            "passed": run_passed,
                            "failed": run_failed,
                            "errored": run_errored,
                            "skipped_first_year": run_skipped_first_year,
                "skipped_no_data": run_skipped_no_data,
                            "total_tokens": run_total_tokens,
                            "total_cost": run_total_cost,
                            "stopped_early": True,
                        })
                        sys.exit(0)

                    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")

                    # Remove any prior entry for this report in rejection log
                    rej_log["entries"] = [
                        e for e in rej_log["entries"]
                        if e["report"] != report_path.stem
                    ]

                    if report_decision == "include":
                        overrides = {}
                        for r in report_results:
                            overrides[r["field"]] = {
                                "model": r.get("final_model"),
                                "value": sanitize_for_json(str(r.get("final_value", ""))),
                            }
                        include_reason = (
                            "Auto-included: all disagreements auto-accepted"
                            if all_auto else
                            "Included with human-approved overrides for disputed fields"
                        )
                        rej_log["entries"].append({
                            "report": report_path.stem,
                            "syndicate": syndicate_num,
                            "year": report_year,
                            "rejected": False,
                            "reason": include_reason,
                            "overrides": overrides,
                            "status": "included_auto_accept" if all_auto else "included_with_override",
                            "human_approved": not all_auto,
                            "date": now,
                        })
                    elif report_decision == "exclude":
                        rej_log["entries"].append({
                            "report": report_path.stem,
                            "syndicate": syndicate_num,
                            "year": report_year,
                            "rejected": True,
                            "reason": report_data,
                            "status": "excluded",
                            "human_approved": True,
                            "date": now,
                        })
                        # Update the output JSON with exclusion flag
                        output_data["excluded"] = True
                        output_data["exclusion_reason"] = report_data
                        output_data["exclusion_date"] = now
                        output_file = OUTPUT_DIR / f"{report_path.stem}.json"
                        with open(output_file, "w") as f:
                            json.dump(sanitize_json_ascii(output_data), f, indent=2, ensure_ascii=True)

                    save_rejection_log(rej_log)
                    print(f"  Rejection log updated for {report_path.stem}")
        else:
            run_passed += 1

    # Summary
    print(f"\n{'=' * 70}")
    print("RUN SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Processed:        {run_processed}")
    print(f"  Passed:           {run_passed}")
    print(f"  Failed:           {run_failed}")
    print(f"  Errored:          {run_errored}")
    print(f"  Skipped (1st yr): {run_skipped_first_year}")
    print(f"  Skipped (no tri): {run_skipped_no_data}")
    print(f"  Remaining:        {len(to_process) - run_processed - run_errored - run_skipped_first_year - run_skipped_no_data}")
    print(f"  Tokens:           {run_total_tokens:,}")
    print(f"  Cost:             ${run_total_cost:.4f}")

    # Write run manifest
    write_run_manifest({
        "total_found": len(reports),
        "already_done": len(already_done),
        "processed": run_processed,
        "passed": run_passed,
        "failed": run_failed,
        "errored": run_errored,
        "skipped_first_year": run_skipped_first_year,
                "skipped_no_data": run_skipped_no_data,
        "total_tokens": run_total_tokens,
        "total_cost": run_total_cost,
    })
